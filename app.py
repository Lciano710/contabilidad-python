import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from datetime import datetime
import re
import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime, timedelta
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# Añade estas importaciones al inicio de tu archivo, junto con las otras importaciones
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# --- Funciones para formato de moneda ---
def validar_monto(texto):
    """Valida que el texto ingresado sea un monto válido"""
    if texto == "":
        return True
    return bool(re.match(r'^[\d,]*\.?\d{0,2}$', texto))
def formatear_monto_durante_escritura(event=None):
    """Formatea el monto mientras se escribe"""
    texto = monto_entry.get().replace(",", "")
    if texto:
        try:
            if "." in texto:
                partes = texto.split(".")
                parte_entera = partes[0]
                parte_decimal = partes[1][:2] if len(partes) > 1 else ""
            else:
                parte_entera = texto
                parte_decimal = ""
            
            if parte_entera:
                parte_entera = "{:,}".format(int(parte_entera))
            
            texto_formateado = parte_entera
            if parte_decimal:
                texto_formateado += f".{parte_decimal}"
            
            if texto_formateado != monto_entry.get():
                monto_entry.delete(0, tk.END)
                monto_entry.insert(0, texto_formateado)
        except:
            pass

def obtener_cuenta_principal(cuenta_completa):
    if not cuenta_completa:
        return ""
    codigo = cuenta_completa.split()[0] if cuenta_completa else ""
    return codigo[:2] if codigo and codigo[:2].isdigit() else ""

def agregar_cuenta_debe():
    cuenta = cuenta_debe_var.get()
    monto = monto_entry.get()
    
    if not cuenta or cuenta == "Seleccionar cuenta...":
        tk.messagebox.showerror("Error", "Debe seleccionar una cuenta para el débito")
        return
    
    if not monto:
        tk.messagebox.showerror("Error", "Debe ingresar un monto")
        return
    
    try:
        monto_float = float(monto.replace(",", ""))
        if monto_float <= 0:
            tk.messagebox.showerror("Error", "El monto debe ser mayor que cero")
            return
    except ValueError:
        tk.messagebox.showerror("Error", "Monto inválido")
        return
    
    # Agregar a la lista de débitos (sin limpiar el monto)
    debe_items.append({
        'cuenta': cuenta,
        'monto': monto_float
    })
    
    # Actualizar el treeview de débitos
    actualizar_treeview_debe()
    
    # Limpiar solo la cuenta pero mantener monto y demás campos
    cuenta_debe_var.set("Seleccionar cuenta...")
    monto_entry.focus()

def agregar_cuenta_haber():
    cuenta = cuenta_haber_var.get()
    monto = monto_entry.get()
    
    if not cuenta or cuenta == "Seleccionar cuenta...":
        tk.messagebox.showerror("Error", "Debe seleccionar una cuenta para el crédito")
        return
    
    if not monto:
        tk.messagebox.showerror("Error", "Debe ingresar un monto")
        return
    
    try:
        monto_float = float(monto.replace(",", ""))
        if monto_float <= 0:
            tk.messagebox.showerror("Error", "El monto debe ser mayor que cero")
            return
    except ValueError:
        tk.messagebox.showerror("Error", "Monto inválido")
        return
    
    # Agregar a la lista de créditos (sin limpiar el monto)
    haber_items.append({
        'cuenta': cuenta,
        'monto': monto_float
    })
    
    # Actualizar el treeview de créditos
    actualizar_treeview_haber()
    
    # Limpiar solo la cuenta pero mantener monto y demás campos
    cuenta_haber_var.set("Seleccionar cuenta...")
    monto_entry.focus()

def actualizar_treeview_debe():
    for item in tree_debe.get_children():
        tree_debe.delete(item)
    
    total_debe = 0
    for i, item in enumerate(debe_items, start=1):
        tree_debe.insert("", "end", values=(i, item['cuenta'], f"S/. {item['monto']:,.2f}"))
        total_debe += item['monto']
    
    lbl_total_debe.config(text=f"Total Debe: S/. {total_debe:,.2f}")
    ajustar_altura_treeview(tree_debe, debe_items)  # Añadir esta línea

def actualizar_treeview_haber():
    for item in tree_haber.get_children():
        tree_haber.delete(item)
    
    total_haber = 0
    for i, item in enumerate(haber_items, start=1):
        tree_haber.insert("", "end", values=(i, item['cuenta'], f"S/. {item['monto']:,.2f}"))
        total_haber += item['monto']
    
    lbl_total_haber.config(text=f"Total Haber: S/. {total_haber:,.2f}")
    ajustar_altura_treeview(tree_haber, haber_items)  # Añadir esta línea

def eliminar_item_debe():
    selected_item = tree_debe.selection()
    if not selected_item:
        return
    
    index = int(tree_debe.item(selected_item[0], "values")[0]) - 1
    debe_items.pop(index)
    actualizar_treeview_debe()

def eliminar_item_haber():
    selected_item = tree_haber.selection()
    if not selected_item:
        return
    
    index = int(tree_haber.item(selected_item[0], "values")[0]) - 1
    haber_items.pop(index)
    actualizar_treeview_haber()

# --- Configuración de tema oscuro premium ---
COLOR_PRIMARIO = "#4fc3f7"       # Azul claro moderno
COLOR_SECUNDARIO = "#2a3f54"     # Azul oscuro elegante
COLOR_FONDO = "#1a1a2e"          # Fondo oscuro premium
COLOR_TEXTO = "#ffffff"          # Texto blanco
COLOR_ACENTO = "#00c292"         # Verde azulado para acentos
COLOR_PANEL = "#16213e"          # Color de paneles


# --- Configuración de la ventana principal con grid ---
root = ttk.Window(themename="superhero")
root.title("Registro Contable Premium")
root.state("zoomed")

# Configurar grid principal
root.grid_columnconfigure(1, weight=1)
root.grid_rowconfigure(1, weight=1)

# Configurar fuente global
estilo_global = ttk.Style()
estilo_global.configure('.', font=('Segoe UI', 11), foreground=COLOR_TEXTO)
estilo_global.configure('Titulo.TLabel', font=('Segoe UI', 18, 'bold'), anchor='center')
estilo_global.configure('Subtitulo.TLabel', font=('Segoe UI', 14))
estilo_global.configure('Boton.TButton', padding=10)
estilo_global.configure('White.TButton', foreground='white')
estilo_global.configure('TCombobox', padding=5)

# Configuración especial para Treeview
estilo_global.configure("Treeview", 
                      background=COLOR_SECUNDARIO,
                      fieldbackground=COLOR_SECUNDARIO,
                      foreground=COLOR_TEXTO,
                      rowheight=35,
                      font=('Segoe UI', 10))
estilo_global.configure("Treeview.Heading", 
                      background=COLOR_ACENTO,
                      foreground="white",
                      font=('Segoe UI', 10, 'bold'))
estilo_global.map("Treeview", background=[('selected', COLOR_PRIMARIO)])

# --- Variables globales ---
operaciones_registradas = []
monto_entry = None
glosa_entry = None
tree_operaciones = None
confirmacion_frame = None
panel_inicio = None
formulario_asientos = None
cuenta_debe_var = None
cuenta_haber_var = None
actividad_var = None
frame_confirmacion = None
frame_botones_accion = None
mostrar_botones_confirmacion = None
confirmar_registro = None
ocultar_botones_confirmacion = None
costos_var = None
actividad_var = None
costos_combobox = None
actividad_combobox = None
traceback = None
frame_agregar_debe = None
frame_agregar_haber = None
mes_actual = datetime.now().strftime("%m/%Y")
meses_disponibles = [mes_actual]  # Inicialmente solo el mes actual
historico_debe_items = []
historico_haber_items = []
frame_confirmacion = None

# --- Nuevas variables para múltiples cuentas ---
debe_items = []    # Lista para cuentas de débito
haber_items = []   # Lista para cuentas de crédito
current_operation_id = None
tree_debe = None   # Treeview para débitos
tree_haber = None  # Treeview para créditos
lbl_total_debe = None  # Label para total débito
lbl_total_haber = None  # Label para total crédito

# --- Configuración inicial de la base de datos ---
def inicializar_base_datos():
    # Crear directorio de datos si no existe
    if not os.path.exists('data'):
        os.makedirs('data')
    
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    # Verificar si la tabla existe
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='operaciones'")
    tabla_existe = cursor.fetchone()
    
    if tabla_existe:
        # Si la tabla existe, verificar su estructura
        cursor.execute("PRAGMA table_info(operaciones)")
        columnas = [col[1] for col in cursor.fetchall()]
        
        if 'operacion_id' not in columnas:
            # Actualizar estructura si es necesario
            conn.close()
            actualizar_estructura_bd()
            return
    else:
        # Crear tabla nueva con la estructura correcta
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS operaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            operacion_id INTEGER NOT NULL,
            fecha TEXT NOT NULL,
            cuenta_debe TEXT NOT NULL,
            cuenta_haber TEXT NOT NULL,
            monto REAL NOT NULL,
            moneda TEXT NOT NULL,
            costos TEXT,
            actividad TEXT,
            glosa TEXT
        )
        ''')
    
    conn.commit()
    conn.close()

def cargar_operaciones_db():
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT 
            id,
            operacion_id, 
            fecha, 
            cuenta_debe, 
            cuenta_haber, 
            monto, 
            moneda, 
            costos, 
            actividad, 
            glosa 
        FROM operaciones 
        ORDER BY operacion_id DESC, fecha DESC
    ''')

    columnas = [desc[0] for desc in cursor.description]
    operaciones = []
    
    for fila in cursor.fetchall():
        operacion = dict(zip(columnas, fila))
        operaciones.append({
            'id': operacion['id'],
            'operacion_id': operacion['operacion_id'],  # Asegúrate de incluir esto
            'fecha': operacion['fecha'],
            'debe': operacion['cuenta_debe'],
            'haber': operacion['cuenta_haber'],
            'cuenta_debe': operacion['cuenta_debe'],
            'cuenta_haber': operacion['cuenta_haber'],
            'monto': "{:,.2f}".format(operacion['monto']),
            'moneda': operacion['moneda'],
            'costos': operacion['costos'],
            'actividad': operacion['actividad'],
            'glosa': operacion['glosa']
        })
    
    conn.close()
    return operaciones

def actualizar_estructura_bd():
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    try:
        # Verificar si la columna operacion_id ya existe
        cursor.execute("PRAGMA table_info(operaciones)")
        columnas = [col[1] for col in cursor.fetchall()]
        
        if 'operacion_id' not in columnas:
            # Paso 1: Crear una tabla temporal con la nueva estructura
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS operaciones_temp (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                operacion_id INTEGER NOT NULL,
                fecha TEXT NOT NULL,
                cuenta_debe TEXT NOT NULL,
                cuenta_haber TEXT NOT NULL,
                monto REAL NOT NULL,
                moneda TEXT NOT NULL,
                costos TEXT,
                actividad TEXT,
                glosa TEXT
            )
            ''')
            
            # Paso 2: Copiar los datos de la tabla antigua a la nueva
            cursor.execute('''
            INSERT INTO operaciones_temp (id, fecha, cuenta_debe, cuenta_haber, monto, moneda, costos, actividad, glosa)
            SELECT id, fecha, cuenta_debe, cuenta_haber, monto, moneda, costos, actividad, glosa FROM operaciones
            ''')
            
            # Paso 3: Asignar números de operación secuenciales
            cursor.execute('''
            UPDATE operaciones_temp
            SET operacion_id = id
            ''')
            
            # Paso 4: Eliminar la tabla antigua
            cursor.execute('DROP TABLE operaciones')
            
            # Paso 5: Renombrar la tabla temporal
            cursor.execute('ALTER TABLE operaciones_temp RENAME TO operaciones')
            
            conn.commit()
            print("Estructura de la base de datos actualizada exitosamente.")
        else:
            print("La estructura de la base de datos ya está actualizada.")
            
    except Exception as e:
        conn.rollback()
        print(f"Error al actualizar la estructura: {e}")
    finally:
        conn.close()

def cargar_datos_operaciones(tree):
    for item in tree.get_children():
        tree.delete(item)
    
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    try:
        # Consulta modificada para obtener operacion_id pero ordenar por fecha
        cursor.execute('''
            SELECT 
                operacion_id,  -- Mostramos el ID de operación agrupado
                fecha,
                cuenta_debe,
                cuenta_haber,
                monto,
                costos,
                actividad,
                glosa
            FROM operaciones
            WHERE glosa != 'Saldo inicial' OR glosa IS NULL
            ORDER BY fecha DESC, id DESC  -- Mantenemos el orden por fecha
        ''')

        for row in cursor.fetchall():
            tree.insert("", "end", values=(
                row[0],  # operacion_id
                row[1],  # Fecha
                row[2] if row[2] else "",  # Cuenta débito
                row[3] if row[3] else "",  # Cuenta haber
                f"{row[4]:,.2f}",  # Monto
                row[5],  # Costos
                row[6],  # Actividad
                row[7] if row[7] else ""  # Glosa
            ))
            
    except sqlite3.Error as e:
        print(f"Error al cargar operaciones: {e}")
        messagebox.showerror("Error", f"No se pudieron cargar las operaciones: {str(e)}")
    finally:
        conn.close()

def obtener_proximo_id():
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    try:
        # Obtener el mes actual en formato MM/YYYY
        mes_actual = datetime.now().strftime("%m/%Y")
        
        # Buscar el máximo ID de operaciones del mes actual
        cursor.execute('''
            SELECT MAX(operacion_id) 
            FROM operaciones 
            WHERE substr(fecha, 4, 7) = ?
        ''', (mes_actual,))
        
        max_id = cursor.fetchone()[0]
        
        # Si no hay operaciones este mes, comenzar desde 1
        return max_id + 1 if max_id is not None else 1
        
    except sqlite3.Error as e:
        print(f"Error al obtener próximo ID: {e}")
        return 1
    finally:
        conn.close()

# --- Estructura completa de cuentas contables ---
def crear_arbol_cuentas():
    """Estructura completa de cuentas contables con todas las subcuentas"""
    return {
        "10": {"nombre": "Efectivo y equivalentes de efectivo", "subcuentas": {
            "101": {"nombre": "Caja"},
            "102": {"nombre": "Fondos fijos"},
            "103": {"nombre": "Efectivo en tránsito"},
            "104": {"nombre": "Cuentas corrientes en instituciones financieras", "subcuentas": {
                "1041": {"nombre": "Cuentas corrientes operativas"},
                "1042": {"nombre": "Cuentas corrientes para fines específicos"}
            }},
            "105": {"nombre": "Otros equivalentes de efectivo", "subcuentas": {
                "1051": {"nombre": "Otros equivalentes de efectivo"}
            }},
            "106": {"nombre": "Depósitos en instituciones financieras", "subcuentas": {
                "1061": {"nombre": "Depósitos de ahorro"},
                "1062": {"nombre": "Depósitos a plazo"}
            }},
            "107": {"nombre": "Fondos sujetos a restricción", "subcuentas": {
                "1071": {"nombre": "Fondos sujetos a restricción"}
            }}
        }},
        "11": {"nombre": "Inversiones financieras", "subcuentas": {
            "111": {"nombre": "Inversiones mantenidas para negociación", "subcuentas": {
                "1111": {"nombre": "Valores emitidos o garantizados por el Estado", "subcuentas": {
                    "11111": {"nombre": "Costo"},
                    "11112": {"nombre": "Valor razonable"}
                }},
                "1112": {"nombre": "Valores emitidos por el sistema financiero", "subcuentas": {
                    "11121": {"nombre": "Costo"},
                    "11122": {"nombre": "Valor razonable"}
                }},
                "1113": {"nombre": "Valores emitidos por empresas", "subcuentas": {
                    "11131": {"nombre": "Costo"},
                    "11132": {"nombre": "Valor razonable"}
                }},
                "1114": {"nombre": "Otros títulos representativos de deuda", "subcuentas": {
                    "11141": {"nombre": "Costo"},
                    "11142": {"nombre": "Valor razonable"}
                }},
                "1115": {"nombre": "Participaciones en entidades", "subcuentas": {
                    "11151": {"nombre": "Costo"},
                    "11152": {"nombre": "Valor razonable"}
                }}
            }},
            "112": {"nombre": "Inversiones disponibles para la venta", "subcuentas": {
                "1121": {"nombre": "Valores emitidos o garantizados por el Estado", "subcuentas": {
                    "11211": {"nombre": "Costo"},
                    "11212": {"nombre": "Valor razonable"}
                }},
                "1122": {"nombre": "Valores emitidos por el sistema financiero", "subcuentas": {
                    "11221": {"nombre": "Costo"},
                    "11222": {"nombre": "Valor razonable"}
                }},
                "1123": {"nombre": "Valores emitidos por empresas", "subcuentas": {
                    "11231": {"nombre": "Costo"},
                    "11232": {"nombre": "Valor razonable"}
                }},
                "1124": {"nombre": "Otros títulos representativos de deuda", "subcuentas": {
                    "11241": {"nombre": "Costo"},
                    "11242": {"nombre": "Valor razonable"}
                }}
            }}
        }},
        "12": {"nombre": "Cuentas por cobrar comerciales - Terceros", "subcuentas": {
            "121": {"nombre": "Facturas, boletas y otros comprobantes por cobrar", "subcuentas": {
                "1211": {"nombre": "No emitidas"},
                "1212": {"nombre": "Emitidas en cartera"},
                "1213": {"nombre": "En cobranza"},
                "1214": {"nombre": "En descuento"}
            }},
            "122": {"nombre": "Anticipos de clientes"},
            "123": {"nombre": "Letras por cobrar", "subcuentas": {
                "1231": {"nombre": "En cartera"},
                "1232": {"nombre": "En cobranza"},
                "1233": {"nombre": "En descuento"}
            }}
        }},
        "13": {"nombre": "Cuentas por cobrar comerciales - Relacionadas", "subcuentas": {
            "131": {"nombre": "Facturas, boletas y otros comprobantes por cobrar", "subcuentas": {
                "1311": {"nombre": "No emitidas", "subcuentas": {
                    "13111": {"nombre": "Matriz"},
                    "13112": {"nombre": "Subsidiarias"},
                    "13113": {"nombre": "Asociadas"},
                    "13114": {"nombre": "Sucursales"},
                    "13115": {"nombre": "Otros"}
                }},
                "1312": {"nombre": "Emitidas en cartera", "subcuentas": {
                    "13121": {"nombre": "Matriz"},
                    "13122": {"nombre": "Subsidiarias"},
                    "13123": {"nombre": "Asociadas"},
                    "13124": {"nombre": "Sucursales"},
                    "13125": {"nombre": "Otros"}
                }},
                "1313": {"nombre": "En cobranza", "subcuentas": {
                    "13131": {"nombre": "Matriz"},
                    "13132": {"nombre": "Subsidiarias"},
                    "13133": {"nombre": "Asociadas"},
                    "13134": {"nombre": "Sucursales"},
                    "13135": {"nombre": "Otros"}
                }},
                "1314": {"nombre": "En descuento", "subcuentas": {
                    "13141": {"nombre": "Matriz"},
                    "13142": {"nombre": "Subsidiarias"},
                    "13143": {"nombre": "Asociadas"},
                    "13144": {"nombre": "Sucursales"},
                    "13145": {"nombre": "Otros"}
                }}
            }},
            "132": {"nombre": "Anticipos recibidos", "subcuentas": {
                "1321": {"nombre": "Anticipos recibidos", "subcuentas": {
                    "13211": {"nombre": "Matriz"},
                    "13212": {"nombre": "Subsidiarias"},
                    "13213": {"nombre": "Asociadas"},
                    "13214": {"nombre": "Sucursales"},
                    "13215": {"nombre": "Otros"}
                }}
            }}
        }},
        "14": {"nombre": "Cuentas por cobrar al personal, a los accionistas (socios), directores y gerentes", "subcuentas": {
            "141": {"nombre": "Personal", "subcuentas": {
                "1411": {"nombre": "Préstamos"},
                "1412": {"nombre": "Adelanto de remuneraciones"},
                "1413": {"nombre": "Entregas a rendir cuenta"},
                "1419": {"nombre": "Otras cuentas por cobrar al personal"}
            }},
            "142": {"nombre": "Accionistas (o socios)", "subcuentas": {
                "1421": {"nombre": "Suscripciones por cobrar a socios o accionistas"},
                "1422": {"nombre": "Préstamos"}
            }},
            "143": {"nombre": "Directores", "subcuentas": {
                "1431": {"nombre": "Préstamos"},
                "1432": {"nombre": "Adelanto de dietas"},
                "1433": {"nombre": "Entregas a rendir cuenta"}
            }},
            "144": {"nombre": "Gerentes", "subcuentas": {
                "1441": {"nombre": "Préstamos"},
                "1442": {"nombre": "Adelanto de remuneraciones"},
                "1443": {"nombre": "Entregas a rendir cuenta"}
            }},
            "148": {"nombre": "Diversas"}
        }},
        "16": {"nombre": "Cuentas por cobrar diversas - Terceros", "subcuentas": {
            "161": {"nombre": "Préstamos", "subcuentas": {
                "1611": {"nombre": "Con garantía"},
                "1612": {"nombre": "Sin garantía"}
            }},
            "162": {"nombre": "Reclamaciones a terceros", "subcuentas": {
                "1621": {"nombre": "Compañías aseguradoras"},
                "1622": {"nombre": "Transportadoras"},
                "1623": {"nombre": "Servicios públicos"},
                "1624": {"nombre": "Tributos"},
                "1629": {"nombre": "Otras"}
            }},
            "163": {"nombre": "Intereses, regalías y dividendos", "subcuentas": {
                "1631": {"nombre": "Intereses"},
                "1632": {"nombre": "Regalías"},
                "1633": {"nombre": "Dividendos"}
            }},
            "164": {"nombre": "Depósitos otorgados en garantía", "subcuentas": {
                "1641": {"nombre": "Préstamos de instituciones financieras"},
                "1642": {"nombre": "Préstamos de instituciones no financieras"},
                "1644": {"nombre": "Depósitos en garantía por alquileres"},
                "1649": {"nombre": "Otros depósitos en garantía"}
            }},
            "165": {"nombre": "Venta de activo inmovilizado", "subcuentas": {
                "1651": {"nombre": "Inversión mobiliaria"},
                "1652": {"nombre": "Inversión inmobiliaria"},
                "1653": {"nombre": "Inmuebles, maquinaria y equipo"},
                "1654": {"nombre": "Intangibles"},
                "1655": {"nombre": "Activos biológicos"}
            }},
            "166": {"nombre": "Activos por instrumentos financieros", "subcuentas": {
                "1661": {"nombre": "Instrumentos financieros primarios"},
                "1662": {"nombre": "Instrumentos financieros derivados", "subcuentas": {
                    "16621": {"nombre": "Cartera de negociación"},
                    "16622": {"nombre": "Instrumentos de cobertura"}
                }}
            }},
            "168": {"nombre": "Otras cuentas por cobrar diversas", "subcuentas": {
                "1681": {"nombre": "Entregas a rendir cuenta a terceros"},
                "1682": {"nombre": "Otras cuentas por cobrar diversas"}
            }}
        }},
        "17": {"nombre": "Cuentas por cobrar diversas - Relacionadas", "subcuentas": {
            "171": {"nombre": "Préstamos", "subcuentas": {
                "1711": {"nombre": "Con garantía", "subcuentas": {
                    "17111": {"nombre": "Matriz"},
                    "17112": {"nombre": "Subsidiarias"},
                    "17113": {"nombre": "Asociadas"},
                    "17114": {"nombre": "Sucursales"},
                    "17115": {"nombre": "Otros"}
                }},
                "1712": {"nombre": "Sin garantía", "subcuentas": {
                    "17121": {"nombre": "Matriz"},
                    "17122": {"nombre": "Subsidiarias"},
                    "17123": {"nombre": "Asociadas"},
                    "17124": {"nombre": "Sucursales"},
                    "17125": {"nombre": "Otros"}
                }}
            }},
            "173": {"nombre": "Intereses, regalías y dividendos", "subcuentas": {
                "1731": {"nombre": "Intereses", "subcuentas": {
                    "17311": {"nombre": "Matriz"},
                    "17312": {"nombre": "Subsidiarias"},
                    "17313": {"nombre": "Asociadas"},
                    "17314": {"nombre": "Sucursales"},
                    "17315": {"nombre": "Otros"}
                }},
                "1732": {"nombre": "Regalías", "subcuentas": {
                    "17321": {"nombre": "Matriz"},
                    "17322": {"nombre": "Subsidiarias"},
                    "17323": {"nombre": "Asociadas"},
                    "17324": {"nombre": "Sucursales"},
                    "17325": {"nombre": "Otros"}
                }},
                "1733": {"nombre": "Dividendos", "subcuentas": {
                    "17331": {"nombre": "Matriz"},
                    "17332": {"nombre": "Subsidiarias"},
                    "17333": {"nombre": "Asociadas"},
                    "17334": {"nombre": "Otros"}
                }}
            }},
            "174": {"nombre": "Depósitos otorgados en garantía"},
            "175": {"nombre": "Venta de activo inmovilizado", "subcuentas": {
                "1751": {"nombre": "Inversión mobiliaria"},
                "1752": {"nombre": "Inversión inmobiliaria"},
                "1753": {"nombre": "Inmuebles, maquinaria y equipo"},
                "1754": {"nombre": "Intangibles"},
                "1755": {"nombre": "Activos biológicos"}
            }},
            "176": {"nombre": "Activos por instrumentos financieros"},
            "178": {"nombre": "Otras cuentas por cobrar diversas"}
        }},
        "18": {"nombre": "Servicios y otros contratados por anticipado", "subcuentas": {
            "181": {"nombre": "Costos financieros"},
            "182": {"nombre": "Seguros"},
            "183": {"nombre": "Alquileres"},
            "184": {"nombre": "Primas pagadas por opciones"},
            "185": {"nombre": "Mantenimiento de activos inmovilizados"},
            "189": {"nombre": "Otros gastos contratados por anticipado"}
        }},
        "19": {"nombre": "Estimación de cuentas de cobranza dudosa", "subcuentas": {
            "191": {"nombre": "Cuentas por cobrar comerciales - Terceros", "subcuentas": {
                "1911": {"nombre": "Facturas, boletas y otros comprobantes por cobrar"},
                "1913": {"nombre": "Letras por cobrar"}
            }},
            "192": {"nombre": "Cuentas por cobrar comerciales - Relacionadas", "subcuentas": {
                "1921": {"nombre": "Facturas, boletas y otros comprobantes por cobrar"},
                "1922": {"nombre": "Letras por cobrar"}
            }},
            "193": {"nombre": "Cuentas por cobrar al personal, a los accionistas (socios), directores y gerentes", "subcuentas": {
                "1931": {"nombre": "Personal"},
                "1932": {"nombre": "Accionistas (o socios)"},
                "1933": {"nombre": "Directores"},
                "1934": {"nombre": "Gerentes"},
                "1938": {"nombre": "Diversas"}
            }},
            "194": {"nombre": "Cuentas por cobrar diversas - Terceros", "subcuentas": {
                "1941": {"nombre": "Préstamos"},
                "1942": {"nombre": "Reclamaciones a terceros"},
                "1943": {"nombre": "Intereses, regalías y dividendos"},
                "1944": {"nombre": "Depósitos otorgados en garantía"},
                "1945": {"nombre": "Venta de activo inmovilizado"},
                "1946": {"nombre": "Activos por instrumentos financieros"},
                "1949": {"nombre": "Otras cuentas por cobrar diversas"}
            }},
            "195": {"nombre": "Cuentas por cobrar diversas - Relacionadas", "subcuentas": {
                "1951": {"nombre": "Préstamos"},
                "1953": {"nombre": "Intereses, regalías y dividendos"},
                "1954": {"nombre": "Depósitos otorgados en garantía"},
                "1955": {"nombre": "Venta de activo inmovilizado"},
                "1956": {"nombre": "Activos por instrumentos financieros"},
                "1958": {"nombre": "Otras cuentas por cobrar diversas"}
            }}
        }},
        "20": {"nombre": "Mercaderías", "subcuentas": {
            "201": {"nombre": "Mercaderías manufacturadas", "subcuentas": {
                "2011": {"nombre": "Costo"},
                "2012": {"nombre": "Valor razonable"}
            }},
            "202": {"nombre": "Mercaderías de extracción"},
            "203": {"nombre": "Mercaderías agropecuarias y piscícolas", "subcuentas": {
                "2031": {"nombre": "De origen animal", "subcuentas": {
                    "20311": {"nombre": "Costo"},
                    "20312": {"nombre": "Valor razonable"}
                }},
                "2032": {"nombre": "De origen vegetal", "subcuentas": {
                    "20321": {"nombre": "Costo"},
                    "20322": {"nombre": "Valor razonable"}
                }}
            }},
            "204": {"nombre": "Mercaderías inmuebles"},
            "208": {"nombre": "Otras mercaderías"}
        }},
        "21": {"nombre": "Productos terminados", "subcuentas": {
            "211": {"nombre": "Productos manufacturados"},
            "212": {"nombre": "Productos de extracción terminados"},
            "213": {"nombre": "Productos agropecuarios y piscícolas terminados", "subcuentas": {
                "2131": {"nombre": "De origen animal", "subcuentas": {
                    "21311": {"nombre": "Costo"},
                    "21312": {"nombre": "Valor razonable"}
                }},
                "2132": {"nombre": "De origen vegetal", "subcuentas": {
                    "21321": {"nombre": "Costo"},
                    "21322": {"nombre": "Valor razonable"}
                }}
            }},
            "214": {"nombre": "Productos inmuebles"},
            "215": {"nombre": "Existencias de servicios terminados"},
            "217": {"nombre": "Otros productos terminados"}
        }},
        "22": {"nombre": "Subproductos, desechos y desperdicios", "subcuentas": {
            "221": {"nombre": "Subproductos"},
            "222": {"nombre": "Desechos y desperdicios"}
        }},
        "23": {"nombre": "Productos en proceso", "subcuentas": {
            "231": {"nombre": "Productos en proceso de manufactura"},
            "232": {"nombre": "Productos extraídos en proceso de transformación"},
            "233": {"nombre": "Productos agropecuarios y piscícolas en proceso", "subcuentas": {
                "2331": {"nombre": "De origen animal", "subcuentas": {
                    "23311": {"nombre": "Costo"},
                    "23312": {"nombre": "Valor razonable"}
                }},
                "2332": {"nombre": "De origen vegetal", "subcuentas": {
                    "23321": {"nombre": "Costo"},
                    "23322": {"nombre": "Valor razonable"}
                }}
            }},
            "234": {"nombre": "Productos inmuebles en proceso"},
            "235": {"nombre": "Existencias de servicios en proceso"},
            "237": {"nombre": "Otros productos en proceso"},
            "238": {"nombre": "Costos de financiación - Productos en proceso"}
        }},
        "24": {"nombre": "Materias primas", "subcuentas": {
            "241": {"nombre": "Materias primas para productos manufacturados"},
            "242": {"nombre": "Materias primas para productos de extracción"},
            "243": {"nombre": "Materias primas para productos agropecuarios y piscícolas"},
            "244": {"nombre": "Materias primas para productos inmuebles"}
        }},
        "25": {"nombre": "Materiales auxiliares, suministros y repuestos", "subcuentas": {
            "251": {"nombre": "Materiales auxiliares"},
            "252": {"nombre": "Suministros", "subcuentas": {
                "2521": {"nombre": "Combustibles"},
                "2522": {"nombre": "Lubricantes"},
                "2523": {"nombre": "Energía"},
                "2524": {"nombre": "Otros suministros"}
            }},
            "253": {"nombre": "Repuestos"}
        }},
        "26": {"nombre": "Envases y embalajes", "subcuentas": {
            "261": {"nombre": "Envases"},
            "262": {"nombre": "Embalajes"}
        }},
        "27": {"nombre": "Activos no corrientes mantenidos para la venta", "subcuentas": {
            "271": {"nombre": "Inversiones inmobiliarias", "subcuentas": {
                "2711": {"nombre": "Terrenos", "subcuentas": {
                    "27111": {"nombre": "Valor razonable"},
                    "27112": {"nombre": "Costo"},
                    "27113": {"nombre": "Revaluación"}
                }},
                "2712": {"nombre": "Edificaciones", "subcuentas": {
                    "27121": {"nombre": "Valor razonable"},
                    "27122": {"nombre": "Costo"},
                    "27123": {"nombre": "Revaluación"},
                    "27124": {"nombre": "Costos de financiación"}
                }}
            }},
            "272": {"nombre": "Inmuebles, maquinaria y equipo", "subcuentas": {
                "2721": {"nombre": "Terrenos", "subcuentas": {
                    "27211": {"nombre": "Valor razonable"},
                    "27212": {"nombre": "Costo"},
                    "27213": {"nombre": "Revaluación"}
                }},
                "2722": {"nombre": "Edificaciones", "subcuentas": {
                    "27221": {"nombre": "Costo de adquisición o construcción"},
                    "27222": {"nombre": "Revaluación"},
                    "27223": {"nombre": "Costo de financiación"}
                }},
                "2723": {"nombre": "Maquinarias y equipos de explotación", "subcuentas": {
                    "27231": {"nombre": "Costo de adquisición o construcción"},
                    "27232": {"nombre": "Revaluación"},
                    "27233": {"nombre": "Costo de financiación"}
                }},
                "2724": {"nombre": "Equipo de transporte", "subcuentas": {
                    "27241": {"nombre": "Costo"},
                    "27242": {"nombre": "Revaluación"}
                }},
                "2725": {"nombre": "Muebles y enseres", "subcuentas": {
                    "27251": {"nombre": "Costo"},
                    "27252": {"nombre": "Revaluación"}
                }},
                "2726": {"nombre": "Equipos diversos", "subcuentas": {
                    "27261": {"nombre": "Costo"},
                    "27262": {"nombre": "Revaluación"}
                }},
                "2727": {"nombre": "Herramientas y unidades de reemplazo", "subcuentas": {
                    "27271": {"nombre": "Costo"},
                    "27272": {"nombre": "Revaluación"}
                }}
            }},
            "273": {"nombre": "Intangibles", "subcuentas": {
                "2731": {"nombre": "Concesiones, licencias y derechos", "subcuentas": {
                    "27311": {"nombre": "Costo"},
                    "27312": {"nombre": "Revaluación"}
                }},
                "2732": {"nombre": "Patentes y propiedad industrial", "subcuentas": {
                    "27321": {"nombre": "Costo"},
                    "27322": {"nombre": "Revaluación"}
                }},
                "2733": {"nombre": "Programas de computadora (software)", "subcuentas": {
                    "27331": {"nombre": "Costo"},
                    "27332": {"nombre": "Revaluación"}
                }},
                "2734": {"nombre": "Costos de exploración y desarrollo", "subcuentas": {
                    "27341": {"nombre": "Costo"},
                    "27342": {"nombre": "Revaluación"}
                }},
                "2735": {"nombre": "Fórmulas, diseños y prototipos", "subcuentas": {
                    "27351": {"nombre": "Costo"},
                    "27352": {"nombre": "Revaluación"}
                }},
                "2736": {"nombre": "Reservas de recursos extraíbles", "subcuentas": {
                    "27361": {"nombre": "Costo"},
                    "27362": {"nombre": "Revaluación"}
                }},
                "2739": {"nombre": "Otros activos intangibles", "subcuentas": {
                    "27391": {"nombre": "Costo"},
                    "27392": {"nombre": "Revaluación"}
                }}
            }}
        }},
        "28": {"nombre": "Existencias por recibir", "subcuentas": {
            "281": {"nombre": "Mercaderías"},
            "284": {"nombre": "Materias primas"},
            "285": {"nombre": "Materiales auxiliares, suministros y repuestos"},
            "286": {"nombre": "Envases y embalajes"}
        }},
        "29": {"nombre": "Desvalorización de existencias", "subcuentas": {
            "291": {"nombre": "Mercaderías", "subcuentas": {
                "2911": {"nombre": "Mercaderías manufacturadas"},
                "2912": {"nombre": "Mercaderías de extracción"},
                "2913": {"nombre": "Mercaderías agropecuarias y piscícolas"},
                "2914": {"nombre": "Mercaderías inmuebles"},
                "2918": {"nombre": "Otras mercaderías"}
            }},
            "292": {"nombre": "Productos terminados", "subcuentas": {
                "2921": {"nombre": "Productos manufacturados"},
                "2922": {"nombre": "Productos de extracción terminados"},
                "2923": {"nombre": "Productos agropecuarios y piscícolas terminados"},
                "2924": {"nombre": "Productos inmuebles"},
                "2925": {"nombre": "Existencias de servicios terminados"},
                "2927": {"nombre": "Otros productos terminados"},
                "2928": {"nombre": "Costos de financiación - Productos terminados"}
            }},
            "293": {"nombre": "Subproductos, desechos y desperdicios", "subcuentas": {
                "2931": {"nombre": "Subproductos"},
                "2932": {"nombre": "Desechos y desperdicios"}
            }},
            "294": {"nombre": "Productos en proceso", "subcuentas": {
                "2941": {"nombre": "Productos en proceso de manufactura"},
                "2942": {"nombre": "Productos extraídos en proceso de transformación"},
                "2943": {"nombre": "Productos agropecuarios y piscícolas en proceso"},
                "2944": {"nombre": "Productos inmuebles en proceso"},
                "2945": {"nombre": "Existencias de servicios en proceso"},
                "2947": {"nombre": "Otros productos en proceso"},
                "2948": {"nombre": "Costos de financiación - Productos en proceso"}
            }},
            "295": {"nombre": "Materias primas", "subcuentas": {
                "2951": {"nombre": "Materias primas para productos manufacturados"},
                "2952": {"nombre": "Materias primas para productos de extracción"},
                "2953": {"nombre": "Materias primas para productos agropecuarios y piscícolas"},
                "2954": {"nombre": "Materias primas para productos inmuebles"}
            }},
            "296": {"nombre": "Materiales auxiliares, suministros y repuestos", "subcuentas": {
                "2961": {"nombre": "Materiales auxiliares"},
                "2962": {"nombre": "Suministros"},
                "2963": {"nombre": "Repuestos"}
            }},
            "297": {"nombre": "Envases y embalajes", "subcuentas": {
                "2971": {"nombre": "Envases"},
                "2972": {"nombre": "Embalajes"}
            }},
            "298": {"nombre": "Existencias por recibir", "subcuentas": {
                "2981": {"nombre": "Mercaderías"},
                "2982": {"nombre": "Materias primas"},
                "2983": {"nombre": "Materiales auxiliares, suministros y repuestos"},
                "2984": {"nombre": "Envases y embalajes"}
            }}
        }},
        "30": {"nombre": "Inversiones mobiliarias", "subcuentas": {
            "301": {"nombre": "Inversiones a ser mantenidas hasta el vencimiento", "subcuentas": {
                "3011": {"nombre": "Instrumentos financieros representativos de deuda", "subcuentas": {
                    "30111": {"nombre": "Valores emitidos o garantizados por el Estado"}, "30112": {"nombre": "Valores emitidos por el sistema financiero"}, "30113": {"nombre": "Valores emitidos por las empresas"}, "30114": {"nombre": "Valores emitidos por otras entidades"}
                }},
            "302": {"nombre": "Instrumentos financieros representativos de derecho patrimonial", "subcuentas": {
                "3021": {"nombre": "Certificados de suscripción preferente"}, "3022": {"nombre": "Acciones representativas de capital social – Comunes", "subcuentas": {"30221": {"nombre": "Costo"}, "30222": {"nombre": "Valor razonable"}, "30223": {"nombre": "Participación patrimonial"}}},
                "3023": {"nombre": "Acciones representativas de capital social – Preferentes", "subcuentas": {"30231": {"nombre": "Costo"}, "30232": {"nombre": "Valor razonable"}, "30233": {"nombre": "Participación patrimonial"}}},
                "3024": {"nombre": "Acciones de inversión", "subcuentas": {"30241": {"nombre": "Costo"}, "30242": {"nombre": "Valor razonable"}, "30243": {"nombre": "Participación patrimonial"}}},
                "3025": {"nombre": "Certificados de participación de fondos de inversión", "subcuentas": {"30251": {"nombre": "Costo"}, "30252": {"nombre": "Valor razonable"}}},
                "3026": {"nombre": "Certificados de participación de fondos mutuos", "subcuentas": {"30261": {"nombre": "Costo"}, "30262": {"nombre": "Valor razonable"}}},
                "3027": {"nombre": "Participaciones en asociaciones en participación y consorcios", "subcuentas": {"30271": {"nombre": "Costo"}, "30272": {"nombre": "Valor razonable"}, "30273": {"nombre": "Participación patrimonial"}}},
                "3028": {"nombre": "Otros títulos representativos de patrimonio", "subcuentas": {"30281": {"nombre": "Costo"}, "30282": {"nombre": "Valor razonable"}}}
                }},
             "308": {"nombre": "Inversiones mobiliarias – Acuerdos de compra", "subcuentas": {
                "3081": {"nombre": "Inversiones a ser mantenidas hasta el vencimiento – Acuerdo de compra", "subcuentas": {"30811": {"nombre": "Costo"}, "30812": {"nombre": "Valor razonable"}}},
                "3082": {"nombre": "Instrumentos financieros representativos de derecho patrimonial – Acuerdo de compra", "subcuentas": {"30821": {"nombre": "Costo"}, "30822": {"nombre": "Valor razonable"}}}
                }}
            }}
        }},
        "31": {"nombre": "Inversiones inmobiliarias", "subcuentas": {
            "311": {"nombre": "Terrenos", "subcuentas": {
                "3111": {"nombre": "Urbanos", "subcuentas": {"31111": {"nombre": "Valor razonable"}, "31112": {"nombre": "Costo"}, "31113": {"nombre": "Revaluación"}}},
                "3112": {"nombre": "Rurales", "subcuentas": {"31121": {"nombre": "Valor razonable"}, "31122": {"nombre": "Costo"}, "31123": {"nombre": "Revaluación"}}}
            }},
            "312": {"nombre": "Edificaciones", "subcuentas": {
                "3121": {"nombre": "Edificaciones administrativas", "subcuentas": {"31211": {"nombre": "Costo de adquisición o construcción"}, "31212": {"nombre": "Revaluación"}, "31213": {"nombre": "Costo de financiación"}}},
                "3122": {"nombre": "Almacenes", "subcuentas": {"31221": {"nombre": "Costo de adquisición o construcción"}, "31222": {"nombre": "Revaluación"}, "31223": {"nombre": "Costo de financiación"}}},
                "3123": {"nombre": "Edificaciones para producción", "subcuentas": {"31231": {"nombre": "Costo de adquisición o construcción"}, "31232": {"nombre": "Revaluación"}, "31233": {"nombre": "Costo de financiación"}}},
                "3124": {"nombre": "Instalaciones", "subcuentas": {"31241": {"nombre": "Costo de adquisición o construcción"}, "31242": {"nombre": "Revaluación"}, "31243": {"nombre": "Costo de financiación"}}}
            }},
            "313": {"nombre": "Maquinarias y equipos de explotación", "subcuentas": {
                "3131": {"nombre": "Maquinarias", "subcuentas": {"31311": {"nombre": "Costo de adquisición o construcción"}, "31312": {"nombre": "Revaluación"}, "31313": {"nombre": "Costo de financiación"}}},
                "3132": {"nombre": "Equipos de explotación", "subcuentas": {"31321": {"nombre": "Costo de adquisición o construcción"}, "31322": {"nombre": "Revaluación"}, "31323": {"nombre": "Costo de financiación"}}}
            }},
            "314": {"nombre": "Equipo de transporte", "subcuentas": {
                "3141": {"nombre": "Vehículos motorizados", "subcuentas": {"31411": {"nombre": "Costo"}, "31412": {"nombre": "Revaluación"}}},
                "3142": {"nombre": "Vehículos no motorizados", "subcuentas": {"31421": {"nombre": "Costo"}, "31422": {"nombre": "Revaluación"}}}
            }},
            "315": {"nombre": "Muebles y enseres", "subcuentas": {
                "3151": {"nombre": "Muebles", "subcuentas": {"31511": {"nombre": "Costo"}, "31512": {"nombre": "Revaluación"}}},
                "3152": {"nombre": "Enseres", "subcuentas": {"31521": {"nombre": "Costo"}, "31522": {"nombre": "Revaluación"}}}
            }},
            "316": {"nombre": "Equipos diversos", "subcuentas": {
                "3161": {"nombre": "Equipo para procesamiento de información (de cómputo)", "subcuentas": {"31611": {"nombre": "Costo"}, "31612": {"nombre": "Revaluación"}}},
                "3162": {"nombre": "Equipo de comunicación", "subcuentas": {"31621": {"nombre": "Costo"}, "31622": {"nombre": "Revaluación"}}},
                "3163": {"nombre": "Equipo de seguridad", "subcuentas": {"31631": {"nombre": "Costo"}, "31632": {"nombre": "Revaluación"}}},
                "3169": {"nombre": "Otros equipos", "subcuentas": {"31691": {"nombre": "Costo"}, "31692": {"nombre": "Revaluación"}}}
            }},
            "317": {"nombre": "Herramientas y unidades de reemplazo", "subcuentas": {
                "3171": {"nombre": "Herramientas", "subcuentas": {"31711": {"nombre": "Costo"}, "31712": {"nombre": "Revaluación"}}},
                "3172": {"nombre": "Unidades de reemplazo", "subcuentas": {"31721": {"nombre": "Costo"}, "31722": {"nombre": "Revaluación"}}}
            }},
            "318": {"nombre": "Construcciones y obras en curso", "subcuentas": {
                "3181": {"nombre": "Adaptación de terrenos", "subcuentas": {"31811": {"nombre": "Costo de adquisición o construcción"}, "31812": {"nombre": "Revaluación"}}},
                "3182": {"nombre": "Construcciones en curso", "subcuentas": {"31821": {"nombre": "Costo de adquisición o construcción"}, "31822": {"nombre": "Revaluación"}}},
                "3183": {"nombre": "Maquinaria en montaje", "subcuentas": {"31831": {"nombre": "Costo de adquisición o construcción"}, "31832": {"nombre": "Revaluación"}}},
                "3184": {"nombre": "Inversión inmobiliaria en curso", "subcuentas": {"31841": {"nombre": "Costo de adquisición o construcción"}, "31842": {"nombre": "Revaluación"}}}
            }}
        }},
        "32": {"nombre": "Activos adquiridos en arrendamiento financiero", "subcuentas": {
            "321": {"nombre": "Inversiones inmobiliarias", "subcuentas": {
                "3211": {"nombre": "Terrenos", "subcuentas": {"32111": {"nombre": "Valor razonable"}, "32112": {"nombre": "Costo"}, "32113": {"nombre": "Revaluación"}}},
                "3212": {"nombre": "Edificaciones", "subcuentas": {"32121": {"nombre": "Costo de adquisición o construcción"}, "32122": {"nombre": "Revaluación"}}}
            }},
            "322": {"nombre": "Inmuebles, maquinaria y equipo", "subcuentas": {
                "3221": {"nombre": "Terrenos", "subcuentas": {"32211": {"nombre": "Costo"}, "32212": {"nombre": "Revaluación"}}},
                "3222": {"nombre": "Edificaciones", "subcuentas": {"32221": {"nombre": "Costo de adquisición o construcción"}, "32222": {"nombre": "Revaluación"}}},
                "3223": {"nombre": "Maquinarias y equipos de explotación", "subcuentas": {"32231": {"nombre": "Costo de adquisición o construcción"}, "32232": {"nombre": "Revaluación"}}},
                "3224": {"nombre": "Equipo de transporte", "subcuentas": {"32241": {"nombre": "Costo"}, "32242": {"nombre": "Revaluación"}}},
                "3225": {"nombre": "Muebles y enseres", "subcuentas": {"32251": {"nombre": "Costo"}, "32252": {"nombre": "Revaluación"}}},
                "3226": {"nombre": "Equipos diversos", "subcuentas": {"32261": {"nombre": "Costo"}, "32262": {"nombre": "Revaluación"}}},
                "3227": {"nombre": "Herramientas y unidades de reemplazo", "subcuentas": {"32271": {"nombre": "Costo"}, "32272": {"nombre": "Revaluación"}}}
            }}
        }},
        "33": {"nombre": "Inmuebles, maquinarias y equipos", "subcuentas": {
            "331": {"nombre": "Terrenos", "subcuentas": {
                "3311": {"nombre": "Terrenos urbanos", "subcuentas": {"33111": {"nombre": "Valor razonable"}, "33112": {"nombre": "Costo"}, "33113": {"nombre": "Revaluación"}}},
                "3312": {"nombre": "Terrenos rurales", "subcuentas": {"33121": {"nombre": "Valor razonable"}, "33122": {"nombre": "Costo"}, "33123": {"nombre": "Revaluación"}}}
            }},
            "332": {"nombre": "Edificaciones", "subcuentas": {
                "3321": {"nombre": "Edificaciones administrativas", "subcuentas": {"33211": {"nombre": "Costo de adquisición o construcción"}, "33212": {"nombre": "Revaluación"}}},
                "3322": {"nombre": "Edificaciones para producción", "subcuentas": {"33221": {"nombre": "Costo de adquisición o construcción"}, "33222": {"nombre": "Revaluación"}}},
                "3323": {"nombre": "Instalaciones", "subcuentas": {"33231": {"nombre": "Costo de adquisición o construcción"}, "33232": {"nombre": "Revaluación"}}}
            }},
            "333": {"nombre": "Maquinarias y equipos de explotación", "subcuentas": {
                "3331": {"nombre": "Maquinarias", "subcuentas": {"33311": {"nombre": "Costo de adquisición o construcción"}, "33312": {"nombre": "Revaluación"}}},
                "3332": {"nombre": "Equipos de explotación", "subcuentas": {"33321": {"nombre": "Costo de adquisición o construcción"}, "33322": {"nombre": "Revaluación"}}}
            }},
            "334": {"nombre": "Equipo de transporte", "subcuentas": {
                "3341": {"nombre": "Vehículos motorizados", "subcuentas": {"33411": {"nombre": "Costo"}, "33412": {"nombre": "Revaluación"}}},
                "3342": {"nombre": "Vehículos no motorizados", "subcuentas": {"33421": {"nombre": "Costo"}, "33422": {"nombre": "Revaluación"}}}
            }},
            "335": {"nombre": "Muebles y enseres", "subcuentas": {
                "3351": {"nombre": "Muebles", "subcuentas": {"33511": {"nombre": "Costo"}, "33512": {"nombre": "Revaluación"}}},
                "3352": {"nombre": "Enseres", "subcuentas": {"33521": {"nombre": "Costo"}, "33522": {"nombre": "Revaluación"}}}
            }},
            "336": {"nombre": "Equipos diversos", "subcuentas": {
                "3361": {"nombre": "Equipo de procesamiento de información (de cómputo)", "subcuentas": {"33611": {"nombre": "Costo"}, "33612": {"nombre": "Revaluación"}}},
                "3362": {"nombre": "Equipo de comunicación", "subcuentas": {"33621": {"nombre": "Costo"}, "33622": {"nombre": "Revaluación"}}},
                "3363": {"nombre": "Equipo de seguridad", "subcuentas": {"33631": {"nombre": "Costo"}, "33632": {"nombre": "Revaluación"}}}
            }},
            "337": {"nombre": "Herramientas y unidades de reemplazo", "subcuentas": {
                "3371": {"nombre": "Herramientas", "subcuentas": {"33711": {"nombre": "Costo"}, "33712": {"nombre": "Revaluación"}}},
                "3372": {"nombre": "Unidades de reemplazo", "subcuentas": {"33721": {"nombre": "Costo"}, "33722": {"nombre": "Revaluación"}}}
            }}
        }},
        "34": {"nombre": "Intangibles", "subcuentas": {
            "341": {"nombre": "Concesiones, licencias y otros derechos", "subcuentas": {
                "3411": {"nombre": "Concesiones", "subcuentas": {"34111": {"nombre": "Costo"}, "34112": {"nombre": "Revaluación"}}},
                "3412": {"nombre": "Licencias", "subcuentas": {"34121": {"nombre": "Costo"}, "34122": {"nombre": "Revaluación"}}},
                "3413": {"nombre": "Otros derechos", "subcuentas": {"34131": {"nombre": "Costo"}, "34132": {"nombre": "Revaluación"}}}
            }},
            "342": {"nombre": "Patentes y propiedad industrial", "subcuentas": {
                "3421": {"nombre": "Patentes", "subcuentas": {"34211": {"nombre": "Costo"}, "34212": {"nombre": "Revaluación"}}},
                "3422": {"nombre": "Marcas", "subcuentas": {"34221": {"nombre": "Costo"}, "34222": {"nombre": "Revaluación"}}}
            }},
            "343": {"nombre": "Programas de computadora (software)", "subcuentas": {
                "3431": {"nombre": "Aplicaciones informáticas", "subcuentas": {"34311": {"nombre": "Costo"}, "34312": {"nombre": "Revaluación"}}},
                "3432": {"nombre": "Software a medida", "subcuentas": {"34321": {"nombre": "Costo"}, "34322": {"nombre": "Revaluación"}}}
            }},
            "344": {"nombre": "Costos de exploración y desarrollo", "subcuentas": {
                "3441": {"nombre": "Costos de exploración", "subcuentas": {"34411": {"nombre": "Costo"}, "34412": {"nombre": "Revaluación"}}},
                "3442": {"nombre": "Costos de desarrollo", "subcuentas": {"34421": {"nombre": "Costo"}, "34422": {"nombre": "Revaluación"}}}
            }},
            "345": {"nombre": "Fórmulas, diseños y prototipos", "subcuentas": {
                "3451": {"nombre": "Fórmulas", "subcuentas": {"34511": {"nombre": "Costo"}, "34512": {"nombre": "Revaluación"}}},
                "3452": {"nombre": "Diseños y prototipos", "subcuentas": {"34521": {"nombre": "Costo"}, "34522": {"nombre": "Revaluación"}}}
            }},
            "346": {"nombre": "Reservas de recursos extraíbles", "subcuentas": {
                "3461": {"nombre": "Minerales", "subcuentas": {"34611": {"nombre": "Costo"}, "34612": {"nombre": "Revaluación"}}},
                "3462": {"nombre": "Petróleo y gas", "subcuentas": {"34621": {"nombre": "Costo"}, "34622": {"nombre": "Revaluación"}}},
                "3463": {"nombre": "Madera", "subcuentas": {"34631": {"nombre": "Costo"}, "34632": {"nombre": "Revaluación"}}},
                "3469": {"nombre": "Otros recursos extraíbles", "subcuentas": {"34691": {"nombre": "Costo"}, "34692": {"nombre": "Revaluación"}}}
            }},
            "347": {"nombre": "Plusvalía mercantil", "subcuentas": {
                "3471": {"nombre": "Plusvalía mercantil", "subcuentas": {"34711": {"nombre": "Costo"}, "34712": {"nombre": "Revaluación"}}}
            }},
            "348": {"nombre": "Otros activos intangibles", "subcuentas": {
                "3481": {"nombre": "Derechos sobre propiedad intelectual", "subcuentas": {"34811": {"nombre": "Costo"}, "34812": {"nombre": "Revaluación"}}},
                "3482": {"nombre": "Otros activos intangibles", "subcuentas": {"34821": {"nombre": "Costo"}, "34822": {"nombre": "Revaluación"}}}
            }}
        }},
        "35": {"nombre": "Activos biológicos", "subcuentas": {
            "351": {"nombre": "Activos biológicos en producción", "subcuentas": {
                "3511": {"nombre": "De origen animal", "subcuentas": {
                    "35111": {"nombre": "Valor razonable"},
                    "35112": {"nombre": "Costo"},
                    "35113": {"nombre": "Costo de financiación"}
                }},
                "3512": {"nombre": "De origen vegetal", "subcuentas": {
                    "35121": {"nombre": "Valor razonable"},
                    "35122": {"nombre": "Costo"},
                    "35123": {"nombre": "Costo de financiación"}
                }}
            }},
            "352": {"nombre": "Activos biológicos en desarrollo", "subcuentas": {
                "3521": {"nombre": "De origen animal", "subcuentas": {
                    "35211": {"nombre": "Valor razonable"},
                    "35212": {"nombre": "Costo"},
                    "35213": {"nombre": "Costo de financiación"}
                }},
                "3522": {"nombre": "De origen vegetal", "subcuentas": {
                    "35221": {"nombre": "Valor razonable"},
                    "35222": {"nombre": "Costo"},
                    "35223": {"nombre": "Costo de financiación"}
                }}
            }}
        }},
        "36": {"nombre": "Desvalorización de activo inmovilizado", "subcuentas": {
            "361": {"nombre": "Desvalorización de inversiones inmobiliarias", "subcuentas": {
                "3611": {"nombre": "Terrenos"},
                "3612": {"nombre": "Edificaciones", "subcuentas": {
                    "36121": {"nombre": "Edificaciones – Costo de adquisición o construcción"},
                    "36122": {"nombre": "Edificaciones – Costo de financiación"}
                }}
            }},
            "363": {"nombre": "Desvalorización de inmuebles, maquinaria y equipo", "subcuentas": {
                "3631": {"nombre": "Terrenos"},
                "3632": {"nombre": "Edificaciones", "subcuentas": {
                    "36321": {"nombre": "Edificaciones – Costo de adquisición o construcción"},
                    "36322": {"nombre": "Edificaciones – Costo de financiación"}
                }},
                "3633": {"nombre": "Maquinarias y equipos de explotación", "subcuentas": {
                    "36331": {"nombre": "Maquinarias y equipos de explotación – Costo de adquisición o construcción"},
                    "36332": {"nombre": "Maquinarias y equipos de explotación – Costo de financiación"}
                }},
                "3634": {"nombre": "Equipo de transporte"},
                "3635": {"nombre": "Muebles y enseres"},
                "3636": {"nombre": "Equipos diversos"},
                "3637": {"nombre": "Herramientas y unidades de reemplazo"}
            }},
            "364": {"nombre": "Desvalorización de intangibles", "subcuentas": {
                "3641": {"nombre": "Concesiones, licencias y otros derechos"},
                "3642": {"nombre": "Patentes y propiedad industrial"},
                "3643": {"nombre": "Programas de computadora (software)"},
                "3644": {"nombre": "Costos de exploración y desarrollo", "subcuentas": {
                    "36441": {"nombre": "Costo"},
                    "36442": {"nombre": "Costo de financiación"}
                }},
                "3645": {"nombre": "Fórmulas, diseños y prototipos"},
                "3647": {"nombre": "Plusvalía mercantil"},
                "3649": {"nombre": "Otros activos intangibles"}
            }},
            "365": {"nombre": "Desvalorización de activos biológicos", "subcuentas": {
                "3651": {"nombre": "Activos biológicos en producción", "subcuentas": {
                    "36511": {"nombre": "Costo"},
                    "36512": {"nombre": "Costo de financiación"}
                }},
                "3652": {"nombre": "Activos biológicos en desarrollo", "subcuentas": {
                    "36521": {"nombre": "Costo"},
                    "36522": {"nombre": "Costo de financiación"}
                }}
            }},
            "366": {"nombre": "Desvalorización de inversiones mobiliarias", "subcuentas": {
                "3661": {"nombre": "Inversiones a ser mantenidas hasta el vencimiento"},
                "3662": {"nombre": "Inversiones financieras representativas de derecho patrimonial"}
            }}
        }},
        "37": {"nombre": "Activos diferidos", "subcuentas": {
            "371": {"nombre": "Impuesto a la renta diferido", "subcuentas": {
                "3711": {"nombre": "Impuesto a la renta diferido – Patrimonio"},
                "3712": {"nombre": "Impuesto a la renta diferido – Resultados"}
            }},
            "372": {"nombre": "Aportes diferidos", "subcuentas": {
                "3721": {"nombre": "Aportes a pensiones"},
                "3722": {"nombre": "Aportes a salud"}
            }},
            "373": {"nombre": "Otros activos diferidos", "subcuentas": {
                "3731": {"nombre": "Intereses diferidos"},
                "3732": {"nombre": "Costos diferidos"}
            }}
        }},
        "38": {"nombre": "Otros activos", "subcuentas": {
            "381": {"nombre": "Bienes de arte y cultura", "subcuentas": {
                "3811": {"nombre": "Obras de arte"},
                "3812": {"nombre": "Biblioteca"},
                "3813": {"nombre": "Otros"}
            }},
            "382": {"nombre": "Diversos", "subcuentas": {
                "3821": {"nombre": "Monedas y joyas"},
                "3822": {"nombre": "Bienes entregados en comodato"},
                "3823": {"nombre": "Bienes recibidos en pago (adjudicados y realizables)"},
                "3829": {"nombre": "Otros"}
            }},    
        }},
       "39": {"nombre": "Depreciación, amortización y agotamiento acumulados", "subcuentas": {
            "391": {"nombre": "Depreciación acumulada", "subcuentas": {
                "3911": {"nombre": "Inversiones inmobiliarias", "subcuentas": {
                    "39111": {"nombre": "Edificaciones – Costo de adquisición o construcción"},
                    "39112": {"nombre": "Edificaciones – Revaluación"},
                    "39113": {"nombre": "Edificaciones – Costo de financiación"}
                }},
                "3912": {"nombre": "Activos adquiridos en arrendamiento financiero", "subcuentas": {
                    "39121": {"nombre": "Inversiones inmobiliarias – Edificaciones"},
                    "39122": {"nombre": "Inmuebles, maquinaria y equipo – Edificaciones"},
                    "39123": {"nombre": "Inmuebles, maquinaria y equipo – Maquinarias y equipos de explotación"},
                    "39124": {"nombre": "Inmuebles, maquinaria y equipo – Equipos de transporte"},
                    "39126": {"nombre": "Inmuebles, maquinaria y equipo – Equipos diversos"}
                }},
                "3913": {"nombre": "Inmuebles, maquinaria y equipo – Costo", "subcuentas": {
                    "39131": {"nombre": "Edificaciones"},
                    "39132": {"nombre": "Maquinarias y equipos de explotación"},
                    "39133": {"nombre": "Equipo de transporte"},
                    "39134": {"nombre": "Muebles y enseres"},
                    "39135": {"nombre": "Equipos diversos"},
                    "39136": {"nombre": "Herramientas y unidades de reemplazo"}
                }},
                "3914": {"nombre": "Inmuebles, maquinaria y equipo – Revaluación", "subcuentas": {
                    "39141": {"nombre": "Edificaciones"},
                    "39142": {"nombre": "Maquinarias y equipos de explotación"},
                    "39143": {"nombre": "Equipo de transporte"},
                    "39144": {"nombre": "Muebles y enseres"},
                    "39145": {"nombre": "Equipos diversos"},
                    "39146": {"nombre": "Herramientas y unidades de reemplazo"}
                }},
                "3915": {"nombre": "Inmuebles, maquinaria y equipo – Costo de financiación", "subcuentas": {
                    "39151": {"nombre": "Edificaciones"},
                    "39152": {"nombre": "Maquinarias y equipos de explotación"}
                }},
                "3916": {"nombre": "Activos biológicos en producción – Costo", "subcuentas": {
                    "39161": {"nombre": "Activos biológicos de origen animal"},
                    "39162": {"nombre": "Activos biológicos de origen vegetal"}
                }},
                "3917": {"nombre": "Activos biológicos en producción – Costo de financiación", "subcuentas": {
                    "39171": {"nombre": "Activos biológicos de origen animal"},
                    "39172": {"nombre": "Activos biológicos de origen vegetal"}
                }}
            }},
            "392": {"nombre": "Amortización acumulada", "subcuentas": {
                "3921": {"nombre": "Intangibles – Costo", "subcuentas": {
                    "39211": {"nombre": "Concesiones, licencias y otros derechos"},
                    "39212": {"nombre": "Patentes y propiedad industrial"},
                    "39213": {"nombre": "Programas de computadora (software)"},
                    "39214": {"nombre": "Costos de exploración y desarrollo"},
                    "39215": {"nombre": "Fórmulas, diseños y prototipos"},
                    "39219": {"nombre": "Otros activos intangibles"}
                }},
                "3922": {"nombre": "Intangibles – Revaluación", "subcuentas": {
                    "39221": {"nombre": "Concesiones, licencias y otros derechos"},
                    "39222": {"nombre": "Patentes y propiedad industrial"},
                    "39223": {"nombre": "Programas de computadora (software)"},
                    "39224": {"nombre": "Costos de exploración y desarrollo"},
                    "39225": {"nombre": "Fórmulas, diseños y prototipos"},
                    "39229": {"nombre": "Otros activos intangibles"}
                }},
                "3923": {"nombre": "Intangibles – Costos de financiación", "subcuentas": {
                    "39234": {"nombre": "Costos de exploración y desarrollo"}
                }}
            }},
            "393": {"nombre": "Agotamiento acumulado", "subcuentas": {
                "3931": {"nombre": "Agotamiento de reservas de recursos extraíbles"}
            }}
        }},
        "40": {"nombre": "Tributos, contrapartes y aportes al sistema de pensiones y de salud por pagar", "subcuentas": {
            "401": {"nombre": "Gobierno central", "subcuentas": {
                "4011": {"nombre": "Impuesto general a las ventas", "subcuentas": {
                    "40111": {"nombre": "IGV – Cuenta propia"},
                    "40112": {"nombre": "IGV – Servicios prestados por no domiciliados"},
                    "40113": {"nombre": "IGV – Régimen de percepciones"},
                    "40114": {"nombre": "IGV – Régimen de retenciones"}
                }},
                "4012": {"nombre": "Impuesto selectivo al consumo"},
                "4015": {"nombre": "Derechos aduaneros", "subcuentas": {
                    "40151": {"nombre": "Derechos arancelarios"},
                    "40152": {"nombre": "Derechos aduaneros por ventas"}
                }},
                "4017": {"nombre": "Impuesto a la renta", "subcuentas": {
                    "40171": {"nombre": "Renta de tercera categoría"},
                    "40172": {"nombre": "Renta de cuarta categoría"},
                    "40173": {"nombre": "Renta de quinta categoría"},
                    "40174": {"nombre": "Renta de no domiciliados"},
                    "40175": {"nombre": "Otras retenciones"}
                }},
                "4018": {"nombre": "Otros impuestos y contraprestaciones", "subcuentas": {
                    "40181": {"nombre": "Impuesto a las transacciones financieras"},
                    "40182": {"nombre": "Impuesto a los juegos de casino y tragamonedas"},
                    "40183": {"nombre": "Tasas por la prestación de servicios públicos"},
                    "40184": {"nombre": "Regalías"},
                    "40185": {"nombre": "Impuesto a los dividendos"},
                    "40186": {"nombre": "Impuesto temporal a los activos netos"},
                    "40189": {"nombre": "Otros impuestos"}
                }},
            "402": {"nombre": "Certificados tributarios"},
            "403": {"nombre": "Instituciones públicas", "subcuentas": {
                    "4031": {"nombre": "ESSALUD"},
                    "4032": {"nombre": "ONP"},
                    "4033": {"nombre": "Contribución al SENATI"},
                    "4034": {"nombre": "Contribución al SENCICO"},
                    "4039": {"nombre": "Otras instituciones"}
                }},
            "405": {"nombre": "Gobiernos regionales"},
            "406": {"nombre": "Gobiernos locales", "subcuentas": {
                    "4061": {"nombre": "Impuestos", "subcuentas": {
                        "40611": {"nombre": "Impuesto al patrimonio vehicular"},
                        "40612": {"nombre": "Impuesto a las apuestas"},
                        "40613": {"nombre": "Impuesto a los juegos"},
                        "40614": {"nombre": "Impuesto de alcabala"},
                        "40615": {"nombre": "Impuesto predial"},
                        "40616": {"nombre": "Impuesto a los espectáculos públicos no deportivos"}
                    }},
                    "4062": {"nombre": "Contribuciones"},
                    "4063": {"nombre": "Tasas", "subcuentas": {
                        "40631": {"nombre": "Licencia de apertura de establecimientos"},
                        "40632": {"nombre": "Transporte público"},
                        "40633": {"nombre": "Estacionamiento de vehículos"},
                        "40634": {"nombre": "Servicios públicos o arbitrios"},
                        "40635": {"nombre": "Servicios administrativos o derechos"}
                    }}
                }},
            "407": {"nombre": "Administradoras de fondos de pensiones"},
            "408": {"nombre": "Empresas prestadoras de servicios de salud", "subcuentas": {
                    "4081": {"nombre": "Cuenta propia"},
                    "4082": {"nombre": "Cuenta de terceros"}
                }},
            "409": {"nombre": "Otros costos administrativos e intereses"}
            }}
        }},
        "41": {"nombre": "Remuneraciones y participaciones por pagar", "subcuentas": {
            "411": {"nombre": "Remuneraciones por pagar", "subcuentas": {
                "4111": {"nombre": "Sueldos y salarios por pagar"},
                "4112": {"nombre": "Comisiones por pagar"},
                "4113": {"nombre": "Remuneraciones en especie por pagar"},
                "4114": {"nombre": "Gratificaciones por pagar"},
                "4115": {"nombre": "Vacaciones por pagar"}
            }},
            "413": {"nombre": "Participaciones de los trabajadores por pagar"},
            "415": {"nombre": "Beneficios sociales de los trabajadores por pagar", "subcuentas": {
                "4151": {"nombre": "Compensación por tiempo de servicios"},
                "4152": {"nombre": "Adelanto de compensación por tiempo de servicios"},
                "4153": {"nombre": "Pensiones y jubilaciones"}
            }},
            "419": {"nombre": "Otras remuneraciones y participaciones por pagar"}
        }},
        "42": {"nombre": "Cuentas por pagar comerciales – Terceros", "subcuentas": {
            "421": {"nombre": "Facturas, boletas y otros comprobantes por pagar", "subcuentas": {
                "4211": {"nombre": "No emitidas"},
                "4212": {"nombre": "Emitidas"}
            }},
            "422": {"nombre": "Anticipos a proveedores"},
            "423": {"nombre": "Letras por pagar"},
            "424": {"nombre": "Honorarios por pagar"}
        }},
        "43": {"nombre": "Cuentas por pagar comerciales – Relacionadas", "subcuentas": {
            "431": {"nombre": "Facturas, boletas y otros comprobantes por pagar", "subcuentas": {
                "4311": {"nombre": "No emitidas", "subcuentas": {
                    "43111": {"nombre": "Matriz"},
                    "43112": {"nombre": "Subsidiarias"},
                    "43113": {"nombre": "Asociadas"},
                    "43114": {"nombre": "Sucursales"},
                    "43115": {"nombre": "Otros"}
                }},
                "4312": {"nombre": "Emitidas", "subcuentas": {
                    "43121": {"nombre": "Matriz"},
                    "43122": {"nombre": "Subsidiarias"},
                    "43123": {"nombre": "Asociadas"},
                    "43124": {"nombre": "Sucursales"},
                    "43125": {"nombre": "Otros"}
                }}
            }},
            "432": {"nombre": "Anticipos otorgados", "subcuentas": {
                "4321": {"nombre": "Anticipos otorgados", "subcuentas": {
                    "43211": {"nombre": "Matriz"},
                    "43212": {"nombre": "Subsidiarias"},
                    "43213": {"nombre": "Asociadas"},
                    "43214": {"nombre": "Sucursales"},
                    "43215": {"nombre": "Otros"}
                }}
            }},
            "433": {"nombre": "Letras por pagar", "subcuentas": {
                "4331": {"nombre": "Letras por pagar", "subcuentas": {
                    "43311": {"nombre": "Matriz"},
                    "43312": {"nombre": "Subsidiarias"},
                    "43313": {"nombre": "Asociadas"},
                    "43314": {"nombre": "Sucursales"},
                    "43315": {"nombre": "Otros"}
                }}
            }},
            "434": {"nombre": "Honorarios por pagar", "subcuentas": {
                "4341": {"nombre": "Honorarios por pagar", "subcuentas": {
                    "43411": {"nombre": "Matriz"},
                    "43412": {"nombre": "Subsidiarias"},
                    "43413": {"nombre": "Asociadas"},
                    "43414": {"nombre": "Sucursales"},
                    "43415": {"nombre": "Otros"}
                }}
            }}
        }},
        "44": {"nombre": "Cuentas por pagar a los accionistas (socios), directores y gerentes", "subcuentas": {
            "441": {"nombre": "Accionistas (o socios)", "subcuentas": {
                "4411": {"nombre": "Préstamos"},
                "4412": {"nombre": "Dividendos"},
                "4419": {"nombre": "Otras cuentas por pagar"}
            }},
            "442": {"nombre": "Directores", "subcuentas": {
                "4421": {"nombre": "Dietas"},
                "4429": {"nombre": "Otras cuentas por pagar"}
            }},
            "443": {"nombre": "Gerentes"}
        }},
        "45": {"nombre": "Obligaciones financieras", "subcuentas": {
            "451": {"nombre": "Préstamos de instituciones financieras y otras entidades", "subcuentas": {
                "4511": {"nombre": "Instituciones financieras"},
                "4512": {"nombre": "Otras entidades"}
            }},
            "452": {"nombre": "Contratos de arrendamiento financiero"},
            "453": {"nombre": "Obligaciones emitidas", "subcuentas": {
                "4531": {"nombre": "Bonos emitidos"},
                "4532": {"nombre": "Bonos titulizados"},
                "4533": {"nombre": "Papeles comerciales"},
                "4539": {"nombre": "Otras obligaciones"}
            }},
            "454": {"nombre": "Otros instrumentos financieros por pagar", "subcuentas": {
                "4541": {"nombre": "Letras"},
                "4542": {"nombre": "Papeles comerciales"},
                "4543": {"nombre": "Bonos"},
                "4544": {"nombre": "Pagarés"},
                "4545": {"nombre": "Facturas conformadas"},
                "4549": {"nombre": "Otras obligaciones financieras"}
            }},
            "455": {"nombre": "Costos de financiación por pagar", "subcuentas": {
                "4551": {"nombre": "Préstamos de instituciones financieras y otras entidades", "subcuentas": {
                    "45511": {"nombre": "Instituciones financieras"},
                    "45512": {"nombre": "Otras entidades"}
                }},
                "4552": {"nombre": "Contratos de arrendamiento financiero"},
                "4553": {"nombre": "Obligaciones emitidas", "subcuentas": {
                    "45531": {"nombre": "Bonos emitidos"},
                    "45532": {"nombre": "Bonos titulizados"},
                    "45533": {"nombre": "Papeles comerciales"},
                    "45539": {"nombre": "Otras obligaciones"}
                }},
                "4554": {"nombre": "Otros instrumentos financieros por pagar", "subcuentas": {
                    "45541": {"nombre": "Letras"},
                    "45542": {"nombre": "Papeles comerciales"},
                    "45543": {"nombre": "Bonos"},
                    "45544": {"nombre": "Pagarés"},
                    "45545": {"nombre": "Facturas conformadas"},
                    "45549": {"nombre": "Otras obligaciones financieras"}
                }}
            }},
            "456": {"nombre": "Préstamos con compromisos de recompra"}
        }},
        "46": {"nombre": "Cuentas por pagar diversas – Terceros", "subcuentas": {
            "461": {"nombre": "Reclamaciones de terceros"},
            "464": {"nombre": "Pasivos por instrumentos financieros", "subcuentas": {
                "4641": {"nombre": "Instrumentos financieros primarios"},
                "4642": {"nombre": "Instrumentos financieros derivados", "subcuentas": {
                    "46421": {"nombre": "Cartera de negociación"},
                    "46422": {"nombre": "Instrumentos de cobertura"}
                }}
            }},
            "465": {"nombre": "Pasivos por compra de activo inmovilizado", "subcuentas": {
                "4651": {"nombre": "Inversiones mobiliarias"},
                "4652": {"nombre": "Inversiones inmobiliarias"},
                "4653": {"nombre": "Activos adquiridos en arrendamiento financiero"},
                "4654": {"nombre": "Inmuebles, maquinaria y equipo"},
                "4655": {"nombre": "Intangibles"},
                "4656": {"nombre": "Activos biológicos"}
            }},
            "467": {"nombre": "Depósitos recibidos en garantía"},
            "469": {"nombre": "Otras cuentas por pagar diversas", "subcuentas": {
                "4691": {"nombre": "Subsidios gubernamentales"},
                "4692": {"nombre": "Donaciones condicionadas"},
                "4699": {"nombre": "Otras cuentas por pagar"}
            }}
        }},
        "47": {"nombre": "Cuentas por pagar diversas – Relacionadas", "subcuentas": {
            "471": {"nombre": "Préstamos", "subcuentas": {
                "4711": {"nombre": "Matriz"},
                "4712": {"nombre": "Subsidiarias"},
                "4713": {"nombre": "Asociadas"},
                "4714": {"nombre": "Sucursales"},
                "4715": {"nombre": "Otras"}
            }},
            "472": {"nombre": "Costos de financiación", "subcuentas": {
                "4721": {"nombre": "Matriz"},
                "4722": {"nombre": "Subsidiarias"},
                "4723": {"nombre": "Asociadas"},
                "4724": {"nombre": "Sucursales"},
                "4725": {"nombre": "Otras"}
            }},
            "473": {"nombre": "Anticipos recibidos", "subcuentas": {
                "4731": {"nombre": "Matriz"},
                "4732": {"nombre": "Subsidiarias"},
                "4733": {"nombre": "Asociadas"},
                "4734": {"nombre": "Sucursales"},
                "4735": {"nombre": "Otras"}
            }},
            "474": {"nombre": "Regalías", "subcuentas": {
                "4741": {"nombre": "Matriz"},
                "4742": {"nombre": "Subsidiarias"},
                "4743": {"nombre": "Asociadas"},
                "4744": {"nombre": "Sucursales"},
                "4745": {"nombre": "Otras"}
            }},
            "475": {"nombre": "Dividendos", "subcuentas": {
                "4751": {"nombre": "Matriz"},
                "4752": {"nombre": "Subsidiarias"},
                "4753": {"nombre": "Asociadas"},
                "4754": {"nombre": "Sucursales"},
                "4755": {"nombre": "Otras"}
            }},
            "477": {"nombre": "Pasivo por compra de activo inmovilizado", "subcuentas": {
                "4771": {"nombre": "Inversiones mobiliarias", "subcuentas": {
                    "47711": {"nombre": "Matriz"},
                    "47712": {"nombre": "Subsidiarias"},
                    "47713": {"nombre": "Asociadas"},
                    "47714": {"nombre": "Sucursales"},
                    "47715": {"nombre": "Otras"}
                }},
                "4772": {"nombre": "Inversiones inmobiliarias", "subcuentas": {
                    "47721": {"nombre": "Matriz"},
                    "47722": {"nombre": "Subsidiarias"},
                    "47723": {"nombre": "Asociadas"},
                    "47724": {"nombre": "Sucursales"},
                    "47725": {"nombre": "Otras"}
                }},
                "4773": {"nombre": "Activos adquiridos en arrendamiento financiero", "subcuentas": {
                    "47731": {"nombre": "Matriz"},
                    "47732": {"nombre": "Subsidiarias"},
                    "47733": {"nombre": "Asociadas"},
                    "47734": {"nombre": "Sucursales"},
                    "47735": {"nombre": "Otras"}
                }},
                "4774": {"nombre": "Inmuebles, maquinaria y equipo", "subcuentas": {
                    "47741": {"nombre": "Matriz"},
                    "47742": {"nombre": "Subsidiarias"},
                    "47743": {"nombre": "Asociadas"},
                    "47744": {"nombre": "Sucursales"},
                    "47745": {"nombre": "Otras"}
                }},
                "4775": {"nombre": "Intangibles", "subcuentas": {
                    "47751": {"nombre": "Matriz"},
                    "47752": {"nombre": "Subsidiarias"},
                    "47753": {"nombre": "Asociadas"},
                    "47754": {"nombre": "Sucursales"},
                    "47755": {"nombre": "Otras"}
                }},
                "4776": {"nombre": "Activos biológicos", "subcuentas": {
                    "47761": {"nombre": "Matriz"},
                    "47762": {"nombre": "Subsidiarias"},
                    "47763": {"nombre": "Asociadas"},
                    "47764": {"nombre": "Sucursales"},
                    "47765": {"nombre": "Otras"}
                }}
            }},
            "479": {"nombre": "Otras cuentas por pagar diversas", "subcuentas": {
                "4791": {"nombre": "Otras cuentas por pagar diversas", "subcuentas": {
                    "47911": {"nombre": "Matriz"},
                    "47912": {"nombre": "Subsidiarias"},
                    "47913": {"nombre": "Asociadas"},
                    "47914": {"nombre": "Sucursales"},
                    "47915": {"nombre": "Otras"}
                }}
            }}
        }},
        "48": {"nombre": "Provisiones", "subcuentas": {
            "481": {"nombre": "Provisión para litigios"},
            "482": {"nombre": "Provisión por desmantelamiento, retiro o rehabilitación del inmovilizado"},
            "483": {"nombre": "Provisión para reestructuraciones"},
            "484": {"nombre": "Provisión para protección y remediación del medio ambiente"},
            "485": {"nombre": "Provisión para gastos de responsabilidad social"},
            "486": {"nombre": "Provisión para garantías"},
            "489": {"nombre": "Otras provisiones"}
        }},
        "49": {"nombre": "Pasivo diferido", "subcuentas": {
            "491": {"nombre": "Impuesto a la renta diferido", "subcuentas": {
                "4911": {"nombre": "Impuesto a la renta diferido – Patrimonio"},
                "4912": {"nombre": "Impuesto a la renta diferido – Resultados"}
            }},
            "492": {"nombre": "Participaciones de los trabajadores diferidas", "subcuentas": {
                "4921": {"nombre": "Participaciones de los trabajadores diferidas – Patrimonio"},
                "4922": {"nombre": "Participaciones de los trabajadores diferidas – Resultados"}
            }},
            "493": {"nombre": "Intereses diferidos", "subcuentas": {
                "4931": {"nombre": "Intereses no devengados en transacciones con terceros"},
                "4932": {"nombre": "Intereses no devengados en medición a valor descontado"}
            }},
            "494": {"nombre": "Ganancia en venta con arrendamiento financiero paralelo"},
            "495": {"nombre": "Subsidios recibidos diferidos"},
            "496": {"nombre": "Ingresos diferidos"}
        }},
        "50": {"nombre": "Capital", "subcuentas": {
            "501": {"nombre": "Capital social", "subcuentas": {
                "5011": {"nombre": "Acciones"},
                "5012": {"nombre": "Participaciones"}
            }},
            "502": {"nombre": "Acciones en tesorería"}
        }},
        "51": {"nombre": "Acciones de inversión", "subcuentas": {
            "511": {"nombre": "Acciones de inversión"},
            "512": {"nombre": "Acciones de inversión en tesorería"}
        }},
        "52": {"nombre": "Capital adicional", "subcuentas": {
            "521": {"nombre": "Primas (descuento) de acciones"},
            "522": {"nombre": "Capitalizaciones en trámite", "subcuentas": {
                "5221": {"nombre": "Aportes"},
                "5222": {"nombre": "Reservas"},
                "5223": {"nombre": "Acreencias"},
                "5224": {"nombre": "Utilidades"}
            }},
            "523": {"nombre": "Reducciones de capital pendientes de formalización"}
        }},
        "56": {"nombre": "Resultados no realizados", "subcuentas": {
            "561": {"nombre": "Diferencia en cambio de inversiones permanentes en entidades extranjeras"},
            "562": {"nombre": "Instrumentos financieros – Cobertura de flujo de efectivo"},
            "563": {"nombre": "Ganancia o pérdida en activos o pasivos financieros disponibles para la venta", "subcuentas": {
                "5631": {"nombre": "Ganancia"},
                "5632": {"nombre": "Pérdida"}
            }},
            "564": {"nombre": "Ganancia o pérdida en activos o pasivos financieros disponibles para la venta – Compra o venta convencional fecha de liquidación", "subcuentas": {
                "5641": {"nombre": "Ganancia"},
                "5642": {"nombre": "Pérdida"}
            }}
        }},
        "57": {"nombre": "Excedente de revaluación", "subcuentas": {
            "571": {"nombre": "Excedente de revaluación", "subcuentas": {
                "5711": {"nombre": "Inversiones inmobiliarias"},
                "5712": {"nombre": "Inmuebles, maquinaria y equipos"},
                "5713": {"nombre": "Intangibles"}
            }},
            "572": {"nombre": "Excedente de revaluación – Acciones liberadas recibidas"},
            "573": {"nombre": "Participación en excedente de revaluación – Inversiones en entidades relacionadas"}
        }},
        "58": {"nombre": "Reservas", "subcuentas": {
            "581": {"nombre": "Reinversión"},
            "582": {"nombre": "Legal"},
            "583": {"nombre": "Contractuales"},
            "584": {"nombre": "Estatutarias"},
            "585": {"nombre": "Facultativas"},
            "589": {"nombre": "Otras reservas"}
        }},
        "59": {"nombre": "Resultados acumulados", "subcuentas": {
            "591": {"nombre": "Utilidades no distribuidas", "subcuentas": {
                "5911": {"nombre": "Utilidades acumuladas"},
                "5912": {"nombre": "Ingresos de años anteriores"}
            }},
            "592": {"nombre": "Pérdidas acumuladas", "subcuentas": {
                "5921": {"nombre": "Pérdidas acumuladas"},
                "5922": {"nombre": "Gastos de años anteriores"}
            }}
        }},
        "60": {"nombre": "Compras", "subcuentas": {
            "601": {"nombre": "Mercaderías", "subcuentas": {
                "6011": {"nombre": "Mercaderías manufacturadas"},
                "6012": {"nombre": "Mercaderías de extracción"},
                "6013": {"nombre": "Mercaderías agropecuarias y piscícolas"},
                "6014": {"nombre": "Mercaderías inmuebles"},
                "6018": {"nombre": "Otras mercaderías"}
            }},
            "602": {"nombre": "Materias primas", "subcuentas": {
                "6021": {"nombre": "Materias primas para productos manufacturados"},
                "6022": {"nombre": "Materias primas para productos de extracción"},
                "6023": {"nombre": "Materias primas para productos agropecuarios y piscícolas"},
                "6024": {"nombre": "Materias primas para productos inmuebles"}
            }},
            "603": {"nombre": "Materiales auxiliares, suministros y repuestos", "subcuentas": {
                "6031": {"nombre": "Materiales auxiliares"},
                "6032": {"nombre": "Suministros"},
                "6033": {"nombre": "Repuestos"}
            }},
            "604": {"nombre": "Envases y embalajes", "subcuentas": {
                "6041": {"nombre": "Envases"},
                "6042": {"nombre": "Embalajes"}
            }},
            "609": {"nombre": "Costos vinculados con las compras", "subcuentas": {
                "6091": {"nombre": "Costos vinculados con las compras de mercaderías", "subcuentas": {
                    "60911": {"nombre": "Transporte"},
                    "60912": {"nombre": "Seguros"},
                    "60913": {"nombre": "Derechos aduaneros"},
                    "60914": {"nombre": "Comisiones"},
                    "60919": {"nombre": "Otros costos vinculados con las compras de mercaderías"}
                }},
                "6092": {"nombre": "Costos vinculados con las compras de materias primas", "subcuentas": {
                    "60921": {"nombre": "Transporte"},
                    "60922": {"nombre": "Seguros"},
                    "60923": {"nombre": "Derechos aduaneros"},
                    "60924": {"nombre": "Comisiones"},
                    "60925": {"nombre": "Otros costos vinculados con las compras de materias primas"}
                }},
                "6093": {"nombre": "Costos vinculados con las compras de materiales, suministros y repuestos", "subcuentas": {
                    "60931": {"nombre": "Transporte"},
                    "60932": {"nombre": "Seguros"},
                    "60933": {"nombre": "Derechos aduaneros"},
                    "60934": {"nombre": "Comisiones"},
                    "60935": {"nombre": "Otros costos vinculados con las compras de materiales, suministros y repuestos"}
                }},
                "6094": {"nombre": "Costos vinculados con las compras de envases y embalajes", "subcuentas": {
                    "60941": {"nombre": "Transporte"},
                    "60942": {"nombre": "Seguros"},
                    "60943": {"nombre": "Derechos aduaneros"},
                    "60944": {"nombre": "Comisiones"},
                    "60945": {"nombre": "Otros costos vinculados con las compras de envases y embalajes"}
                }}
            }}
        }},
        "61": {"nombre": "Variación de existencias", "subcuentas": {
            "611": {"nombre": "Mercaderías", "subcuentas": {
                "6111": {"nombre": "Mercaderías manufacturadas"},
                "6112": {"nombre": "Mercaderías de extracción"},
                "6113": {"nombre": "Mercaderías agropecuarias y piscícolas"},
                "6114": {"nombre": "Mercaderías inmuebles"},
                "6115": {"nombre": "Otras mercaderías"}
            }},
            "612": {"nombre": "Materias primas", "subcuentas": {
                "6121": {"nombre": "Materias primas para productos manufacturados"},
                "6122": {"nombre": "Materias primas para productos de extracción"},
                "6123": {"nombre": "Materias primas para productos agropecuarios y piscícolas"},
                "6124": {"nombre": "Materias primas para productos inmuebles"}
            }},
            "613": {"nombre": "Materiales auxiliares, suministros y repuestos", "subcuentas": {
                "6131": {"nombre": "Materiales auxiliares"},
                "6132": {"nombre": "Suministros"},
                "6133": {"nombre": "Repuestos"}
            }},
            "614": {"nombre": "Envases y embalajes", "subcuentas": {
                "6141": {"nombre": "Envases"},
                "6142": {"nombre": "Embalajes"}
            }}
        }},
        "62": {"nombre": "Gastos de personal, directores y gerentes", "subcuentas": {
            "621": {"nombre": "Remuneraciones", "subcuentas": {
                "6211": {"nombre": "Sueldos y salarios"},
                "6212": {"nombre": "Comisiones"},
                "6213": {"nombre": "Remuneraciones en especie"},
                "6214": {"nombre": "Gratificaciones"},
                "6215": {"nombre": "Vacaciones"}
            }},
            "622": {"nombre": "Otras remuneraciones"},
            "623": {"nombre": "Indemnizaciones al personal"},
            "624": {"nombre": "Capacitación"},
            "625": {"nombre": "Atención al personal"},
            "626": {"nombre": "Gerentes"},
            "627": {"nombre": "Seguridad, previsión social y otras contribuciones", "subcuentas": {
                "6271": {"nombre": "Régimen de prestaciones de salud"},
                "6272": {"nombre": "Régimen de pensiones"},
                "6273": {"nombre": "Seguro complementario de trabajo de riesgo, accidentes de trabajo y enfermedades profesionales"},
                "6274": {"nombre": "Seguro de vida"},
                "6275": {"nombre": "Seguros particulares de prestaciones de salud – EPS y otros particulares"},
                "6276": {"nombre": "Caja de beneficios de seguridad social del pescador"},
                "6277": {"nombre": "Contribuciones al SENCICO y el SENATI"}
            }},
            "628": {"nombre": "Retribuciones al directorio"},
            "629": {"nombre": "Beneficios sociales de los trabajadores", "subcuentas": {
                "6291": {"nombre": "Compensación por tiempo de servicio"},
                "6292": {"nombre": "Pensiones y jubilaciones"},
                "6293": {"nombre": "Otros beneficios post-empleo"}
            }}
        }},
        "63": {"nombre": "Gastos de servicios prestados por terceros", "subcuentas": {
            "631": {"nombre": "Transporte, correos y gastos de viaje", "subcuentas": {
                "6311": {"nombre": "Transporte", "subcuentas": {
                    "63111": {"nombre": "De carga"},
                    "63112": {"nombre": "De pasajeros"}
                }},
                "6312": {"nombre": "Correos"},
                "6313": {"nombre": "Alojamiento"},
                "6314": {"nombre": "Alimentación"},
                "6315": {"nombre": "Otros gastos de viaje"}
            }},
            "632": {"nombre": "Asesoría y consultoría", "subcuentas": {
                "6321": {"nombre": "Administrativa"},
                "6322": {"nombre": "Legal y tributaria"},
                "6323": {"nombre": "Auditoría y contable"},
                "6324": {"nombre": "Mercadotecnia"},
                "6325": {"nombre": "Medioambiental"},
                "6326": {"nombre": "Investigación y desarrollo"},
                "6327": {"nombre": "Producción"},
                "6329": {"nombre": "Otros"}
            }},
            "633": {"nombre": "Producción encargada a terceros"},
            "634": {"nombre": "Mantenimiento y reparaciones", "subcuentas": {
                "6341": {"nombre": "Inversión inmobiliaria"},
                "6342": {"nombre": "Activos adquiridos en arrendamiento financiero"},
                "6343": {"nombre": "Inmuebles, maquinaria y equipo"},
                "6344": {"nombre": "Intangibles"},
                "6345": {"nombre": "Activos biológicos"}
            }},
            "635": {"nombre": "Alquileres", "subcuentas": {
                "6351": {"nombre": "Terrenos"},
                "6352": {"nombre": "Edificaciones"},
                "6353": {"nombre": "Maquinarias y equipos de explotación"},
                "6354": {"nombre": "Equipo de transporte"},
                "6356": {"nombre": "Equipos diversos"}
            }},
            "636": {"nombre": "Servicios básicos", "subcuentas": {
                "6361": {"nombre": "Energía eléctrica"},
                "6362": {"nombre": "Gas"},
                "6363": {"nombre": "Agua"},
                "6364": {"nombre": "Teléfono"},
                "6365": {"nombre": "Internet"},
                "6366": {"nombre": "Radio"},
                "6367": {"nombre": "Cable"}
            }},
            "637": {"nombre": "Publicidad, publicaciones, relaciones públicas", "subcuentas": {
                "6371": {"nombre": "Publicidad"},
                "6372": {"nombre": "Publicaciones"},
                "6373": {"nombre": "Relaciones públicas"}
            }},
            "638": {"nombre": "Servicios de contratistas"},
            "639": {"nombre": "Otros servicios prestados por terceros", "subcuentas": {
                "6391": {"nombre": "Gastos bancarios"},
                "6392": {"nombre": "Gastos de laboratorio"}
            }}
        }},
        "64": {"nombre": "Gastos por tributos", "subcuentas": {
            "641": {"nombre": "Gobierno central", "subcuentas": {
                "6411": {"nombre": "Impuesto general a las ventas y selectivo al consumo"},
                "6412": {"nombre": "Impuesto a las transacciones financieras"},
                "6413": {"nombre": "Impuesto temporal a los activos netos"},
                "6414": {"nombre": "Impuesto a los juegos de casino y máquinas tragamonedas"},
                "6415": {"nombre": "Regalías mineras"},
                "6416": {"nombre": "Cánones"},
                "6419": {"nombre": "Otros"}
            }},
            "642": {"nombre": "Gobierno regional"},
            "643": {"nombre": "Gobierno local", "subcuentas": {
                "6431": {"nombre": "Impuesto predial"},
                "6432": {"nombre": "Arbitrios municipales y seguridad ciudadana"},
                "6433": {"nombre": "Impuesto al patrimonio vehicular"},
                "6434": {"nombre": "Licencia de funcionamiento"},
                "6439": {"nombre": "Otros"}
            }},
            "644": {"nombre": "Otros gastos por tributos", "subcuentas": {
                "6441": {"nombre": "Contribución al SENATI"},
                "6442": {"nombre": "Contribución al SENCICO"},
                "6443": {"nombre": "Otros"}
            }}
        }},
        "65": {"nombre": "Otros gastos de gestión", "subcuentas": {
            "651": {"nombre": "Seguros"},
            "652": {"nombre": "Regalías"},
            "653": {"nombre": "Suscripciones"},
            "654": {"nombre": "Licencias y derechos de vigencia"},
            "655": {"nombre": "Costo neto de enajenación de activos inmovilizados y operaciones discontinuadas", "subcuentas": {
                "6551": {"nombre": "Costo neto de enajenación de activos inmovilizados", "subcuentas": {
                    "65511": {"nombre": "Inversiones inmobiliarias"},
                    "65512": {"nombre": "Activos adquiridos en arrendamiento financiero"},
                    "65513": {"nombre": "Inmuebles, maquinaria y equipo"},
                    "65514": {"nombre": "Intangibles"},
                    "65515": {"nombre": "Activos biológicos"}
                }},
                "6552": {"nombre": "Operaciones discontinuadas – Abandono de activos", "subcuentas": {
                    "65521": {"nombre": "Inversiones inmobiliarias"},
                    "65522": {"nombre": "Activos adquiridos en arrendamiento financiero"},
                    "65523": {"nombre": "Inmuebles, maquinaria y equipo"},
                    "65524": {"nombre": "Intangibles"},
                    "65525": {"nombre": "Activos biológicos"}
                }}
            }},
            "656": {"nombre": "Suministros"},
            "658": {"nombre": "Gestión medioambiental"},
            "659": {"nombre": "Otros gastos de gestión", "subcuentas": {
                "6591": {"nombre": "Donaciones"},
                "6592": {"nombre": "Sanciones administrativas"}
            }}
        }},
        "66": {"nombre": "Pérdida por medición de activos no financieros al valor razonable", "subcuentas": {
            "661": {"nombre": "Activo realizable", "subcuentas": {
                "6611": {"nombre": "Mercaderías"},
                "6612": {"nombre": "Productos terminados"},
                "6613": {"nombre": "Activos no corrientes mantenidos para la venta", "subcuentas": {
                    "66131": {"nombre": "Inversión inmobiliaria"},
                    "66132": {"nombre": "Inmuebles, maquinaria y equipo"},
                    "66133": {"nombre": "Intangibles"},
                    "66134": {"nombre": "Activos biológicos"}
                }}
            }},
            "662": {"nombre": "Activo inmovilizado", "subcuentas": {
                "6621": {"nombre": "Inversiones inmobiliarias"},
                "6622": {"nombre": "Activos biológicos"}
            }}
        }},
        "67": {"nombre": "Gastos financieros", "subcuentas": {
            "671": {"nombre": "Gastos en operaciones de endeudamiento y otros", "subcuentas": {
                "6711": {"nombre": "Préstamos de instituciones financieras y otras entidades"},
                "6712": {"nombre": "Contratos de arrendamiento financiero"},
                "6713": {"nombre": "Emisión y colocación de instrumentos representativos de deuda y patrimonio"},
                "6714": {"nombre": "Documentos vendidos o descontados"}
            }},
            "672": {"nombre": "Pérdida por instrumentos financieros derivados"},
            "673": {"nombre": "Intereses por préstamos y otras obligaciones", "subcuentas": {
                "6731": {"nombre": "Préstamos de instituciones financieras y otras entidades", "subcuentas": {
                    "67311": {"nombre": "Instituciones financieras"},
                    "67312": {"nombre": "Otras entidades"}
                }},
                "6732": {"nombre": "Contratos de arrendamiento financiero"},
                "6733": {"nombre": "Otros instrumentos financieros por pagar"},
                "6734": {"nombre": "Documentos vendidos o descontados"},
                "6735": {"nombre": "Obligaciones emitidas"},
                "6736": {"nombre": "Obligaciones comerciales"},
                "6737": {"nombre": "Obligaciones tributarias"}
            }},
            "674": {"nombre": "Gastos en operaciones de factoraje (factoring)", "subcuentas": {
                "6741": {"nombre": "Gastos por menor valor"}
            }},
            "675": {"nombre": "Descuentos concedidos por pronto pago"},
            "676": {"nombre": "Diferencia de cambio"},
            "677": {"nombre": "Pérdida por medición de activos y pasivos financieros al valor razonable", "subcuentas": {
                "6771": {"nombre": "Inversiones para negociación"},
                "6772": {"nombre": "Inversiones disponibles para la venta"},
                "6773": {"nombre": "Otros"}
            }},
            "678": {"nombre": "Participación en resultados de entidades relacionadas", "subcuentas": {
                "6781": {"nombre": "Participación en los resultados de subsidiarias y asociadas bajo el método del valor patrimonial"},
                "6782": {"nombre": "Participaciones en negocios conjuntos"}
            }},
            "679": {"nombre": "Otros gastos financieros", "subcuentas": {
                "6791": {"nombre": "Primas por opciones"},
                "6792": {"nombre": "Gastos financieros en medición a valor descontado"}
            }}
        }},
        "68": {"nombre": "Valuación y deterioro de activos y provisiones", "subcuentas": {
            "681": {"nombre": "Depreciación", "subcuentas": {
                "6811": {"nombre": "Depreciación de inversiones inmobiliarias", "subcuentas": {
                    "68111": {"nombre": "Edificaciones – Costo"},
                    "68112": {"nombre": "Edificaciones – Revaluación"},
                    "68113": {"nombre": "Edificaciones – Costo de financiación"}
                }},
                "6812": {"nombre": "Depreciación de activos adquiridos en arrendamiento financiero – Inversiones inmobiliarias", "subcuentas": {
                    "68121": {"nombre": "Edificaciones"}
                }},
                "6813": {"nombre": "Depreciación de activos adquiridos en arrendamiento financiero – Inmuebles, maquinaria y equipo", "subcuentas": {
                    "68131": {"nombre": "Edificaciones"},
                    "68132": {"nombre": "Maquinarias y equipos de explotación"},
                    "68133": {"nombre": "Equipo de transporte"},
                    "68134": {"nombre": "Equipos diversos"}
                }},
                "6814": {"nombre": "Depreciación de inmuebles, maquinaria y equipo – Costo", "subcuentas": {
                    "68141": {"nombre": "Edificaciones"},
                    "68142": {"nombre": "Maquinarias y equipos de explotación"},
                    "68143": {"nombre": "Equipo de transporte"},
                    "68144": {"nombre": "Muebles y enseres"},
                    "68145": {"nombre": "Equipos diversos"},
                    "68146": {"nombre": "Herramientas y unidades de reemplazo"}
                }},
                "6815": {"nombre": "Depreciación de inmuebles, maquinaria y equipo – Revaluación", "subcuentas": {
                    "68151": {"nombre": "Edificaciones"},
                    "68152": {"nombre": "Maquinarias y equipos de explotación"},
                    "68153": {"nombre": "Equipo de transporte"},
                    "68154": {"nombre": "Muebles y enseres"},
                    "68155": {"nombre": "Equipos diversos"},
                    "68156": {"nombre": "Herramientas y unidades de reemplazo"}
                }},
                "6816": {"nombre": "Depreciación de inmuebles, maquinaria y equipo – Costos de financiación", "subcuentas": {
                    "68161": {"nombre": "Edificaciones"},
                    "68162": {"nombre": "Maquinarias y equipos de explotación"}
                }},
                "6817": {"nombre": "Depreciación de activos biológicos en producción – Costo", "subcuentas": {
                    "68171": {"nombre": "Activos biológicos de origen animal"},
                    "68172": {"nombre": "Activos biológicos de origen vegetal"}
                }},
                "6818": {"nombre": "Depreciación de activos biológicos en producción – Costo de financiación", "subcuentas": {
                    "68181": {"nombre": "Activos biológicos de origen animal"},
                    "68182": {"nombre": "Activos biológicos de origen vegetal"}
                }},
            }},
            "682": {"nombre": "Amortización de intangibles", "subcuentas": {
                "6821": {"nombre": "Amortización de intangibles – Costo", "subcuentas": {
                    "68211": {"nombre": "Concesiones, licencias y otros derechos"},
                    "68212": {"nombre": "Patentes y propiedad industrial"},
                    "68213": {"nombre": "Programas de computadora (software)"},
                    "68214": {"nombre": "Costos de exploración y desarrollo"},
                    "68215": {"nombre": "Fórmulas, diseños y prototipos"},
                    "68219": {"nombre": "Otros activos intangibles"}
                }},
                "6822": {"nombre": "Amortización de intangibles – Revaluación", "subcuentas": {
                    "68221": {"nombre": "Concesiones, licencias y otros derechos"},
                    "68222": {"nombre": "Patentes y propiedad industrial"},
                    "68223": {"nombre": "Programas de computadora (software)"},
                    "68224": {"nombre": "Costos de exploración y desarrollo"},
                    "68225": {"nombre": "Fórmulas, diseños y prototipos"},
                    "68229": {"nombre": "Otros activos intangibles"}
                }}
            }},
            "683": {"nombre": "Agotamiento", "subcuentas": {
                "6831": {"nombre": "Agotamiento de recursos naturales adquiridos"}
            }},
            "684": {"nombre": "Valuación de activos", "subcuentas": {
                "6841": {"nombre": "Estimación de cuentas de cobranza dudosa", "subcuentas": {
                    "68411": {"nombre": "Cuentas por cobrar comerciales – Terceros"},
                    "68412": {"nombre": "Cuentas por cobrar comerciales – Relacionadas"},
                    "68413": {"nombre": "Cuentas por cobrar al personal, a los accionistas (socios), directores y gerentes"},
                    "68414": {"nombre": "Cuentas por cobrar diversas – Terceros"},
                    "68415": {"nombre": "Cuentas por cobrar diversas – Relacionadas"}
                }},
                "6843": {"nombre": "Desvalorización de inversiones mobiliarias", "subcuentas": {
                    "68431": {"nombre": "Inversiones a ser mantenidas hasta el vencimiento"},
                    "68432": {"nombre": "Instrumentos financieros representativos de derecho patrimonial"}
                }}
            }},
            "685": {"nombre": "Deterioro del valor de los activos", "subcuentas": {
                "6851": {"nombre": "Desvalorización de inversiones inmobiliarias", "subcuentas": {
                    "68511": {"nombre": "Edificaciones"}
                }},
                "6852": {"nombre": "Desvalorización de inmuebles maquinaria y equipo", "subcuentas": {
                    "68521": {"nombre": "Edificaciones"},
                    "68522": {"nombre": "Maquinarias y equipos de explotación"},
                    "68523": {"nombre": "Equipo de transporte"},
                    "68524": {"nombre": "Muebles y enseres"},
                    "68525": {"nombre": "Equipos diversos"},
                    "68526": {"nombre": "Herramientas y unidades de reemplazo"}
                }},
                "6853": {"nombre": "Desvalorización de intangibles", "subcuentas": {
                    "68531": {"nombre": "Concesiones, licencias y otros derechos"},
                    "68532": {"nombre": "Patentes y propiedad industrial"},
                    "68533": {"nombre": "Programas de computadora (software)"},
                    "68534": {"nombre": "Costos de exploración y desarrollo"},
                    "68535": {"nombre": "Fórmulas, diseños y prototipos"},
                    "68536": {"nombre": "Otros activos intangibles"},
                    "68537": {"nombre": "Plusvalía mercantil"}
                }},
                "6854": {"nombre": "Desvalorización de activos biológicos en producción", "subcuentas": {
                    "68541": {"nombre": "Activos biológicos de origen animal"},
                    "68542": {"nombre": "Activos biológicos de origen vegetal"}
                }}
            }},
            "686": {"nombre": "Provisiones", "subcuentas": {
                "6861": {"nombre": "Provisión para litigios", "subcuentas": {
                    "68611": {"nombre": "Provisión para litigios – Costo"},
                    "68612": {"nombre": "Provisión para litigios – Actualización financiera"}
                }},
                "6862": {"nombre": "Provisión por desmantelamiento, retiro o rehabilitación del inmovilizado", "subcuentas": {
                    "68621": {"nombre": "Provisión por desmantelamiento, retiro o rehabilitación del inmovilizado – Costo"},
                    "68622": {"nombre": "Provisión por desmantelamiento, retiro o rehabilitación del inmovilizado – Actualización financiera"}
                }},
                "6863": {"nombre": "Provisión para reestructuraciones"},
                "6864": {"nombre": "Provisión para protección y remediación del medio ambiente", "subcuentas": {
                    "68641": {"nombre": "Provisión para protección y remediación del medio ambiente – Costo"},
                    "68642": {"nombre": "Provisión para protección y remediación del medio ambiente – Actualización financiera"}
                }},
                "6866": {"nombre": "Provisión para garantías", "subcuentas": {
                    "68661": {"nombre": "Provisión para garantías – Costo"},
                    "68662": {"nombre": "Provisión para garantías – Actualización financiera"}
                }},
                "6869": {"nombre": "Otras provisiones"}
            }}
        }},
        "69": {"nombre": "Costo de ventas", "subcuentas": {
            "691": {"nombre": "Mercaderías", "subcuentas": {
                "6911": {"nombre": "Mercaderías manufacturadas", "subcuentas": {
                    "69111": {"nombre": "Terceros"},
                    "69112": {"nombre": "Relacionadas"}
                }},
                "6912": {"nombre": "Mercaderías de extracción", "subcuentas": {
                    "69121": {"nombre": "Terceros"},
                    "69122": {"nombre": "Relacionadas"}
                }},
                "6913": {"nombre": "Mercaderías agropecuarias y piscícolas", "subcuentas": {
                    "69131": {"nombre": "Terceros"},
                    "69132": {"nombre": "Relacionadas"}
                }},
                "6914": {"nombre": "Mercaderías inmuebles", "subcuentas": {
                    "69141": {"nombre": "Terceros"},
                    "69142": {"nombre": "Relacionadas"}
                }},
                "6915": {"nombre": "Otras mercaderías", "subcuentas": {
                    "69151": {"nombre": "Terceros"},
                    "69152": {"nombre": "Relacionadas"}
                }}
            }},
            "692": {"nombre": "Productos terminados", "subcuentas": {
                "6921": {"nombre": "Productos manufacturados", "subcuentas": {
                    "69211": {"nombre": "Terceros"},
                    "69212": {"nombre": "Relacionadas"}
                }},
                "6922": {"nombre": "Productos de extracción terminados", "subcuentas": {
                    "69221": {"nombre": "Terceros"},
                    "69222": {"nombre": "Relacionadas"}
                }},
                "6923": {"nombre": "Productos agropecuarios y piscícolas terminados", "subcuentas": {
                    "69231": {"nombre": "Terceros"},
                    "69232": {"nombre": "Relacionadas"}
                }},
                "6924": {"nombre": "Productos inmuebles terminados", "subcuentas": {
                    "69241": {"nombre": "Terceros"},
                    "69242": {"nombre": "Relacionadas"}
                }},
                "6925": {"nombre": "Existencias de servicios terminados", "subcuentas": {
                    "69251": {"nombre": "Terceros"},
                    "69252": {"nombre": "Relacionadas"}
                }},
                "6926": {"nombre": "Costos de financiación – Productos terminados", "subcuentas": {
                    "69261": {"nombre": "Terceros"},
                    "69262": {"nombre": "Relacionadas"}
                }},
                "6927": {"nombre": "Costos de producción no absorbido – Productos terminados"},
                "6928": {"nombre": "Costo de ineficiencia – Productos terminados"}
            }},
            "693": {"nombre": "Subproductos, desechos y desperdicios", "subcuentas": {
                "6931": {"nombre": "Subproductos", "subcuentas": {
                    "69311": {"nombre": "Terceros"},
                    "69312": {"nombre": "Relacionadas"}
                }},
                "6932": {"nombre": "Desechos y desperdicios", "subcuentas": {
                    "69321": {"nombre": "Terceros"},
                    "69322": {"nombre": "Relacionadas"}
                }}
            }},
            "694": {"nombre": "Servicios", "subcuentas": {
                "6941": {"nombre": "Terceros"},
                "6942": {"nombre": "Relacionadas"}
            }},
            "695": {"nombre": "Gastos por desvalorización de existencias", "subcuentas": {
                "6951": {"nombre": "Mercaderías"},
                "6952": {"nombre": "Productos terminados"},
                "6953": {"nombre": "Subproductos, desechos y desperdicios"},
                "6954": {"nombre": "Productos en proceso"},
                "6955": {"nombre": "Materias primas"},
                "6956": {"nombre": "Materiales auxiliares, suministros y repuestos"},
                "6957": {"nombre": "Envases y embalajes"},
                "6958": {"nombre": "Existencias por recibir"}
            }}
        }},
        "70": {"nombre": "Ventas", "subcuentas": {
            "701": {"nombre": "Mercaderías", "subcuentas": {
                "7011": {"nombre": "Mercaderías manufacturadas", "subcuentas": {
                    "70111": {"nombre": "Terceros"},
                    "70112": {"nombre": "Relacionadas"}
                }},
                "7012": {"nombre": "Mercaderías de extracción", "subcuentas": {
                "70121": {"nombre": "Terceros"},
                    "70122": {"nombre": "Relacionadas"}
                }},
                "7013": {"nombre": "Mercaderías agropecuarias y piscícolas", "subcuentas": {
                    "70131": {"nombre": "Terceros"},
                    "70132": {"nombre": "Relacionadas"}
                }},
                "7014": {"nombre": "Mercaderías inmuebles", "subcuentas": {
                    "70141": {"nombre": "Terceros"},
                    "70142": {"nombre": "Relacionadas"}
                }},
                "7015": {"nombre": "Mercaderías – Otras", "subcuentas": {
                    "70151": {"nombre": "Terceros"},
                    "70152": {"nombre": "Relacionadas"}
                }}
            }},
            "702": {"nombre": "Productos terminados", "subcuentas": {
                "7021": {"nombre": "Productos manufacturados", "subcuentas": {
                    "70211": {"nombre": "Terceros"},
                    "70212": {"nombre": "Relacionadas"}
                }},
                "7022": {"nombre": "Productos de extracción terminados", "subcuentas": {
                    "70221": {"nombre": "Terceros"},
                    "70222": {"nombre": "Relacionadas"}
                }},
                "7023": {"nombre": "Productos agropecuarios y piscícolas terminados", "subcuentas": {
                    "70231": {"nombre": "Terceros"},
                    "70232": {"nombre": "Relacionadas"}
                }},
                "7024": {"nombre": "Productos inmuebles terminados", "subcuentas": {
                    "70241": {"nombre": "Terceros"},
                    "70242": {"nombre": "Relacionadas"}
                }},
                "7025": {"nombre": "Existencias de servicios terminados", "subcuentas": {
                    "70251": {"nombre": "Terceros"},
                    "70252": {"nombre": "Relacionadas"}
                }}
            }},
            "703": {"nombre": "Subproductos, desechos y desperdicios", "subcuentas": {
                "7031": {"nombre": "Subproductos", "subcuentas": {
                    "70311": {"nombre": "Terceros"},
                    "70312": {"nombre": "Relacionadas"}
                }},
                "7032": {"nombre": "Desechos y desperdicios", "subcuentas": {
                    "70321": {"nombre": "Terceros"},
                    "70322": {"nombre": "Relacionadas"}
                }}
            }},
            "704": {"nombre": "Prestación de servicios", "subcuentas": {
                "7041": {"nombre": "Terceros"},
                "7042": {"nombre": "Relacionadas"}
            }},
            "709": {"nombre": "Devoluciones sobre ventas", "subcuentas": {
                "7091": {"nombre": "Mercaderías – Terceros", "subcuentas": {
                    "70911": {"nombre": "Mercaderías manufacturadas"},
                    "70912": {"nombre": "Mercaderías de extracción"},
                    "70913": {"nombre": "Mercaderías agropecuarias y piscícolas"},
                    "70914": {"nombre": "Mercaderías inmuebles"},
                    "70915": {"nombre": "Mercaderías – Otras"}
                }},
                "7092": {"nombre": "Mercaderías – Relacionadas", "subcuentas": {
                    "70921": {"nombre": "Mercaderías manufacturadas"},
                    "70922": {"nombre": "Mercaderías de extracción"},
                    "70923": {"nombre": "Mercaderías agropecuarias y piscícolas"},
                    "70924": {"nombre": "Mercaderías inmuebles"},
                    "70925": {"nombre": "Mercaderías – Otras"}
                }},
                "7093": {"nombre": "Productos terminados – Terceros", "subcuentas": {
                    "70931": {"nombre": "Productos manufacturados"},
                    "70932": {"nombre": "Productos de extracción terminados"},
                    "70933": {"nombre": "Productos agropecuarios y piscícolas terminados"},
                    "70934": {"nombre": "Productos inmuebles terminados"},
                    "70935": {"nombre": "Existencias de servicios terminados"}
                }},
                "7094": {"nombre": "Productos terminados – Relacionadas", "subcuentas": {
                    "70941": {"nombre": "Productos manufacturados"},
                    "70942": {"nombre": "Productos de extracción terminados"},
                    "70943": {"nombre": "Productos agropecuarios y piscícolas terminados"},
                    "70944": {"nombre": "Productos inmuebles terminados"},
                    "70945": {"nombre": "Existencias de servicios terminados"}
                }},
                "7095": {"nombre": "Subproductos, desechos y desperdicios – Terceros", "subcuentas": {
                    "70951": {"nombre": "Subproductos"},
                 "70952": {"nombre": "Desechos y desperdicios"}
                }},
                "7096": {"nombre": "Subproductos, desechos y desperdicios – Relacionadas", "subcuentas": {
                    "70961": {"nombre": "Subproductos"},
                    "70962": {"nombre": "Desechos y desperdicios"}
                }}
            }}
        }},
        "71": {"nombre": "Variación de la producción almacenada", "subcuentas": {
            "711": {"nombre": "Variación de productos terminados", "subcuentas": {
                "7111": {"nombre": "Productos manufacturados"},
                "7112": {"nombre": "Productos de extracción terminados"},
                "7113": {"nombre": "Productos agropecuarios y piscícolas terminados"},
                "7114": {"nombre": "Productos inmuebles terminados"},
                "7115": {"nombre": "Existencias de servicios terminados"}
            }},
            "712": {"nombre": "Variación de subproductos, desechos y desperdicios", "subcuentas": {
                "7121": {"nombre": "Subproductos"},
                "7122": {"nombre": "Desechos y desperdicios"}
            }},
            "713": {"nombre": "Variación de productos en proceso", "subcuentas": {
                "7131": {"nombre": "Productos en proceso de manufactura"},
                "7132": {"nombre": "Productos extraídos en proceso de transformación"},
                "7133": {"nombre": "Productos agropecuarios y piscícolas en proceso"},
                "7134": {"nombre": "Productos inmuebles en proceso"},
                "7135": {"nombre": "Existencias de servicios en proceso"},
                "7138": {"nombre": "Otros productos en proceso"}
            }},
            "714": {"nombre": "Variación de envases y embalajes", "subcuentas": {
                "7141": {"nombre": "Envases"},
                "7142": {"nombre": "Embalajes"}
            }},
            "715": {"nombre": "Variación de existencias de servicios"}
        }},
        "72": {"nombre": "Producción de activo inmovilizado", "subcuentas": {
            "721": {"nombre": "Inversiones inmobiliarias", "subcuentas": {
                "7211": {"nombre": "Edificaciones"}
            }},
            "722": {"nombre": "Inmuebles, maquinaria y equipo", "subcuentas": {
                "7221": {"nombre": "Edificaciones"},
                "7222": {"nombre": "Maquinarias y otros equipos de explotación"},
                "7223": {"nombre": "Equipo de transporte"},
                "7224": {"nombre": "Muebles y enseres"},
                "7225": {"nombre": "Equipos diversos"}
            }},
            "723": {"nombre": "Intangibles", "subcuentas": {
                "7231": {"nombre": "Programas de computadora (software)"},
                "7232": {"nombre": "Costos de exploración y desarrollo"},
                "7233": {"nombre": "Fórmulas, diseños y prototipos"}
            }},
            "724": {"nombre": "Activos biológicos", "subcuentas": {
                "7241": {"nombre": "Activos biológicos en desarrollo de origen animal"},
                "7242": {"nombre": "Activos biológicos en desarrollo de origen vegetal"}
            }},
            "725": {"nombre": "Costos de financiación capitalizados", "subcuentas": {
                "7251": {"nombre": "Costos de financiación – Inversiones inmobiliarias", "subcuentas": {
                    "72511": {"nombre": "Edificaciones"}
                }},
                "7252": {"nombre": "Costos de financiación – Inmuebles, maquinaria y equipo", "subcuentas": {
                    "72521": {"nombre": "Edificaciones"},
                    "72522": {"nombre": "Maquinarias y otros equipos de explotación"}
                }}
            }}
        }},
        "73": {"nombre": "Descuentos, rebajas y bonificaciones obtenidos", "subcuentas": {
            "731": {"nombre": "Descuentos, rebajas y bonificaciones obtenidos", "subcuentas": {
                "7311": {"nombre": "Terceros"},
                "7312": {"nombre": "Relacionadas"}
            }}
        }},
        "74": {"nombre": "Descuentos, rebajas y bonificaciones concedidos", "subcuentas": {
            "741": {"nombre": "Descuentos, rebajas y bonificaciones concedidos", "subcuentas": {
                "7411": {"nombre": "Terceros"},
                "7412": {"nombre": "Relacionadas"}
            }}
        }},
        "75": {"nombre": "Otros ingresos de gestión", "subcuentas": {
            "751": {"nombre": "Servicios en beneficio del personal"},
            "752": {"nombre": "Comisiones y corretajes"},
            "753": {"nombre": "Regalías"},
            "754": {"nombre": "Alquileres", "subcuentas": {
                "7541": {"nombre": "Terrenos"},
                "7542": {"nombre": "Edificaciones"},
                "7543": {"nombre": "Maquinarias y equipos de explotación"},
                "7544": {"nombre": "Equipo de transporte"},
                "7545": {"nombre": "Equipos diversos"}
            }},
            "755": {"nombre": "Recuperación de cuentas de valuación", "subcuentas": {
                "7551": {"nombre": "Recuperación – Cuentas de cobranza dudosa"},
                "7552": {"nombre": "Recuperación – Desvalorización de existencias"},
                "7553": {"nombre": "Recuperación – Desvalorización de inversiones mobiliarias"}
            }},
            "756": {"nombre": "Enajenación de activos inmovilizados", "subcuentas": {
                "7561": {"nombre": "Inversiones mobiliarias"},
                "7562": {"nombre": "Inversiones inmobiliarias"},
                "7563": {"nombre": "Activos adquiridos en arrendamiento financiero"},
                "7564": {"nombre": "Inmuebles, maquinaria y equipo"},
                "7565": {"nombre": "Intangibles"},
                "7566": {"nombre": "Activos biológicos"}
            }},
            "757": {"nombre": "Recuperación de deterioro de cuentas de activos inmovilizados", "subcuentas": {
                "7571": {"nombre": "Recuperación de deterioro de inversiones inmobiliarias"},
                "7572": {"nombre": "Recuperación de deterioro de inmuebles, maquinaria y equipo"},
                "7573": {"nombre": "Recuperación de deterioro de intangibles"},
                "7574": {"nombre": "Recuperación de deterioro de activos biológicos"}
            }},
            "759": {"nombre": "Otros ingresos de gestión", "subcuentas": {
                "7591": {"nombre": "Subsidios gubernamentales"},
                "7592": {"nombre": "Reclamos al seguro"},
                "7593": {"nombre": "Donaciones"},
                "7599": {"nombre": "Otros ingresos de gestión"}
            }}
        }},
        "76": {"nombre": "Ganancia por medición de activos no financieros al valor razonable", "subcuentas": {
            "761": {"nombre": "Activo realizable", "subcuentas": {
                "7611": {"nombre": "Mercaderías"},
                "7612": {"nombre": "Productos terminados"},
                "7613": {"nombre": "Activos no corrientes mantenidos para la venta", "subcuentas": {
                    "76131": {"nombre": "Inversión inmobiliaria"},
                    "76132": {"nombre": "Inmuebles, maquinaria y equipo"},
                    "76133": {"nombre": "Intangibles"},
                    "76134": {"nombre": "Activos biológicos"}
                }}
            }},
            "762": {"nombre": "Activo inmovilizado", "subcuentas": {
                "7621": {"nombre": "Inversiones inmobiliarias"},
                "7622": {"nombre": "Activos biológicos"}
            }}
        }},
        "77": {"nombre": "Ingresos financieros", "subcuentas": {
            "771": {"nombre": "Ganancia por instrumento financiero derivado"},
            "772": {"nombre": "Rendimientos ganados", "subcuentas": {
                "7721": {"nombre": "Depósitos en instituciones financieras"},
                "7722": {"nombre": "Cuentas por cobrar comerciales"},
                "7723": {"nombre": "Préstamos otorgados"},
                "7724": {"nombre": "Inversiones a ser mantenidas hasta el vencimiento"},
                "7725": {"nombre": "Instrumentos financieros representativos de derecho patrimonial"}
            }},
            "773": {"nombre": "Dividendos"},
            "774": {"nombre": "Ingresos en operaciones de factoraje (factoring)"},
            "775": {"nombre": "Descuentos obtenidos por pronto pago"},
            "776": {"nombre": "Diferencia en cambio"},
            "777": {"nombre": "Ganancia por medición de activos y pasivos financieros al valor razonable", "subcuentas": {
                "7771": {"nombre": "Inversiones mantenidas para negociación"},
                "7772": {"nombre": "Inversiones disponibles para la venta"},
                "7773": {"nombre": "Otros"}
            }},
            "778": {"nombre": "Participación en resultados de entidades relacionadas", "subcuentas": {
                "7781": {"nombre": "Participación en los resultados de subsidiarias y asociadas bajo el método del valor patrimonial"},
                "7782": {"nombre": "Ingresos por participaciones en negocios conjuntos"}
            }},
            "779": {"nombre": "Otros ingresos financieros", "subcuentas": {
                "7792": {"nombre": "Ingresos financieros en medición a valor descontado"}
            }}
        }},
        "78": {"nombre": "Cargas cubiertas por provisiones", "subcuentas": {
            "781": {"nombre": "Cargas cubiertas por provisiones"}
        }},
        "79": {"nombre": "Cargas imputables a cuentas de costos y gastos", "subcuentas": {
            "791": {"nombre": "Cargas imputables a cuentas de costos y gastos"},
            "792": {"nombre": "Gastos financieros imputables a cuentas de existencias"}
        }},
        "80": {"nombre": "Margen comercial", "subcuentas": {
            "801": {"nombre": "Margen comercial"}
        }},
        "81": {"nombre": "Producción del ejercicio", "subcuentas": {
            "811": {"nombre": "Producción de bienes"},
            "812": {"nombre": "Producción de servicios"},
            "813": {"nombre": "Producción de activo inmovilizado"}
        }},
        "82": {"nombre": "Valor agregado", "subcuentas": {
            "821": {"nombre": "Valor agregado"}
        }},
        "83": {"nombre": "Excedente bruto (insuficiencia bruta) de explotación", "subcuentas": {
            "831": {"nombre": "Excedente bruto (insuficiencia bruta) de explotación"}
        }},
        "84": {"nombre": "Resultado de explotación", "subcuentas": {
            "841": {"nombre": "Resultado de explotación"}
        }},
        "85": {"nombre": "Resultado antes de participaciones e impuestos", "subcuentas": {
            "851": {"nombre": "Resultado antes de participaciones e impuestos"}
        }},
        "87": {"nombre": "Participaciones de los trabajadores", "subcuentas": {
            "871": {"nombre": "Participación de los trabajadores – Corriente"},
            "872": {"nombre": "Participación de los trabajadores – Diferida"}
        }},
        "88": {"nombre": "Impuesto a la renta", "subcuentas": {
            "881": {"nombre": "Impuesto a la renta – Corriente"},
            "882": {"nombre": "Impuesto a la renta – Diferido"}
        }},
        "89": {"nombre": "Determinación del resultado del ejercicio", "subcuentas": {
            "891": {"nombre": "Utilidad"},
            "892": {"nombre": "Pérdida"}
        }},
        "90": {"nombre": "Cuentas reflejas", "subcuentas": {
        }},
        "91": {"nombre": "Costos por distribuir", "subcuentas": {
        }},
        "92": {"nombre": "Costos de producción", "subcuentas": {
        }},
        "93": {"nombre": "Centro de costos", "subcuentas": {
        }},
        "94": {"nombre": "Gastos de administración", "subcuentas": {
        }},
        "95": {"nombre": "Gastos de ventas", "subcuentas": {
        }},
        "96": {"nombre": "Inventario permanentes", "subcuentas": {
        }},
        "97": {"nombre": "Gastos financieros", "subcuentas": {
        }},
        "98": {"nombre": "Gastos de créditos y cobranzas", "subcuentas": {
        }},
        "99": {"nombre": "Gastos de almacén", "subcuentas": {
        }},
        "01": {"nombre": "Bienes y valores entregados", "subcuentas": {
            "011": {"nombre": "Bienes en préstamo, custodia y no capitalizables", "subcuentas": {
                "0111": {"nombre": "Bienes en préstamo"},
                "0112": {"nombre": "Bienes en custodia"}
            }},
            "012": {"nombre": "Valores y bienes entregados en garantía", "subcuentas": {
                "0121": {"nombre": "Cartas fianza"},
                "0122": {"nombre": "Cuentas por cobrar"},
                "0123": {"nombre": "Existencias"},
                "0124": {"nombre": "Inversión mobiliaria"},
                "0125": {"nombre": "Inversión inmobiliaria"},
                "0126": {"nombre": "Inmuebles, maquinaria y equipo"},
                "0127": {"nombre": "Intangibles"},
                "0128": {"nombre": "Activos biológicos"}
            }},
            "013": {"nombre": "Activos realizables entregados en consignación"}
        }},
        "02": {"nombre": "Derechos sobre instrumentos financieros", "subcuentas": {
            "021": {"nombre": "Primarios", "subcuentas": {
                "0211": {"nombre": "Inversiones mantenidas para negociación"},
                "0212": {"nombre": "Inversiones disponibles para la venta"},
                "0213": {"nombre": "Inversiones a ser mantenidas hasta el vencimiento"}
            }},
            "022": {"nombre": "Derivados", "subcuentas": {
                "0221": {"nombre": "Contratos a futuro"},
                "0222": {"nombre": "Contratos a término (forward)"},
                "0223": {"nombre": "Permutas financieras (swap)"},
                "0224": {"nombre": "Contratos de opción"}
            }}
        }},
        "03": {"nombre": "Otras cuentas de orden deudoras", "subcuentas": {
            "031": {"nombre": "Contratos aprobados", "subcuentas": {
                "0311": {"nombre": "Contratos en ejecución"},
                "0312": {"nombre": "Contratos en trámite"}
            }},
            "032": {"nombre": "Bienes dados de baja", "subcuentas": {
                "0321": {"nombre": "Suministros"},
                "0322": {"nombre": "Inmuebles, maquinaria y equipo"}
            }},
            "039": {"nombre": "Diversas"}
        }},
        "04": {"nombre": "Deudoras por contra"},
        "06": {"nombre": "Bienes y valores recibidos", "subcuentas": {
            "061": {"nombre": "Bienes recibidos en préstamo y custodia", "subcuentas": {
                "0611": {"nombre": "Bienes recibidos en préstamo"},
                "0612": {"nombre": "Bienes recibidos en custodia"}
            }},
            "062": {"nombre": "Valores y bienes recibidos en garantía", "subcuentas": {
                "0621": {"nombre": "Cartas fianza"},
                "0622": {"nombre": "Cuentas por cobrar"},
                "0623": {"nombre": "Existencias"},
                "0624": {"nombre": "Inversión mobiliaria"},
                "0625": {"nombre": "Inversión inmobiliaria"},
                "0626": {"nombre": "Inmuebles, maquinaria y equipo"},
                "0627": {"nombre": "Intangibles"},
                "0628": {"nombre": "Activos biológicos"}
            }},
            "063": {"nombre": "Activos realizables recibidos en consignación"}
        }},
        "07": {"nombre": "Compromisos sobre instrumentos financieros", "subcuentas": {
            "071": {"nombre": "Primarios", "subcuentas": {
                "0711": {"nombre": "Inversiones mantenidas para negociación"},
                "0712": {"nombre": "Inversiones disponibles para la venta"},
                "0713": {"nombre": "Inversiones a ser mantenidas hasta el vencimiento"}
            }},
            "072": {"nombre": "Derivados", "subcuentas": {
                "0721": {"nombre": "Contratos a futuro"},
                "0722": {"nombre": "Contratos a término (forward)"},
                "0723": {"nombre": "Permutas financieras (swap)"},
                "0724": {"nombre": "Contratos de opción"}
            }}
        }},
        "08": {"nombre": "Otras cuentas de orden acreedoras", "subcuentas": {
            "089": {"nombre": "Diversas"}
        }},
        "09": {"nombre": "Acreedoras por contra"},

     }

def mostrar_seleccion_cuentas(parent, debe_haber):
    def seleccionar_cuenta():
        item = tree.focus()
        if item:
            valores = tree.item(item, "values")
            if valores and len(valores) >= 2:
                codigo = valores[0]
                nombre = valores[1]
                cuenta_completa = f"{codigo} - {nombre}"
                
                if debe_haber == "debe":
                    cuenta_debe_var.set(cuenta_completa)
                else:
                    cuenta_haber_var.set(cuenta_completa)
                
                seleccion_window.destroy()
            else:
                messagebox.showerror("Error", "Seleccione una cuenta válida")
        else:
            messagebox.showerror("Error", "Seleccione una cuenta de la lista")

    # Crear ventana de selección
    seleccion_window = tk.Toplevel(parent)
    seleccion_window.title(f"Seleccionar Cuenta ({'DEBE' if debe_haber == 'debe' else 'HABER'})")
    seleccion_window.geometry("900x700")
    
    # Configurar grid principal
    seleccion_window.grid_columnconfigure(0, weight=1)
    seleccion_window.grid_rowconfigure(0, weight=1)
    
    # Frame principal
    main_frame = ttk.Frame(seleccion_window)
    main_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
    
    # Treeview con scrollbar
    tree_frame = ttk.Frame(main_frame)
    tree_frame.pack(fill="both", expand=True, pady=(0, 10))
    
    # Configurar Treeview
    tree = ttk.Treeview(
        tree_frame,
        columns=("codigo", "nombre"),
        show="tree headings",  # Mostrar estructura de árbol y encabezados
        selectmode="browse"
    )
    
    # Configurar columnas
    tree.heading("codigo", text="Código", anchor="center")
    tree.heading("nombre", text="Nombre de Cuenta", anchor="w")
    tree.column("codigo", width=150, anchor="center")
    tree.column("nombre", width=700, anchor="w")
    
    # Scrollbar
    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    scrollbar.pack(side="right", fill="y")
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(fill="both", expand=True)
    
    # Función recursiva para cargar cuentas y subcuentas
    def cargar_cuentas_recursivo(parent_id, cuentas_dict):
        for codigo, datos in cuentas_dict.items():
            # Determinar si tiene subcuentas para mostrar flecha
            has_children = "subcuentas" in datos and bool(datos["subcuentas"])
            
            # Insertar la cuenta (inicialmente colapsada)
            current_id = tree.insert(
                parent_id, "end", 
                values=(codigo, datos["nombre"]),
                open=False  # Mostrar colapsado por defecto
            )
            
            # Si tiene subcuentas, cargarlas (pero permanecerán ocultas hasta expandir)
            if has_children:
                cargar_cuentas_recursivo(current_id, datos["subcuentas"])
    
    # Cargar cuentas de forma recursiva (inicialmente colapsadas)
    arbol_cuentas = crear_arbol_cuentas()
    cargar_cuentas_recursivo("", arbol_cuentas)
    
    # Frame de botones
    btn_frame = ttk.Frame(main_frame)
    btn_frame.pack(pady=(5, 0))
    
    ttk.Button(
        btn_frame,
        text="Seleccionar",
        command=seleccionar_cuenta,
        bootstyle=SUCCESS
    ).pack(side="left", padx=5)
    
    ttk.Button(
        btn_frame,
        text="Cancelar",
        command=seleccion_window.destroy,
        bootstyle=DANGER
    ).pack(side="left", padx=5)
    
    # Evento doble clic
    tree.bind("<Double-1>", lambda e: seleccionar_cuenta())

def configurar_cuenta_debe():
    mostrar_seleccion_cuentas(root, "debe")

def configurar_cuenta_haber():
    mostrar_seleccion_cuentas(root, "haber")
    
def crear_formulario_asientos():
    global tree_debe, tree_haber, monto_entry, glosa_entry
    global cuenta_debe_var, cuenta_haber_var, costos_combobox, actividad_combobox
    global frame_botones_accion, frame_confirmacion
    
    # --- Contenedor principal con scroll invisible ---
    contenedor_scroll = ttk.Frame(contenedor_principal)
    contenedor_scroll.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
    contenedor_scroll.grid_rowconfigure(0, weight=1)
    contenedor_scroll.grid_columnconfigure(0, weight=1)
    
    # Canvas para el scroll
    canvas = tk.Canvas(contenedor_scroll, highlightthickness=0)
    canvas.grid(row=0, column=3, sticky="nsew")
    
    # Scrollbar (completamente invisible)
    scrollbar = ttk.Scrollbar(contenedor_scroll, orient="vertical", command=canvas.yview)
    scrollbar.grid(row=3, column=1, sticky="ns")
    
    # Configuración para hacer la scrollbar invisible (versión corregida)
    style = ttk.Style()
    style.configure("Invisible.Vertical.TScrollbar", 
                  troughcolor=contenedor_scroll['style'],
                  background=contenedor_scroll['style'],
                  bordercolor=contenedor_scroll['style'],
                  arrowcolor=contenedor_scroll['style'])
    scrollbar.configure(style="Invisible.Vertical.TScrollbar")
    
    # Configurar el canvas
    canvas.configure(yscrollcommand=scrollbar.set)
    
    # Frame interno para todos los widgets
    formulario_asientos = ttk.Frame(canvas)
    canvas.create_window((0, 0), window=formulario_asientos, anchor="nw")
    
    # Configurar el scroll con rueda del ratón
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    # Actualizar el scrollregion cuando cambie el tamaño
    def _configure_canvas(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
        if formulario_asientos.winfo_reqwidth() != canvas.winfo_width():
            canvas.config(width=formulario_asientos.winfo_reqwidth())
    
    formulario_asientos.bind("<Configure>", _configure_canvas)
    
    # --- Variables ---
    cuenta_debe_var = tk.StringVar(value="Seleccionar cuenta...")
    cuenta_haber_var = tk.StringVar(value="Seleccionar cuenta...")
    
    # ========== SECCIÓN DEBE ==========
    debe_frame = ttk.LabelFrame(formulario_asientos, text="DEBE", bootstyle="info", padding=10)
    debe_frame.grid(row=0, column=0, sticky="ew", pady=5)
    
    # Treeview para débitos
    tree_debe = ttk.Treeview(
        debe_frame,
        columns=("num", "cuenta", "monto"),
        show="headings",
        height=1,
        selectmode="browse",
        style='Dynamic.Treeview'
    )
    tree_debe.heading("num", text="#", anchor="center")
    tree_debe.heading("cuenta", text="Cuenta", anchor="w")
    tree_debe.heading("monto", text="Monto", anchor="e")
    tree_debe.column("num", width=40, anchor="center")
    tree_debe.column("cuenta", width=250, anchor="w")
    tree_debe.column("monto", width=120, anchor="e")
    tree_debe.pack(fill="x", pady=(0, 5))
    
    # Frame para agregar débitos
    frame_agregar_debe = ttk.Frame(debe_frame)
    frame_agregar_debe.pack(fill="x", pady=(5, 0))
    
    # Botón para seleccionar cuenta débito
    btn_seleccionar_debe = ttk.Button(
        frame_agregar_debe,
        text="Seleccionar Cuenta",
        command=configurar_cuenta_debe,
        bootstyle=(OUTLINE, INFO),
        width=15
    )
    btn_seleccionar_debe.pack(side="left", padx=2)
    
    # Etiqueta que muestra la cuenta seleccionada
    ttk.Label(
        frame_agregar_debe,
        text="Cuenta:"
    ).pack(side="left", padx=5)
    
    ttk.Label(
        frame_agregar_debe,
        textvariable=cuenta_debe_var,
        width=30,
        relief="solid",
        padding=5
    ).pack(side="left", fill="x", expand=True, padx=5)
    
    # Botones de acción para débito
    frame_botones_debe = ttk.Frame(debe_frame)
    frame_botones_debe.pack(fill="x", pady=(5, 0))
    
    ttk.Button(
        frame_botones_debe,
        text="Agregar Débito",
        command=agregar_cuenta_debe,
        bootstyle=INFO,
        width=15
    ).pack(side="left", padx=2)
    
    ttk.Button(
        frame_botones_debe,
        text="Borrar",
        command=eliminar_item_debe,
        bootstyle=(OUTLINE, INFO),
        width=8
    ).pack(side="left", padx=2)
    
    # Total débitos
    global lbl_total_debe
    lbl_total_debe = ttk.Label(debe_frame, text="Total Debe: S/. 0.00", font=('Segoe UI', 9, 'bold'))
    lbl_total_debe.pack(anchor="e", pady=(5, 0))
    
    # ========== SECCIÓN HABER ==========
    haber_frame = ttk.LabelFrame(formulario_asientos, text="HABER", bootstyle="danger", padding=10)
    haber_frame.grid(row=1, column=0, sticky="ew", pady=5)
    
    # Treeview para créditos
    tree_haber = ttk.Treeview(
        haber_frame,
        columns=("num", "cuenta", "monto"),
        show="headings",
        height=1,
        selectmode="browse",
        style='Dynamic.Treeview'
    )
    tree_haber.heading("num", text="#", anchor="center")
    tree_haber.heading("cuenta", text="Cuenta", anchor="w")
    tree_haber.heading("monto", text="Monto", anchor="e")
    tree_haber.column("num", width=40, anchor="center")
    tree_haber.column("cuenta", width=250, anchor="w")
    tree_haber.column("monto", width=120, anchor="e")
    tree_haber.pack(fill="x", pady=(0, 5))
    
    # Frame para agregar créditos
    frame_agregar_haber = ttk.Frame(haber_frame)
    frame_agregar_haber.pack(fill="x", pady=(5, 0))
    
    # Botón para seleccionar cuenta crédito
    btn_seleccionar_haber = ttk.Button(
        frame_agregar_haber,
        text="Seleccionar Cuenta",
        command=configurar_cuenta_haber,
        bootstyle=(OUTLINE, DANGER),
        width=15
    )
    btn_seleccionar_haber.pack(side="left", padx=2)
    
    # Etiqueta que muestra la cuenta seleccionada
    ttk.Label(
        frame_agregar_haber,
        text="Cuenta:"
    ).pack(side="left", padx=5)
    
    ttk.Label(
        frame_agregar_haber,
        textvariable=cuenta_haber_var,
        width=30,
        relief="solid",
        padding=5
    ).pack(side="left", fill="x", expand=True, padx=5)
    
    # Botones de acción para crédito
    frame_botones_haber = ttk.Frame(haber_frame)
    frame_botones_haber.pack(fill="x", pady=(5, 0))
    
    ttk.Button(
        frame_botones_haber,
        text="Agregar Crédito",
        command=agregar_cuenta_haber,
        bootstyle=DANGER,
        width=15
    ).pack(side="left", padx=2)
    
    ttk.Button(
        frame_botones_haber,
        text="Borrar",
        command=eliminar_item_haber,
        bootstyle=(OUTLINE, DANGER),
        width=8
    ).pack(side="left", padx=2)
    
    # Total créditos
    global lbl_total_haber
    lbl_total_haber = ttk.Label(haber_frame, text="Total Haber: S/. 0.00", font=('Segoe UI', 9, 'bold'))
    lbl_total_haber.pack(anchor="e", pady=(5, 0))
    
    # ========== SECCIÓN DE MONTO ==========
    campos_frame = ttk.Frame(formulario_asientos)
    campos_frame.grid(row=2, column=0, sticky="ew", pady=10)
    campos_frame.grid_columnconfigure(1, weight=1)

    # Monto con moneda fija en Soles
    ttk.Label(campos_frame, text="Monto:").grid(row=0, column=0, sticky="w", pady=5)
    
    # Frame para monto y símbolo de moneda
    monto_frame = ttk.Frame(campos_frame)
    monto_frame.grid(row=0, column=1, sticky="ew")
    
    # Símbolo de moneda fijo
    ttk.Label(monto_frame, text="S/.", font=('Segoe UI', 10)).pack(side="left")
    
    # Entry para monto
    validacion = root.register(validar_monto)
    monto_entry = ttk.Entry(
        monto_frame,
        validate="key",
        validatecommand=(validacion, '%P')
    )
    monto_entry.pack(side="left", fill="x", expand=True, padx=5)
    monto_entry.bind("<KeyRelease>", formatear_monto_durante_escritura)

    # Separador visual
    ttk.Separator(campos_frame, orient="horizontal").grid(row=1, column=0, columnspan=3, sticky="ew", pady=10)

    # ===== SECCIÓN COSTOS Y ACTIVIDAD =====
    # Configuración de variables
    global costos_var, actividad_var
    costos_var = tk.StringVar()
    actividad_var = tk.StringVar()

    # 1. Combobox de Costos
    ttk.Label(campos_frame, text="Costos:").grid(row=2, column=0, sticky="w", pady=5)
    costos_combobox = ttk.Combobox(
        campos_frame,
        textvariable=costos_var,
        values=["Productivos", "Administrativos", "Ventas"],
        state="readonly"
    )
    costos_combobox.grid(row=2, column=1, sticky="ew", pady=5)
    costos_combobox.current(0)

    # 2. Combobox de Actividad
    ttk.Label(campos_frame, text="Actividad:").grid(row=3, column=0, sticky="w", pady=5)
    actividad_combobox = ttk.Combobox(
        campos_frame,
        textvariable=actividad_var,
        values=["Operación", "Inversión", "Financiamiento"],
        state="readonly"
    )
    actividad_combobox.grid(row=3, column=1, sticky="ew", pady=5)
    actividad_combobox.current(0)

    # Glosa
    ttk.Label(campos_frame, text="Glosa:").grid(row=4, column=0, sticky="w", pady=5)
    glosa_entry = ttk.Entry(campos_frame)
    glosa_entry.grid(row=4, column=1, columnspan=2, sticky="ew", pady=5)
    
    # Separador
    ttk.Separator(formulario_asientos, orient="horizontal").grid(row=4, column=0, sticky="ew", pady=10)
    
    # ========== BOTONES DE ACCIÓN ==========
    frame_botones = ttk.Frame(formulario_asientos)
    frame_botones.grid(row=5, column=0, pady=(0, 15), sticky="ew")
    frame_botones.grid_columnconfigure(0, weight=1)
    
    # Frame para botones de acción
    frame_botones_accion = ttk.Frame(frame_botones)
    frame_botones_accion.grid(row=0, column=0)
    
    # Botón principal de Registrar
    btn_registrar = ttk.Button(
        frame_botones_accion,
        text="Registrar Operación",
        bootstyle=SUCCESS,
        command=mostrar_botones_confirmacion,
        width=20
    )
    btn_registrar.pack(side="left", padx=5)
    
    # Frame para botones de confirmación (inicialmente oculto)
    frame_confirmacion = ttk.Frame(frame_botones)
    
    # Botón Confirmar
    btn_confirmar = ttk.Button(
        frame_confirmacion,
        text="Confirmar",
        bootstyle=(SUCCESS, OUTLINE),
        command=confirmar_registro,
        width=10
    )
    btn_confirmar.pack(side="left", padx=5)
    
    # Botón Cancelar
    btn_cancelar = ttk.Button(
        frame_confirmacion,
        text="Cancelar",
        bootstyle=(DANGER, OUTLINE),
        command=ocultar_botones_confirmacion,
        width=10
    )
    btn_cancelar.pack(side="left", padx=5)
    
    # Asegurar que el formulario se expanda correctamente
    formulario_asientos.grid_columnconfigure(0, weight=1)
    
    return contenedor_scroll

def ajustar_altura_treeview(tree, items):
    """Ajusta dinámicamente la altura del treeview basado en los items"""
    # Calcular nueva altura (mínimo 1, máximo 8 filas visibles)
    nueva_altura = min(max(len(items), 1), 8)
    
    # Configurar nueva altura
    tree.configure(height=nueva_altura)
    
    # Manejar scrollbar solo si es necesario
    for widget in tree.master.winfo_children():
        if isinstance(widget, ttk.Scrollbar):
            widget.destroy()
    
    if len(items) > 5:  # Mostrar scrollbar si hay más de 5 elementos
        scrollbar = ttk.Scrollbar(tree.master, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)

def mostrar_ajustes():
    # Limpiar el main_frame
    for widget in main_frame.winfo_children():
        widget.destroy()
    
    # Frame principal
    frame_principal = ttk.Frame(main_frame)
    frame_principal.pack(fill="both", expand=True, padx=40, pady=20)
    
    # Título
    ttk.Label(
        frame_principal,
        text="AJUSTES DEL SISTEMA",
        style='Titulo.TLabel',
        font=('Helvetica', 20, 'bold')
    ).pack(pady=(10, 30))
    
    # Botón Histórico Inicial con estilo consistente
    btn_historico = ttk.Button(
        frame_principal,
        text="Histórico Inicial",
        bootstyle=(OUTLINE, PRIMARY),
        style='Boton.TButton',
        width=15,
        command=mostrar_historico_inicial
    )
    btn_historico.pack(pady=15)
    
    # Botón regresar
    ttk.Button(
        frame_principal,
        text="Regresar",
        bootstyle=(OUTLINE, SECONDARY),
        command=lambda: mostrar_seccion("inicio")
    ).pack(pady=10)

def mostrar_seleccion_cuenta_historico(parent, entry_widget):
    def seleccionar_cuenta():
        item = tree.focus()
        if item:
            valores = tree.item(item, "values")
            if valores and len(valores) >= 2 and len(valores[0]) == 2:
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, valores[0])
                seleccion_window.destroy()
            else:
                messagebox.showerror("Error", "Seleccione una cuenta principal de 2 dígitos")

    # Ventana de selección
    seleccion_window = tk.Toplevel(parent)
    seleccion_window.title("Seleccionar Cuenta Principal")
    seleccion_window.geometry("900x600")
    
    # Frame principal
    main_frame = ttk.Frame(seleccion_window)
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    # Treeview
    tree_frame = ttk.Frame(main_frame)
    tree_frame.pack(fill="both", expand=True)
    
    tree = ttk.Treeview(
        tree_frame,
        columns=("codigo", "nombre"),
        show="headings",
        selectmode="browse"
    )
    tree.heading("codigo", text="Código (2 dígitos)", anchor="center")
    tree.heading("nombre", text="Nombre de Cuenta", anchor="w")
    tree.column("codigo", width=150, anchor="center")
    tree.column("nombre", width=700, anchor="w")
    
    # Cargar solo cuentas principales (2 dígitos)
    arbol_cuentas = crear_arbol_cuentas()
    for codigo, datos in arbol_cuentas.items():
        if len(codigo) == 2:
            tree.insert("", "end", values=(codigo, datos["nombre"]))
    
    # Scrollbar
    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    scrollbar.pack(side="right", fill="y")
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(fill="both", expand=True)
    
    # Botones
    btn_frame = ttk.Frame(main_frame)
    btn_frame.pack(pady=10)
    
    ttk.Button(
        btn_frame,
        text="Seleccionar",
        bootstyle=SUCCESS,
        command=seleccionar_cuenta
    ).pack(side="left", padx=5)
    
    ttk.Button(
        btn_frame,
        text="Cancelar",
        bootstyle=DANGER,
        command=seleccion_window.destroy
    ).pack(side="left", padx=5)
    
    tree.bind("<Double-1>", lambda e: seleccionar_cuenta())

def guardar_historico():
    global historico_debe_items, historico_haber_items
    
    try:
        # 1. Validar que los totales coincidan
        total_debe = sum(item['monto'] for item in historico_debe_items)
        total_haber = sum(item['monto'] for item in historico_haber_items)
        
        if abs(total_debe - total_haber) > 0.01:
            messagebox.showerror("Error", "Los totales no coinciden")
            return
        
        conn = sqlite3.connect('data/contabilidad.db')
        cursor = conn.cursor()
        
        # 2. Eliminar TODOS los históricos anteriores
        cursor.execute("DELETE FROM operaciones WHERE glosa = 'Saldo inicial'")
        
        # 3. Configurar fecha (primer día del mes actual)
        primer_dia_mes_actual = datetime.now().replace(day=1)
        fecha_historico = primer_dia_mes_actual.strftime("%d/%m/%Y") + " 00:00"
        
        # 4. Registrar nuevos históricos (DEBE)
        for item in historico_debe_items:
            cuenta_codigo = item['cuenta'].split()[0]  # Extraer solo el código
            nombre_completo = obtener_nombre_cuenta(cuenta_codigo)
            cuenta_completa = f"{cuenta_codigo} {nombre_completo}"
            
            cursor.execute('''
                INSERT INTO operaciones 
                (operacion_id, fecha, cuenta_debe, cuenta_haber, monto, moneda, costos, actividad, glosa)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                0,  # ID FIJO 0 para todos los históricos
                fecha_historico,
                cuenta_completa,  # Ej: "10 Efectivo y equivalentes de efectivo"
                "",  # Cuenta haber vacía
                item['monto'],
                "S/",  # Moneda en soles
                "Ajuste Inicial",
                "Ajuste",
                "Saldo inicial"  # Glosa identificadora
            ))
        
        # 5. Registrar nuevos históricos (HABER)
        for item in historico_haber_items:
            cuenta_codigo = item['cuenta'].split()[0]  # Extraer solo el código
            nombre_completo = obtener_nombre_cuenta(cuenta_codigo)
            cuenta_completa = f"{cuenta_codigo} {nombre_completo}"
            
            cursor.execute('''
                INSERT INTO operaciones 
                (operacion_id, fecha, cuenta_debe, cuenta_haber, monto, moneda, costos, actividad, glosa)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                0,  # ID FIJO 0 para todos los históricos
                fecha_historico,
                "",  # Cuenta debe vacía
                cuenta_completa,  # Ej: "10 Efectivo y equivalentes de efectivo"
                item['monto'],
                "S/",  # Moneda en soles
                "Ajuste Inicial",
                "Ajuste",
                "Saldo inicial"  # Glosa identificadora
            ))
        
        conn.commit()
        conn.close()
        
        # 6. Limpiar y actualizar interfaz
        historico_debe_items.clear()
        historico_haber_items.clear()
        mostrar_ajustes()
        
        messagebox.showinfo("Éxito", "Histórico inicial guardado correctamente\n"
                                f"Fecha: {primer_dia_mes_actual.strftime('%B %Y')}")
        
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el histórico: {str(e)}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()

def mostrar_historico_inicial():
    global historico_debe_items, historico_haber_items, historico_tree, btn_frame, frame_confirmacion, btn_confirmar, btn_regresar
    
    # Limpiar el main_frame
    for widget in main_frame.winfo_children():
        widget.destroy()
    
    historico_debe_items = []
    historico_haber_items = []
    
    # Frame principal
    frame_principal = ttk.Frame(main_frame)
    frame_principal.pack(fill="both", expand=True, padx=20, pady=20)
    
    # Título
    ttk.Label(
        frame_principal,
        text="HISTÓRICO INICIAL - SALDOS DE CUENTAS",
        style='Titulo.TLabel',
        font=('Helvetica', 16, 'bold')
    ).pack(pady=(0, 20))
    
    # Formulario
    form_frame = ttk.Frame(frame_principal)
    form_frame.pack(fill="x", pady=10)
    
    # Cuenta con selector
    ttk.Label(form_frame, text="Cuenta (2 dígitos):").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    cuenta_var = tk.StringVar()
    cuenta_entry = ttk.Entry(form_frame, textvariable=cuenta_var, width=10)
    cuenta_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
    
    ttk.Button(
        form_frame,
        text="Seleccionar Cuenta",
        bootstyle=(OUTLINE, INFO),
        command=lambda: mostrar_seleccion_cuenta_historico(frame_principal, cuenta_entry),
        width=15
    ).grid(row=0, column=2, padx=5, sticky="w")
    
    # Tipo (Debe/Haber)
    ttk.Label(form_frame, text="Tipo:").grid(row=0, column=3, padx=5, pady=5, sticky="w")
    tipo_var = tk.StringVar(value="Debe")
    ttk.Radiobutton(form_frame, text="Debe", variable=tipo_var, value="Debe").grid(row=0, column=4, padx=5, sticky="w")
    ttk.Radiobutton(form_frame, text="Haber", variable=tipo_var, value="Haber").grid(row=0, column=5, padx=5, sticky="w")
    
    # Monto
    ttk.Label(form_frame, text="Monto:").grid(row=0, column=6, padx=5, pady=5, sticky="w")
    monto_var = tk.StringVar()
    monto_entry = ttk.Entry(form_frame, textvariable=monto_var, width=15)
    monto_entry.grid(row=0, column=7, padx=5, pady=5, sticky="w")
    
    # Validación de monto
    validacion = frame_principal.register(validar_monto)
    monto_entry.config(validate="key", validatecommand=(validacion, '%P'))
    
    # Función para agregar
    def agregar_historico():
        cuenta = cuenta_var.get()
        tipo = tipo_var.get()
        monto = monto_var.get().replace(",", "")
        
        if len(cuenta) != 2:
            messagebox.showerror("Error", "La cuenta debe tener exactamente 2 dígitos")
            return

        # Validar que la cuenta exista
        if obtener_nombre_cuenta(cuenta) == "Cuenta no encontrada":
            messagebox.showerror("Error", "La cuenta no existe en el plan contable")
            return

        if not monto:
            messagebox.showerror("Error", "Ingrese un monto válido")
            return
            
        try:
            monto_float = float(monto)
            if monto_float <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "Monto debe ser un número positivo")
            return
            
        # Eliminar si ya existe
        historico_debe_items[:] = [item for item in historico_debe_items if item['cuenta'] != cuenta]
        historico_haber_items[:] = [item for item in historico_haber_items if item['cuenta'] != cuenta]
        
        # Agregar nuevo
        nuevo_item = {'cuenta': cuenta, 'monto': monto_float}
        if tipo == "Debe":
            historico_debe_items.append(nuevo_item)
        else:
            historico_haber_items.append(nuevo_item)
            
        actualizar_treeview()
        cuenta_var.set("")
        monto_var.set("")
        cuenta_entry.focus()
    
    ttk.Button(
        form_frame,
        text="Agregar",
        bootstyle=INFO,
        command=agregar_historico
    ).grid(row=0, column=8, padx=10)
    
    # Treeview
    tree_frame = ttk.Frame(frame_principal)
    tree_frame.pack(fill="both", expand=True, pady=10)
    
    historico_tree = ttk.Treeview(
        tree_frame,
        columns=("cuenta", "tipo", "monto"),
        show="headings",
        height=8
    )
    historico_tree.heading("cuenta", text="Cuenta y Nombre")
    historico_tree.heading("tipo", text="Tipo")
    historico_tree.heading("monto", text="Monto (S/.)")
    historico_tree.column("cuenta", width=300, anchor="w")
    historico_tree.column("tipo", width=100, anchor="center")
    historico_tree.column("monto", width=150, anchor="e")
    
    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=historico_tree.yview)
    scrollbar.pack(side="right", fill="y")
    historico_tree.configure(yscrollcommand=scrollbar.set)
    historico_tree.pack(fill="both", expand=True)
    
    def actualizar_treeview():
        historico_tree.delete(*historico_tree.get_children())
        
        # Limpiar frame de totales si existe
        for widget in frame_principal.winfo_children():
            if isinstance(widget, ttk.Frame) and hasattr(widget, 'es_totales'):
                widget.destroy()
        
        # Insertar items
        for item in historico_debe_items:
            nombre_cuenta = obtener_nombre_cuenta(item['cuenta'])
            historico_tree.insert("", "end", values=(
                f"{item['cuenta']} - {nombre_cuenta}",
                "Debe", 
                f"S/. {item['monto']:,.2f}"
            ))
            
        for item in historico_haber_items:
            nombre_cuenta = obtener_nombre_cuenta(item['cuenta'])
            historico_tree.insert("", "end", values=(
                f"{item['cuenta']} - {nombre_cuenta}",
                "Haber", 
                f"S/. {item['monto']:,.2f}"
            ))
        
        # Crear frame de totales (solo uno)
        totales_frame = ttk.Frame(frame_principal)
        totales_frame.pack(fill="x", pady=(5, 10))
        totales_frame.es_totales = True
        
        total_debe = sum(item['monto'] for item in historico_debe_items)
        total_haber = sum(item['monto'] for item in historico_haber_items)
        
        ttk.Label(totales_frame, text=f"Total Debe: S/. {total_debe:,.2f}", 
                font=('Segoe UI', 10, 'bold')).pack(side="left", padx=10)
        ttk.Label(totales_frame, text=f"Total Haber: S/. {total_haber:,.2f}", 
                font=('Segoe UI', 10, 'bold')).pack(side="left", padx=10)
        
        if abs(total_debe - total_haber) > 0.01:
            ttk.Label(totales_frame, text=f"Diferencia: S/. {abs(total_debe - total_haber):,.2f}", 
                    bootstyle=DANGER, font=('Segoe UI', 10, 'bold')).pack(side="left", padx=10)

    # Botones finales - IMPORTANTE: Esta es la sección clave que estaba fallando
    btn_frame = ttk.Frame(frame_principal)
    btn_frame.pack(pady=(10, 0))
    
    # Botón Confirmar (PRINCIPAL) - debe ser declarado como variable global
    btn_confirmar = ttk.Button(
        btn_frame,
        text="Confirmar",
        bootstyle=SUCCESS,
        command=lambda: mostrar_confirmacion_historico()
    )
    btn_confirmar.pack(side="left", padx=5)
    
    # Botón Regresar
    btn_regresar = ttk.Button(
        btn_frame,
        text="Regresar a Ajustes",
        bootstyle=(OUTLINE, SECONDARY),
        command=mostrar_ajustes
    )
    btn_regresar.pack(side="left", padx=5)
    
    # Frame de confirmación (inicialmente oculto)
    frame_confirmacion = ttk.Frame(btn_frame)
    
    # Botón Aceptar
    btn_aceptar = ttk.Button(
        frame_confirmacion,
        text="Aceptar",
        bootstyle=(SUCCESS, OUTLINE),
        command=lambda: guardar_historico(),
        width=10
    )
    btn_aceptar.pack(side="left", padx=5)
    
    # Botón Cancelar
    btn_cancelar = ttk.Button(
        frame_confirmacion,
        text="Cancelar",
        bootstyle=(DANGER, OUTLINE),
        command=lambda: ocultar_confirmacion_historico(),
        width=10
    )
    btn_cancelar.pack(side="left", padx=5)

    # Función para mostrar confirmación
    def mostrar_confirmacion_historico():
        total_debe = sum(item['monto'] for item in historico_debe_items)
        total_haber = sum(item['monto'] for item in historico_haber_items)
        
        if abs(total_debe - total_haber) > 0.01:
            messagebox.showerror("Error", f"Debe nivelar los totales\nDiferencia: S/. {abs(total_debe - total_haber):,.2f}")
            return
        
        # Ocultar botones principales
        btn_confirmar.pack_forget()
        btn_regresar.pack_forget()
        
        # Mostrar botones de confirmación
        frame_confirmacion.pack(side="left", padx=5)

    # Función para ocultar confirmación
    def ocultar_confirmacion_historico():
        # Ocultar confirmación
        frame_confirmacion.pack_forget()
        
        # Mostrar botones principales nuevamente
        btn_confirmar.pack(side="left", padx=5)
        btn_regresar.pack(side="left", padx=5)

def mostrar_confirmacion_historico():
    global btn_frame, frame_confirmacion
    
    total_debe = sum(item['monto'] for item in historico_debe_items)
    total_haber = sum(item['monto'] for item in historico_haber_items)
    
    if abs(total_debe - total_haber) > 0.01:
        messagebox.showerror("Error", f"Debe nivelar los totales\nDiferencia: S/. {abs(total_debe - total_haber):,.2f}")
        return
    
    # Ocultar botón Confirmar
    for widget in btn_frame.winfo_children():
        if isinstance(widget, ttk.Button) and widget.cget("text") == "Confirmar":
            widget.pack_forget()
    
    # Crear frame de confirmación si no existe
    if 'frame_confirmacion' not in globals():
        frame_confirmacion = ttk.Frame(btn_frame)
        
        ttk.Button(
            frame_confirmacion,
            text="Aceptar",
            bootstyle=(SUCCESS, OUTLINE),
            command=guardar_historico,
            width=10
        ).pack(side="left", padx=5)
        
        ttk.Button(
            frame_confirmacion,
            text="Cancelar",
            bootstyle=(DANGER, OUTLINE),
            command=ocultar_confirmacion_historico,
            width=10
        ).pack(side="left", padx=5)
    
    # Mostrar confirmación
    frame_confirmacion.pack(side="left", padx=5)

def ocultar_confirmacion_historico():
    global btn_frame, frame_confirmacion
    
    # Ocultar confirmación
    frame_confirmacion.pack_forget()
    
    # Mostrar botón Confirmar
    for widget in btn_frame.winfo_children():
        if isinstance(widget, ttk.Button) and widget.cget("text") == "Regresar a Ajustes":
            ttk.Button(
                btn_frame,
                text="Confirmar",
                bootstyle=SUCCESS,
                command=mostrar_confirmacion_historico
            ).pack(side="left", padx=5)
            break

# --- Barra superior premium con grid ---
barra_superior = ttk.Frame(root, bootstyle="dark")
barra_superior.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=(0, 5))
barra_superior.grid_columnconfigure(1, weight=1)

frame_titulo = ttk.Frame(barra_superior, bootstyle="dark")
frame_titulo.grid(row=0, column=0, padx=20, sticky="w")

ttk.Label(
    frame_titulo,
    text="REGISTRO DE CUENTAS",
    style='Titulo.TLabel',
    bootstyle="inverse-dark"
).pack(side="left")

secciones_nav = ["Operaciones", "Libros", "Estados Financieros", "Costos", "Tablas", "Reportes", "Ajustes"]
nav_container = ttk.Frame(barra_superior)
nav_container.grid(row=0, column=1, sticky="ew", padx=10)  # Cambiado a "ew" para expansión

for i, seccion in enumerate(secciones_nav):
    nav_container.grid_columnconfigure(i, weight=1, uniform="nav_btns")
    btn = ttk.Button(
        nav_container,
        text=seccion,
        bootstyle=(OUTLINE, PRIMARY),
        style='Boton.TButton',
        width=10,  # Puedes ajustar este número
        command=lambda s=seccion.lower(): mostrar_seccion(s)
    )
    btn.grid(row=0, column=i, sticky="ew", padx=4)  # Mismo ancho para todas las columnas



# --- Contenedor principal con grid ---
contenedor_principal = ttk.Frame(root)
contenedor_principal.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)
contenedor_principal.grid_columnconfigure(1, weight=1)
contenedor_principal.grid_rowconfigure(0, weight=1)

# --- Formulario de asientos contables con grid ---
formulario_asientos = ttk.Frame(contenedor_principal)
formulario_asientos.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
formulario_asientos.grid_rowconfigure(8, weight=1)


#Campos adicionales
campos_frame = ttk.Frame(formulario_asientos)


# Monto y Moneda
ttk.Label(campos_frame, text="Monto:").grid(row=0, column=0, sticky="w", pady=5)

validacion = root.register(validar_monto)
monto_entry = ttk.Entry(
    campos_frame,
    validate="key",
    validatecommand=(validacion, '%P')
)
monto_entry.grid(row=0, column=1, sticky="ew", pady=5)
monto_entry.bind("<KeyRelease>", formatear_monto_durante_escritura)

moneda_var = tk.StringVar(value="S/")
moneda_frame = ttk.Frame(campos_frame)
moneda_frame.grid(row=0, column=2, sticky="ew", padx=10)
ttk.Radiobutton(moneda_frame, text="Soles (S/)", variable=moneda_var, value="S/").pack(side="left")
ttk.Radiobutton(moneda_frame, text="Dólares ($)", variable=moneda_var, value="$").pack(side="left", padx=10)

# Costos
ttk.Label(campos_frame, text="Costos:").grid(row=1, column=0, sticky="w", pady=5)
costos_var = tk.StringVar()
costos_combobox = ttk.Combobox(
    campos_frame,
    textvariable=costos_var,
    values=["Productivos", "Administrativos", "Ventas"],
    state="readonly"
)
costos_combobox.grid(row=1, column=1, columnspan=2, sticky="ew", pady=5)

# --- Área principal de contenido con grid ---
main_frame = ttk.Frame(contenedor_principal)
main_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
main_frame.grid_columnconfigure(0, weight=1)
main_frame.grid_rowconfigure(0, weight=1)

def obtener_nombre_cuenta(codigo):
    arbol = crear_arbol_cuentas()
    
    # Buscar en cuentas principales
    for cuenta_principal, datos in arbol.items():
        if cuenta_principal == codigo:
            return datos['nombre']
        
        # Buscar en subcuentas si existen
        if 'subcuentas' in datos:
            for subcuenta, subdatos in datos['subcuentas'].items():
                if subcuenta == codigo:
                    return subdatos['nombre']
    
    return "Cuenta no encontrada"
# --- Funcionalidad principal ---
def mostrar_seccion(seccion):
    global panel_inicio, formulario_asientos, tree_operaciones
    
    # Limpiar el main_frame
    for widget in main_frame.winfo_children():
        widget.destroy()
    
    # Ocultar formulario_asientos si no estamos en operaciones
    if formulario_asientos:
        formulario_asientos.grid_forget()
    
    if seccion == "operaciones":
        # Limpiar listas al cambiar a sección de operaciones
        debe_items.clear()
        haber_items.clear()
        
        # Mostrar formulario de asientos
        formulario_asientos.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        
        # Cargar operaciones en el treeview
        mostrar_operaciones()

    elif seccion == "ajustes":
        mostrar_ajustes()  # Añade este caso
    elif seccion == "libros":
        mostrar_libros()
    elif seccion == "estados financieros":
        mostrar_estados_financieros()
    elif seccion == "inicio":
        panel_inicio = ttk.Frame(main_frame)
        panel_inicio.grid(row=0, column=0, sticky="nsew")
        panel_inicio.grid_columnconfigure(0, weight=1)
        panel_inicio.grid_rowconfigure(1, weight=1)
        
        ttk.Label(
            panel_inicio,
            text="SISTEMA DE CONTABILIDAD DE COSTOS",
            style='Titulo.TLabel'
        ).grid(row=0, column=0, pady=(50, 20))
        
        ttk.Label(
            panel_inicio,
            text="Bienvenido al sistema premium de gestión contable",
            style='Subtitulo.TLabel'
        ).grid(row=1, column=0, pady=(0, 40))
    elif seccion == "reportes":
        mostrar_reportes()   

def crear_combobox_confiables():
    global costos_var, actividad_var, costos_combobox, actividad_combobox
    
    # Frame para campos adicionales
    campos_frame = ttk.Frame(formulario_asientos)
    campos_frame.grid(row=2, column=0, sticky="ew", pady=10)
    campos_frame.grid_columnconfigure(1, weight=1)
    
    # Etiquetas
    ttk.Label(campos_frame, text="Costos:").grid(row=1, column=0, sticky="w", pady=5)
    ttk.Label(campos_frame, text="Actividad:").grid(row=2, column=0, sticky="w", pady=5)
    
    # Configuración de variables
    costos_var = tk.StringVar(value="")  # Valor inicial vacío
    actividad_var = tk.StringVar(value="")  # Valor inicial vacío
    
    # Combobox de Costos
    costos_combobox = ttk.Combobox(
        campos_frame,
        textvariable=costos_var,
        values=["Productivos", "Administrativos", "Ventas"],
        state="readonly"
    )
    costos_combobox.grid(row=1, column=1, columnspan=2, sticky="ew", pady=5)
    
    # Combobox de Actividad
    actividad_combobox = ttk.Combobox(
        campos_frame,
        textvariable=actividad_var,
        values=["Operación", "Inversión", "Financiamiento"],
        state="readonly"
    )
    actividad_combobox.grid(row=2, column=1, columnspan=2, sticky="ew", pady=5)
    
    # Establecer valores por defecto
    costos_combobox.set("Productivos")  # Valor por defecto
    actividad_combobox.set("Operación")  # Valor por defecto
    
    # Configurar eventos
    def on_combobox_select(event):
        widget = event.widget
        print(f"{'Costos' if widget == costos_combobox else 'Actividad'} seleccionado: {widget.get()}")
    
    costos_combobox.bind("<<ComboboxSelected>>", on_combobox_select)
    actividad_combobox.bind("<<ComboboxSelected>>", on_combobox_select)

def mostrar_libros():
    # Limpiar el main_frame
    for widget in main_frame.winfo_children():
        widget.destroy()
    
    frame_libros = ttk.Frame(main_frame)
    frame_libros.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
    frame_libros.grid_columnconfigure(0, weight=1)
    frame_libros.grid_rowconfigure(1, weight=1)
    
    # Título
    ttk.Label(
        frame_libros,
        text="LIBROS CONTABLES",
        style='Titulo.TLabel'
    ).grid(row=0, column=0, pady=(0, 10))
    
    # Contenedor de botones de libros
    libros_frame = ttk.Frame(frame_libros)
    libros_frame.grid(row=1, column=0, sticky="nsew")
    libros_frame.grid_columnconfigure(0, weight=1, uniform="libros_col")
    libros_frame.grid_columnconfigure(1, weight=1, uniform="libros_col")
    
    libros = [
        ("Libro Diario", "📒", mostrar_libro_diario),
        ("Libro Mayor", "📊", mostrar_libro_mayor),
        ("Libro de Inventarios y Balances", "⚖️", lambda: mostrar_libro("Libro de Inventarios y Balances")),
        ("Libro de Compras", "🛒", lambda: mostrar_libro("Libro de Compras")),
        ("Libro de Ventas", "💰", lambda: mostrar_libro("Libro de Ventas")),
        ("Libro de Cajas", "💵", lambda: mostrar_libro("Libro de Cajas")),
        ("Libro de Planillas", "👥", lambda: mostrar_libro("Libro de Planillas"))
    ]
    
    for i, (nombre, icono, comando) in enumerate(libros):
        col = i % 2
        row = i // 2
        
        estilo = (OUTLINE, SUCCESS) if nombre == "Libro Mayor" else (OUTLINE, INFO)
        
        btn = ttk.Button(
            libros_frame,
            text=f"{icono}  {nombre}",
            style='Libros.TButton',
            bootstyle=estilo,
            width=20,
            command=comando
        )
        btn.grid(row=row, column=col, pady=5, padx=10, sticky="ew")
    
    # Botón regresar
    ttk.Button(
        frame_libros,
        text="Regresar",
        bootstyle=(OUTLINE, SECONDARY),
        command=lambda: mostrar_seccion("inicio")
    ).grid(row=2, column=0, pady=10)

def mostrar_libro(nombre_libro):
    if nombre_libro == "Libro Diario":
        mostrar_libro_diario()
    else:
        for widget in main_frame.winfo_children():
            widget.destroy()
        
        frame_libro = ttk.Frame(main_frame)
        frame_libro.grid(row=0, column=0, sticky="nsew", padx=20, pady=10)
        frame_libro.grid_columnconfigure(0, weight=1)
        frame_libro.grid_rowconfigure(1, weight=1)
        
        ttk.Label(
            frame_libro,
            text=nombre_libro.upper(),
            style='Titulo.TLabel'
        ).grid(row=0, column=0, pady=(0, 20))
        
        ttk.Label(
            frame_libro,
            text=f"Contenido del {nombre_libro} se mostrará aquí",
            font=('Segoe UI', 12)
        ).grid(row=1, column=0, pady=50)
        
        ttk.Button(
            frame_libro,
            text="Regresar a Libros",
            bootstyle=(OUTLINE, SECONDARY),
            command=lambda: mostrar_seccion("libros")
        ).grid(row=2, column=0, pady=20, sticky="e")

def mostrar_libro_diario(mes=None):
    global meses_disponibles, mes_actual, total_debe, total_haber, totales_frame
    
    # Si no se especifica mes, usar el actual
    mes = mes or mes_actual
    
    # Actualizar meses disponibles
    actualizar_meses_disponibles()
    
    # Limpiar el main_frame
    for widget in main_frame.winfo_children():
        widget.destroy()
    
    # Frame principal responsivo
    frame_libro = ttk.Frame(main_frame)
    frame_libro.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
    frame_libro.grid_columnconfigure(0, weight=1)
    frame_libro.grid_rowconfigure(3, weight=1)  # Hacer el treeview responsivo

    # Título centrado y responsivo
    title_frame = ttk.Frame(frame_libro)
    title_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
    title_frame.grid_columnconfigure(0, weight=1)
    
    ttk.Label(
        title_frame,
        text="LIBRO DIARIO",
        style='Titulo.TLabel'
    ).grid(row=0, column=0)

    # Frame para controles de mes responsivo
    meses_frame = ttk.Frame(frame_libro)
    meses_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
    meses_frame.grid_columnconfigure(1, weight=1)
    
    ttk.Label(meses_frame, text="Mes:", font=('Segoe UI', 10, 'bold')).grid(row=0, column=0, padx=(0, 5), sticky="w")
    
    # Contenedor scrollable para meses
    meses_container = ttk.Frame(meses_frame)
    meses_container.grid(row=0, column=1, sticky="nsew")
    meses_container.grid_columnconfigure(0, weight=1)
    
    canvas = tk.Canvas(meses_container, height=40, highlightthickness=0, bg='#333333')
    scrollbar = ttk.Scrollbar(meses_container, orient="horizontal", command=canvas.xview)
    scrollable_frame = ttk.Frame(canvas)
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all"),
            width=e.width if e.width > frame_libro.winfo_width()-50 else frame_libro.winfo_width()-50
        )
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(xscrollcommand=scrollbar.set)
    
    canvas.grid(row=0, column=0, sticky="ew")
    scrollbar.grid(row=1, column=0, sticky="ew")
    
    # Configurar grid para botones de meses
    scrollable_frame.grid_rowconfigure(0, weight=1)
    
    # Crear botones de meses (se añaden de izquierda a derecha)
    for i, mes_disponible in enumerate(meses_disponibles):
        estilo = 'primary' if mes_disponible == mes else 'secondary'
        
        btn = ttk.Button(
            scrollable_frame,
            text=datetime.strptime(mes_disponible, "%m/%Y").strftime("%B %Y").upper(),
            bootstyle=estilo,
            width=15,
            command=lambda m=mes_disponible: mostrar_libro_diario(m)
        )
        btn.grid(row=0, column=i, padx=3, pady=2, sticky="nsew")
    
    # Controles de búsqueda responsivos
    controles_frame = ttk.Frame(frame_libro)
    controles_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 10))
    controles_frame.grid_columnconfigure(1, weight=1)
    
    ttk.Label(controles_frame, text="Buscar:").grid(row=0, column=0, padx=5, sticky="w")
    
    busqueda_var = tk.StringVar()
    busqueda_entry = ttk.Entry(controles_frame, textvariable=busqueda_var)
    busqueda_entry.grid(row=0, column=1, padx=5, sticky="ew")
    
    ttk.Button(
        controles_frame,
        text="Buscar",
        bootstyle='info',
        command=lambda: filtrar_libro_diario(tree_libro, busqueda_var.get())
    ).grid(row=0, column=2, padx=5, sticky="ew")
    
    # Treeview responsivo
    tree_frame = ttk.Frame(frame_libro)
    tree_frame.grid(row=3, column=0, sticky="nsew", pady=(0, 10))
    tree_frame.grid_columnconfigure(0, weight=1)
    tree_frame.grid_rowconfigure(0, weight=1)
    
    columns = ("id", "fecha", "cuenta", "debe", "haber")
    tree_libro = ttk.Treeview(
        tree_frame,
        columns=columns,
        show="headings",
        height=15,
        selectmode="extended"
    )
    
    # Configurar columnas responsivas
    tree_libro.heading("id", text="ID", anchor="center")
    tree_libro.heading("fecha", text="Fecha", anchor="center")
    tree_libro.heading("cuenta", text="Cuenta", anchor="w")
    tree_libro.heading("debe", text="Debe", anchor="e")
    tree_libro.heading("haber", text="Haber", anchor="e")
    
    tree_libro.column("id", width=50, anchor="center", stretch=False, minwidth=50)
    tree_libro.column("fecha", width=90, anchor="center", stretch=False, minwidth=90)
    tree_libro.column("cuenta", width=250, anchor="w", stretch=True, minwidth=150)
    tree_libro.column("debe", width=120, anchor="e", stretch=True, minwidth=80)
    tree_libro.column("haber", width=120, anchor="e", stretch=True, minwidth=80)
    
    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree_libro.yview)
    scrollbar.grid(row=0, column=1, sticky="ns")
    tree_libro.configure(yscrollcommand=scrollbar.set)
    tree_libro.grid(row=0, column=0, sticky="nsew")
    
    # Totales responsivos
    totales_frame = ttk.Frame(frame_libro)
    totales_frame.grid(row=4, column=0, sticky="nsew", pady=(10, 0))
    totales_frame.grid_columnconfigure(1, weight=1)
    
    ttk.Label(totales_frame, text="Totales:", font=('Segoe UI', 10, 'bold')).grid(row=0, column=0, padx=5, sticky="w")
    
    total_debe = ttk.Label(totales_frame, text="Debe: S/. 0.00", font=('Segoe UI', 10, 'bold'))
    total_debe.grid(row=0, column=1, padx=20, sticky="w")
    
    total_haber = ttk.Label(totales_frame, text="Haber: S/. 0.00", font=('Segoe UI', 10, 'bold'))
    total_haber.grid(row=0, column=2, padx=20, sticky="w")
    
    # Botón regresar centrado
    btn_frame = ttk.Frame(frame_libro)
    btn_frame.grid(row=5, column=0, pady=10, sticky="nsew")
    btn_frame.grid_columnconfigure(0, weight=1)
    
    ttk.Button(
        btn_frame,
        text="Regresar a Libros",
        bootstyle='secondary',
        command=lambda: mostrar_seccion("libros")
    ).grid(row=0, column=0)
    
    # Cargar datos del mes seleccionado
    filtrar_libro_diario_por_mes(tree_libro, mes)
    
    # Asegurar scroll al final
    canvas.xview_moveto(1.0)

def cargar_datos_libro_diario(tree):
    for item in tree.get_children():
        tree.delete(item)
    
    for operacion in operaciones_registradas:
        fecha = operacion["fecha"].split()[0]
        
        tree.insert("", "end", values=(
            fecha,
            operacion["debe"],
            operacion["monto"] if operacion["monto"] else "",
            ""
        ))
        
        tree.insert("", "end", values=(
            fecha,
            operacion["haber"],
            "",
            operacion["monto"] if operacion["monto"] else ""
        ))

def filtrar_libro_diario(tree, texto_busqueda):
    texto = texto_busqueda.lower()
    
    if not texto:
        for item in tree.get_children():
            tree.item(item, tags=())
        return
    
    for item in tree.get_children():
        valores = tree.item(item, "values")
        if (texto in valores[0].lower() or
            texto in valores[1].lower()):
            tree.item(item, tags=())
        else:
            tree.item(item, tags=("hidden",))
    
    tree.tag_configure("hidden", foreground="gray80")

def mostrar_libro_mayor(mes=None):
    global meses_disponibles, mes_actual
    
    # Establecer mes actual si no se especifica
    mes = mes or datetime.now().strftime("%m/%Y")
    
    # Actualizar meses disponibles sin borrar los anteriores
    actualizar_meses_disponibles()
    
    # Limpiar el frame principal
    for widget in main_frame.winfo_children():
        widget.destroy()
    
    # Configuración del frame principal con scroll
    frame_principal = ttk.Frame(main_frame)
    frame_principal.pack(fill="both", expand=True, padx=0, pady=10)
    
    main_canvas = tk.Canvas(frame_principal, highlightthickness=0)
    main_scrollbar = ttk.Scrollbar(frame_principal, orient="vertical", command=main_canvas.yview)
    scrollable_main_frame = ttk.Frame(main_canvas)
    
    scrollable_main_frame.bind(
        "<Configure>",
        lambda e: main_canvas.configure(
            scrollregion=main_canvas.bbox("all"),
            width=e.width
        )
    )
    
    main_canvas.create_window((0, 0), window=scrollable_main_frame, anchor="nw")
    main_canvas.configure(yscrollcommand=main_scrollbar.set)
    
    main_canvas.grid(row=0, column=0, sticky="nsew")
    main_scrollbar.grid(row=0, column=1, sticky="ns")
    frame_principal.grid_columnconfigure(0, weight=1)
    frame_principal.grid_rowconfigure(0, weight=1)
    
    # Título del libro mayor
    title_frame = ttk.Frame(scrollable_main_frame)
    title_frame.pack(fill="x", pady=(10, 20))
    
    ttk.Label(
        title_frame,
        text=f"LIBRO MAYOR - {datetime.strptime(mes, '%m/%Y').strftime('%B %Y').upper()}",
        style='Titulo.TLabel',
        font=('Helvetica', 16, 'bold')
    ).pack()
    
    # Controles de selección de mes
    meses_frame = ttk.Frame(scrollable_main_frame)
    meses_frame.pack(fill="x", pady=(0, 10), padx=10)
    
    ttk.Label(meses_frame, text="Mes:", font=('Segoe UI', 10, 'bold')).grid(row=0, column=0, padx=(0, 5), sticky="w")
    
    meses_container = ttk.Frame(meses_frame)
    meses_container.grid(row=0, column=1, sticky="nsew")
    
    canvas = tk.Canvas(meses_container, height=40, highlightthickness=0, bg='#333333')
    scrollbar = ttk.Scrollbar(meses_container, orient="horizontal", command=canvas.xview)
    scrollable_frame = ttk.Frame(canvas)
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all"),
            width=e.width if e.width > frame_principal.winfo_width()-50 else frame_principal.winfo_width()-50
        )
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(xscrollcommand=scrollbar.set)
    
    canvas.grid(row=0, column=0, sticky="ew")
    scrollbar.grid(row=1, column=0, sticky="ew")
    
    # Botones de meses disponibles (todos los meses registrados)
    for i, mes_disponible in enumerate(sorted(meses_disponibles, key=lambda x: datetime.strptime(x, "%m/%Y"), reverse=True)):
        estilo = 'primary' if mes_disponible == mes else 'secondary'
        
        btn = ttk.Button(
            scrollable_frame,
            text=datetime.strptime(mes_disponible, "%m/%Y").strftime("%B %Y").upper(),
            bootstyle=estilo,
            width=15,
            command=lambda m=mes_disponible: mostrar_libro_mayor(m)
        )
        btn.grid(row=0, column=i, padx=3, pady=2, sticky="nsew")
    
    # Conexión a la base de datos
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    # Obtener el mes y año actual
    mes_actual_num, año_actual = map(int, mes.split('/'))
    
    # Calcular mes anterior
    if mes_actual_num == 1:
        mes_anterior = 12
        año_anterior = año_actual - 1
    else:
        mes_anterior = mes_actual_num - 1
        año_anterior = año_actual
    
    mes_anterior_str = f"{mes_anterior:02d}/{año_anterior}"
    
    # Consulta para obtener saldos acumulados hasta el mes anterior (INCLUYENDO HISTORIAL INICIAL)
    cursor.execute('''
        SELECT 
            cuenta_principal,
            SUM(CASE WHEN tipo = 'debe' THEN monto ELSE 0 END) as total_debe,
            SUM(CASE WHEN tipo = 'haber' THEN monto ELSE 0 END) as total_haber,
            MAX(nombre_cuenta) as nombre_cuenta
        FROM (
            SELECT 
                substr(cuenta_debe, 1, 2) as cuenta_principal,
                cuenta_debe as nombre_cuenta,
                monto,
                'debe' as tipo,
                substr(fecha, 4, 7) as mes_operacion
            FROM operaciones
            WHERE (substr(fecha, 4, 7) < ? OR substr(fecha, 4, 7) = ?)
            AND cuenta_debe != ''
            
            UNION ALL
            
            SELECT 
                substr(cuenta_haber, 1, 2) as cuenta_principal,
                cuenta_haber as nombre_cuenta,
                monto,
                'haber' as tipo,
                substr(fecha, 4, 7) as mes_operacion
            FROM operaciones
            WHERE (substr(fecha, 4, 7) < ? OR substr(fecha, 4, 7) = ?)
            AND cuenta_haber != ''
        )
        GROUP BY cuenta_principal
    ''', (mes, mes_anterior_str, mes, mes_anterior_str))
    
    saldos_acumulados = cursor.fetchall()
    
    # Procesar los saldos acumulados hasta el mes anterior
    saldos_iniciales = {}
    for cuenta, total_debe, total_haber, nombre in saldos_acumulados:
        diferencia = total_debe - total_haber
        if abs(diferencia) > 0.01:  # Solo si hay saldo significativo
            nombre_cuenta = nombre.split(' - ')[1] if ' - ' in nombre else nombre
            saldos_iniciales[cuenta] = {
                'monto': abs(diferencia),
                'tipo': 'debe' if diferencia > 0 else 'haber',
                'nombre': nombre_cuenta,
                'saldo_acumulado': diferencia  # Guardamos el saldo real para cálculos
            }
    
    # Consulta para obtener operaciones del mes seleccionado
    cursor.execute('''
        SELECT 
            operacion_id,
            fecha,
            cuenta_debe,
            cuenta_haber,
            monto,
            glosa
        FROM operaciones
        WHERE substr(fecha, 4, 7) = ?
        ORDER BY 
            CASE WHEN glosa = 'Saldo inicial' THEN 0 ELSE 1 END,
            fecha ASC
    ''', (mes,))
    
    # Procesamiento de los datos
    cuentas_t = {}
    
    # Primero agregamos los saldos acumulados hasta el mes anterior
    for cuenta, datos in saldos_iniciales.items():
        if cuenta not in cuentas_t:
            cuentas_t[cuenta] = {
                'debe': [],
                'haber': [],
                'nombre': datos['nombre'],
                'codigo': cuenta,
                'saldo_acumulado': datos['saldo_acumulado']  # Guardamos el saldo acumulado
            }
        
        cuentas_t[cuenta][datos['tipo']].append({
            'id': "SI",
            'monto': datos['monto'],
            'nombre': f"{cuenta} - {datos['nombre']}",
            'fecha': datetime.strptime(f"01/{mes}", "%d/%m/%Y").strftime("%d/%m/%Y"),
            'glosa': f"Saldo acumulado hasta {datetime.strptime(mes_anterior_str, '%m/%Y').strftime('%B %Y')}",
            'es_historico': True,
            'fecha_orden': datetime.strptime(f"01/{mes}", "%d/%m/%Y")
        })
    
    # Luego procesamos las operaciones normales del mes actual
    for op in cursor.fetchall():
        operacion_id, fecha, cuenta_debe, cuenta_haber, monto, glosa = op
        cuenta = cuenta_debe if cuenta_debe else cuenta_haber
        tipo = 'debe' if cuenta_debe else 'haber'
        cuenta_principal = obtener_cuenta_principal(cuenta)
        
        if cuenta_principal not in cuentas_t:
            nombre_completo = cuenta.split(" - ")[1] if " - " in cuenta else cuenta
            cuentas_t[cuenta_principal] = {
                'debe': [],
                'haber': [],
                'nombre': nombre_completo,
                'codigo': cuenta_principal,
                'saldo_acumulado': 0  # Inicializamos saldo acumulado en 0 para nuevas cuentas
            }
        
        # Actualizar saldo acumulado
        if tipo == 'debe':
            cuentas_t[cuenta_principal]['saldo_acumulado'] += float(monto)
        else:
            cuentas_t[cuenta_principal]['saldo_acumulado'] -= float(monto)
        
        # Configurar visualización del ID
        id_display = "SI" if glosa == "Saldo inicial" else operacion_id
        
        cuentas_t[cuenta_principal][tipo].append({
            'id': id_display,
            'monto': float(monto),
            'nombre': cuenta,
            'fecha': fecha.split()[0],
            'glosa': glosa,
            'es_historico': glosa == "Saldo inicial",
            'fecha_orden': datetime.strptime(fecha.split()[0], "%d/%m/%Y") if fecha else datetime.min
        })
    
    conn.close()
    
    # Ordenar cuentas por código
    cuentas_ordenadas = sorted(cuentas_t.items(), key=lambda x: x[0])
    
    # Contenedor para las cuentas
    cuentas_container = ttk.Frame(scrollable_main_frame)
    cuentas_container.pack(fill="both", expand=True, padx=10)
    
    # Mostrar cada cuenta
    for codigo, datos in cuentas_ordenadas:
        # Solo mostrar si tiene movimientos
        if not datos['debe'] and not datos['haber']:
            continue
            
        frame_cuenta = ttk.LabelFrame(
            cuentas_container,
            text=f"CUENTA {codigo} - {datos['nombre'].upper()}",
            bootstyle="info",
            padding=10
        )
        frame_cuenta.pack(fill="x", pady=5)
        
        frame_contenedor = ttk.Frame(frame_cuenta)
        frame_contenedor.pack(fill="x")
        frame_contenedor.grid_columnconfigure(0, weight=1, uniform="cols")
        frame_contenedor.grid_columnconfigure(1, weight=1, uniform="cols")
        
        # Columna DEBE
        frame_debe = ttk.Frame(frame_contenedor)
        frame_debe.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        
        ttk.Label(
            frame_debe, 
            text="DEBE", 
            bootstyle="inverse-info", 
            font=('Segoe UI', 10, 'bold'),
            anchor="center"
        ).pack(fill="x", pady=(0, 5))
        
        debe_content_frame = ttk.Frame(frame_debe)
        debe_content_frame.pack(fill="x")
        
        total_debe = 0
        max_lineas_debe = 0
        
        # Ordenar: históricos primero, luego por fecha
        datos['debe'].sort(key=lambda x: (not x['es_historico'], x['fecha_orden']))
        
        if datos['debe']:
            for mov in datos['debe']:
                frame_mov = ttk.Frame(debe_content_frame)
                frame_mov.pack(fill="x", pady=2)
                
                ttk.Label(
                    frame_mov, 
                    text=f" {mov['id']} ",
                    bootstyle="inverse-dark",
                    font=('Segoe UI', 8, 'bold')
                ).pack(side="left", padx=(0, 5))
                
                ttk.Label(
                    frame_mov, 
                    text=mov['nombre'], 
                    anchor="w"
                ).pack(side="left", fill="x", expand=True)
                
                ttk.Label(
                    frame_mov, 
                    text=f"S/. {mov['monto']:,.2f}", 
                    anchor="e"
                ).pack(side="right")
                
                total_debe += mov['monto']
                max_lineas_debe += 1
                
                if mov['glosa']:
                    ttk.Label(
                        debe_content_frame,
                        text=f"Glosa: {mov['glosa']}",
                        font=('Segoe UI', 8),
                        foreground="#aaaaaa"
                    ).pack(anchor="w", padx=25, pady=(0, 2))
                    max_lineas_debe += 0.5
        else:
            ttk.Label(
                debe_content_frame, 
                text="No hay operaciones", 
                foreground="#888888"
            ).pack(pady=5)
        
        # Total DEBE
        ttk.Separator(frame_debe, orient="horizontal").pack(fill="x", pady=(5, 0))
        frame_total_debe = ttk.Frame(frame_debe)
        frame_total_debe.pack(fill="x")
        
        ttk.Label(
            frame_total_debe, 
            text="TOTAL DEBE:", 
            font=('Segoe UI', 9, 'bold')
        ).pack(side="left")
        
        ttk.Label(
            frame_total_debe, 
            text=f"S/. {total_debe:,.2f}", 
            font=('Segoe UI', 9, 'bold')
        ).pack(side="right")
        
        # Columna HABER
        frame_haber = ttk.Frame(frame_contenedor)
        frame_haber.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        
        ttk.Label(
            frame_haber, 
            text="HABER", 
            bootstyle="inverse-danger", 
            font=('Segoe UI', 10, 'bold'),
            anchor="center"
        ).pack(fill="x", pady=(0, 5))
        
        haber_content_frame = ttk.Frame(frame_haber)
        haber_content_frame.pack(fill="x")
        
        total_haber = 0
        max_lineas_haber = 0
        
        # Ordenar: históricos primero, luego por fecha
        datos['haber'].sort(key=lambda x: (not x['es_historico'], x['fecha_orden']))
        
        if datos['haber']:
            for mov in datos['haber']:
                frame_mov = ttk.Frame(haber_content_frame)
                frame_mov.pack(fill="x", pady=2)
                
                ttk.Label(
                    frame_mov, 
                    text=f" {mov['id']} ",
                    bootstyle="inverse-dark",
                    font=('Segoe UI', 8, 'bold')
                ).pack(side="left", padx=(0, 5))
                
                ttk.Label(
                    frame_mov, 
                    text=mov['nombre'], 
                    anchor="w"
                ).pack(side="left", fill="x", expand=True)
                
                ttk.Label(
                    frame_mov, 
                    text=f"S/. {mov['monto']:,.2f}", 
                    anchor="e"
                ).pack(side="right")
                
                total_haber += mov['monto']
                max_lineas_haber += 1
                
                if mov['glosa']:
                    ttk.Label(
                        haber_content_frame,
                        text=f"Glosa: {mov['glosa']}",
                        font=('Segoe UI', 8),
                        foreground="#aaaaaa"
                    ).pack(anchor="w", padx=25, pady=(0, 2))
                    max_lineas_haber += 0.5
        else:
            ttk.Label(
                haber_content_frame, 
                text="No hay operaciones", 
                foreground="#888888"
            ).pack(pady=5)
        
        # Total HABER
        ttk.Separator(frame_haber, orient="horizontal").pack(fill="x", pady=(5, 0))
        frame_total_haber = ttk.Frame(frame_haber)
        frame_total_haber.pack(fill="x")
        
        ttk.Label(
            frame_total_haber, 
            text="TOTAL HABER:", 
            font=('Segoe UI', 9, 'bold')
        ).pack(side="left")
        
        ttk.Label(
            frame_total_haber, 
            text=f"S/. {total_haber:,.2f}", 
            font=('Segoe UI', 9, 'bold')
        ).pack(side="right")
        
        # Ajustar altura de columnas
        diferencia_lineas = max_lineas_debe - max_lineas_haber
        if diferencia_lineas > 0:
            for _ in range(int(diferencia_lineas)):
                ttk.Frame(haber_content_frame, height=1).pack(fill="x")
        elif diferencia_lineas < 0:
            for _ in range(int(abs(diferencia_lineas))):
                ttk.Frame(debe_content_frame, height=1).pack(fill="x")
        
        # Saldo de la cuenta (usando el saldo acumulado)
        saldo_final = datos['saldo_acumulado']
        frame_saldo = ttk.Frame(frame_cuenta)
        frame_saldo.pack(fill="x", pady=(5, 0))
        
        if saldo_final > 0:
            ttk.Label(
                frame_saldo, 
                text=f"SALDO DEUDOR: S/. {saldo_final:,.2f}",
                bootstyle="warning",
                font=('Segoe UI', 9, 'bold')
            ).pack(side="left")
        elif saldo_final < 0:
            ttk.Label(
                frame_saldo, 
                text=f"SALDO ACREEDOR: S/. {abs(saldo_final):,.2f}",
                bootstyle="warning",
                font=('Segoe UI', 9, 'bold')
            ).pack(side="right")
        else:
            ttk.Label(
                frame_saldo, 
                text="SALDO FINAL = S/. 0",
                bootstyle="warning",
                font=('Segoe UI', 9, 'bold')
            ).pack(fill="x")
    
    # Botón de regreso
    btn_frame = ttk.Frame(scrollable_main_frame)
    btn_frame.pack(pady=10)
    
    ttk.Button(
        btn_frame,
        text="Regresar a Libros",
        bootstyle='secondary',
        command=lambda: mostrar_seccion("libros")
    ).pack()
    
    # Ajustar scroll
    main_canvas.yview_moveto(0.0)
    canvas.xview_moveto(1.0)

# --- Funciones auxiliares para manejo de meses ---
def obtener_mes_de_fecha(fecha_str):
    """Extrae el mes/año de una fecha en formato dd/mm/yyyy"""
    try:
        fecha = datetime.strptime(fecha_str.split()[0], "%d/%m/%Y")
        return fecha.strftime("%m/%Y")
    except:
        return None

def actualizar_meses_disponibles():
    """Actualiza la lista de meses disponibles sin borrar los anteriores"""
    global meses_disponibles
    
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    try:
        # Obtener todos los meses distintos con datos
        cursor.execute('''
            SELECT DISTINCT substr(fecha, 4, 7) as mes 
            FROM operaciones 
            ORDER BY substr(fecha, 7, 4) || substr(fecha, 4, 2)
        ''')
        
        # Agregar todos los meses encontrados (sin duplicados)
        nuevos_meses = [row[0] for row in cursor.fetchall()]
        
        # Combinar con los meses existentes, eliminando duplicados
        todos_meses = list(set(meses_disponibles + nuevos_meses))
        
        # Ordenar los meses cronológicamente
        meses_disponibles = sorted(todos_meses, 
                                 key=lambda x: datetime.strptime(x, "%m/%Y"))
        
    finally:
        conn.close()

def cargar_operaciones_por_mes(mes):
    """Carga operaciones filtradas por mes específico"""
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT 
            id,
            operacion_id, 
            fecha, 
            cuenta_debe, 
            cuenta_haber, 
            monto, 
            moneda, 
            costos, 
            actividad, 
            glosa 
        FROM operaciones 
        WHERE substr(fecha, 4, 7) = ?
        ORDER BY operacion_id DESC, fecha DESC
    ''', (mes,))

    columnas = [desc[0] for desc in cursor.description]
    operaciones = []
    
    for fila in cursor.fetchall():
        operacion = dict(zip(columnas, fila))
        operaciones.append({
            'id': operacion['id'],
            'operacion_id': operacion['operacion_id'],
            'fecha': operacion['fecha'],
            'debe': operacion['cuenta_debe'],
            'haber': operacion['cuenta_haber'],
            'cuenta_debe': operacion['cuenta_debe'],
            'cuenta_haber': operacion['cuenta_haber'],
            'monto': "{:,.2f}".format(operacion['monto']),
            'moneda': operacion['moneda'],
            'costos': operacion['costos'],
            'actividad': operacion['actividad'],
            'glosa': operacion['glosa']
        })
    
    conn.close()
    return operaciones

def filtrar_libro_diario_por_mes(tree, mes):
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    for item in tree.get_children():
        tree.delete(item)
    
    suma_debe = 0.0
    suma_haber = 0.0
    
    # Excluye explícitamente operaciones con glosa "Saldo inicial"
    cursor.execute('''
        SELECT 
            operacion_id,
            fecha,
            cuenta_debe,
            cuenta_haber,
            monto
        FROM operaciones
        WHERE substr(fecha, 4, 7) = ?
        AND (glosa != 'Saldo inicial' OR glosa IS NULL)
        ORDER BY operacion_id DESC, fecha DESC
    ''', (mes,))
    
    for operacion in cursor.fetchall():
        op_dict = {
            "id": operacion[0],
            "fecha": operacion[1],
            "cuenta_debe": operacion[2],
            "cuenta_haber": operacion[3],
            "monto": float(operacion[4].replace(",", "")) if isinstance(operacion[4], str) else operacion[4]
        }
        
        fecha = op_dict["fecha"].split()[0] if "fecha" in op_dict else ""
        
        if op_dict["cuenta_debe"]:
            tree.insert("", "end", values=(
                op_dict["id"],
                fecha,
                op_dict["cuenta_debe"],
                f"{op_dict['monto']:,.2f}",
                ""
            ))
            suma_debe += op_dict['monto']
        
        if op_dict["cuenta_haber"]:
            tree.insert("", "end", values=(
                op_dict["id"],
                fecha,
                op_dict["cuenta_haber"],
                "",
                f"{op_dict['monto']:,.2f}"
            ))
            suma_haber += op_dict['monto']
    
    conn.close()
    
    # Actualizar totales si existen
    if 'total_debe' in globals() and 'total_haber' in globals():
        total_debe.config(text=f"Debe: S/. {suma_debe:,.2f}")
        total_haber.config(text=f"Haber: S/. {suma_haber:,.2f}")
        
        if abs(suma_debe - suma_haber) > 0.01 and 'totales_frame' in globals():
            for widget in totales_frame.winfo_children():
                if isinstance(widget, ttk.Label) and "ADVERTENCIA" in widget.cget("text"):
                    widget.destroy()
            
            ttk.Label(
                totales_frame,
                text="¡ADVERTENCIA: Los totales no coinciden!",
                bootstyle=DANGER,
                font=('Segoe UI', 10, 'bold')
            ).grid(row=0, column=3, padx=10, sticky="e")

def ver_detalle_operacion(tree):
    item = tree.focus()
    if not item:
        return
    
    valores = tree.item(item, "values")
    fecha = valores[0]
    cuenta = valores[1]
    
    for operacion in operaciones_registradas:
        if (operacion["fecha"].startswith(fecha) and 
            (cuenta in operacion["debe"] or cuenta in operacion["haber"])):
            mostrar_detalle_operacion(operacion)
            break

def mostrar_detalle_operacion(operacion):
    ventana_detalle = tk.Toplevel(root)
    ventana_detalle.title("Detalle de Operación")
    ventana_detalle.geometry("500x300")
    
    ttk.Label(
        ventana_detalle,
        text="Detalle Completo de la Operación",
        style='Titulo.TLabel'
    ).grid(row=0, column=0, pady=10)
    
    frame_detalle = ttk.Frame(ventana_detalle)
    frame_detalle.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
    
    campos = [
        ("Fecha:", operacion["fecha"]),
        ("Cuenta Debe:", operacion["debe"]),
        ("Cuenta Haber:", operacion["haber"]),
        ("Monto:", operacion["monto"]),
        ("Moneda:", operacion["moneda"]),
        ("Tipo de Costos:", operacion.get("costos", "")),
        ("Actividad:", operacion.get("actividad", "")),
        ("Glosa:", operacion.get("glosa", ""))
    ]
    
    for i, (texto, valor) in enumerate(campos):
        frame_campo = ttk.Frame(frame_detalle)
        frame_campo.grid(row=i, column=0, sticky="ew", pady=2)
        
        ttk.Label(frame_campo, text=texto, width=15, bootstyle="dark").grid(row=0, column=0, sticky="w")
        ttk.Label(frame_campo, text=valor if valor else "N/A").grid(row=0, column=1, padx=5, sticky="w")
    
    ttk.Button(
        ventana_detalle,
        text="Cerrar",
        bootstyle=DANGER,
        command=ventana_detalle.destroy
    ).grid(row=2, column=0, pady=10)

def mostrar_operaciones():
    global tree_operaciones
    
    frame_operaciones = ttk.Frame(main_frame)
    frame_operaciones.grid(row=0, column=0, sticky="nsew", padx=10, pady=20)
    frame_operaciones.grid_columnconfigure(0, weight=1)
    frame_operaciones.grid_rowconfigure(1, weight=1)
    
    ttk.Label(
        frame_operaciones,
        text="REGISTRO DE OPERACIONES CONTABLES",
        style='Titulo.TLabel'
    ).grid(row=0, column=0, pady=(0, 20), sticky="w")

    tree_frame = ttk.Frame(frame_operaciones)
    tree_frame.grid(row=1, column=0, sticky="nsew")
    tree_frame.grid_columnconfigure(0, weight=1)
    tree_frame.grid_rowconfigure(0, weight=1)
    
    # Columnas actualizadas (sin moneda)
    columns = ("id", "fecha", "debe", "haber", "monto", "costos", "actividad", "glosa")
    tree_operaciones = ttk.Treeview(
        tree_frame,
        columns=columns,
        show="headings",
        height=12,
        selectmode="browse"
    )
    
    # Configurar encabezados (sin moneda)
    tree_operaciones.heading("id", text="ID", anchor="center")
    tree_operaciones.heading("fecha", text="Fecha", anchor="center")
    tree_operaciones.heading("debe", text="Cuenta Débito", anchor="w")
    tree_operaciones.heading("haber", text="Cuenta Crédito", anchor="w")
    tree_operaciones.heading("monto", text="Monto (S/.)", anchor="e")
    tree_operaciones.heading("costos", text="Costos", anchor="center")
    tree_operaciones.heading("actividad", text="Actividad", anchor="center")
    tree_operaciones.heading("glosa", text="Glosa", anchor="w")
    
    # Configurar columnas (con redistribución de espacio)
    tree_operaciones.column("id", width=60, anchor="center")
    tree_operaciones.column("fecha", width=100, anchor="center")
    tree_operaciones.column("debe", width=220, anchor="w")  # Más ancho
    tree_operaciones.column("haber", width=220, anchor="w")  # Más ancho
    tree_operaciones.column("monto", width=120, anchor="e")
    tree_operaciones.column("costos", width=100, anchor="center")
    tree_operaciones.column("actividad", width=100, anchor="center")
    tree_operaciones.column("glosa", width=280, anchor="w")  # Más ancho
    
    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree_operaciones.yview)
    scrollbar.grid(row=0, column=1, sticky="ns")
    tree_operaciones.configure(yscrollcommand=scrollbar.set)
    tree_operaciones.grid(row=0, column=0, sticky="nsew")
    
    cargar_datos_operaciones_sin_moneda(tree_operaciones)

def cargar_datos_operaciones_sin_moneda(tree):
    cargar_datos_operaciones(tree_operaciones)

def mostrar_confirmacion(debe_items, haber_items, moneda, costos, actividad, glosa):
    global confirmacion_frame
    
    if not debe_items or not haber_items:
        tk.messagebox.showerror("Error", "Debe agregar al menos un débito y un crédito")
        return
    
    total_debe = sum(item['monto'] for item in debe_items)
    total_haber = sum(item['monto'] for item in haber_items)
    
    # Crear mensaje detallado
    mensaje = "¿Confirmar operación?\n\n"
    mensaje += "DÉBITOS:\n"
    for item in debe_items:
        mensaje += f"- {item['cuenta']}: S/. {item['monto']:,.2f}\n"
    
    mensaje += "\nCRÉDITOS:\n"
    for item in haber_items:
        mensaje += f"- {item['cuenta']}: S/. {item['monto']:,.2f}\n"
    
    mensaje += f"\nTOTAL DÉBITO: S/. {total_debe:,.2f}\n"
    mensaje += f"TOTAL CRÉDITO: S/. {total_haber:,.2f}\n"
    
    if abs(total_debe - total_haber) > 0.01:
        mensaje += "\n¡ADVERTENCIA! Los totales no coinciden"
        tk.messagebox.showwarning("Confirmación", mensaje)
        return
    
    # Mostrar diálogo de confirmación
    confirmar = tk.messagebox.askyesno("Confirmar operación", mensaje)
    if confirmar:
        registrar_operacion(debe_items, haber_items, moneda, costos, actividad, glosa)
    
def registrar_operacion():
    try:
        # Validación inicial de combobox
        if not costos_var.get():
            costos_combobox.current(0)
        if not actividad_var.get():
            actividad_combobox.current(0)

        # Obtener valores con verificación
        costos = costos_combobox.get()
        actividad = actividad_combobox.get()
        
        print(f"\nValores a registrar - Costos: {costos}, Actividad: {actividad}")  # Diagnóstico

        # Resto de validaciones...
        if not costos or costos not in costos_combobox['values']:
            messagebox.showerror("Error", "Seleccione un tipo de Costos válido")
            costos_combobox.focus()
            return
            
        if not actividad or actividad not in actividad_combobox['values']:
            messagebox.showerror("Error", "Seleccione un tipo de Actividad válido")
            actividad_combobox.focus()
            return

        # Resto del código de registro...
        conn = sqlite3.connect('data/contabilidad.db')
        cursor = conn.cursor()
        
        operacion_id = obtener_proximo_id()
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        # Insertar débitos
        for debe in debe_items:
            cursor.execute('''
                INSERT INTO operaciones 
                (operacion_id, fecha, cuenta_debe, cuenta_haber, monto, moneda, costos, actividad, glosa)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                operacion_id,
                fecha_actual,
                debe['cuenta'],
                "",
                debe['monto'],
                moneda_var.get(),
                costos,  # Usamos el valor validado
                actividad,  # Usamos el valor validado
                glosa_entry.get()
            ))
        
        # Insertar créditos
        for haber in haber_items:
            cursor.execute('''
                INSERT INTO operaciones 
                (operacion_id, fecha, cuenta_debe, cuenta_haber, monto, moneda, costos, actividad, glosa)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                operacion_id,
                fecha_actual,
                "",
                haber['cuenta'],
                haber['monto'],
                moneda_var.get(),
                costos,  # Mismo valor para toda la operación
                actividad,  # Mismo valor para toda la operación
                glosa_entry.get()
            ))
        
        conn.commit()
        
        # Verificación en consola
        print("Registro exitoso. Valores guardados:")
        print(f"- Costos: {costos}")
        print(f"- Actividad: {actividad}")
        
        # Limpiar formulario
        limpiar_formulario()
        
        # Actualizar vista
        if 'tree_operaciones' in globals() and tree_operaciones:
            cargar_datos_operaciones(tree_operaciones)
        
        messagebox.showinfo("Éxito", "Operación registrada correctamente")
        
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo registrar la operación: {str(e)}")
        print(f"Error completo: {traceback.format_exc()}")
        if 'conn' in locals():
            conn.rollback()
    finally:
        if 'conn' in locals():
            conn.close()

def limpiar_formulario():
    # Limpiar listas y entradas
    debe_items.clear()
    haber_items.clear()
    monto_entry.delete(0, tk.END)
    glosa_entry.delete(0, tk.END)
    cuenta_debe_var.set("Seleccionar cuenta...")
    cuenta_haber_var.set("Seleccionar cuenta...")
    
    # Restablecer Combobox a valores por defecto
    costos_combobox.current(0)
    actividad_combobox.current(0)
    
    # Actualizar treeviews
    actualizar_treeview_debe()
    actualizar_treeview_haber()

def verificar_componentes():
    componentes = {
        'debe_items': debe_items,
        'haber_items': haber_items,
        'monto_entry': monto_entry,
        'glosa_entry': glosa_entry,
        'cuenta_debe_var': cuenta_debe_var,
        'cuenta_haber_var': cuenta_haber_var,
        'moneda_var': moneda_var,
        'costos_var': costos_var,
        'actividad_var': actividad_var,
        'tree_debe': tree_debe,
        'tree_haber': tree_haber
    }
    
    for nombre, componente in componentes.items():
        estado = "OK" if componente is not None else "FALTA"
        print(f"{nombre}: {estado}")

def mostrar_botones_confirmacion():
    """Muestra los botones de confirmación/cancelación"""
    global frame_botones_accion, frame_confirmacion
    
    # Validación básica
    if not debe_items or not haber_items:
        messagebox.showerror("Error", "Debe agregar al menos un débito y un crédito")
        return
    
    # Validar equilibrio contable
    total_debe = sum(item['monto'] for item in debe_items)
    total_haber = sum(item['monto'] for item in haber_items)
    
    if abs(total_debe - total_haber) > 0.01:
        messagebox.showerror("Error", 
            f"Los totales no coinciden\nDebe: S/. {total_debe:,.2f}\nHaber: S/. {total_haber:,.2f}")
        return
    
    # Ocultar botón principal y mostrar botones de confirmación
    frame_botones_accion.grid_forget()
    frame_confirmacion.grid(row=0, column=0)  # Ahora en la misma posición

def ocultar_botones_confirmacion():
    """Oculta los botones de confirmación y muestra el botón principal"""
    frame_confirmacion.grid_forget()
    frame_botones_accion.grid(row=0, column=0)

def confirmar_registro():
    """Registra la operación y restablece los botones a su estado inicial"""
    try:
        # Validar que haya al menos un débito y un crédito
        if not debe_items or not haber_items:
            messagebox.showerror("Error", "Debe agregar al menos un débito y un crédito")
            return

        # Validar equilibrio contable
        total_debe = sum(item['monto'] for item in debe_items)
        total_haber = sum(item['monto'] for item in haber_items)
        
        if abs(total_debe - total_haber) > 0.01:
            messagebox.showerror("Error", 
                f"Los totales no coinciden\nDebe: S/. {total_debe:,.2f}\nHaber: S/. {total_haber:,.2f}")
            return

        # Registrar en la base de datos
        operacion_id = obtener_proximo_id()
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        conn = sqlite3.connect('data/contabilidad.db')
        cursor = conn.cursor()
        
        # Registrar débitos
        for debe in debe_items:
            cursor.execute('''
                INSERT INTO operaciones 
                (operacion_id, fecha, cuenta_debe, cuenta_haber, monto, moneda, costos, actividad, glosa)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                operacion_id,
                fecha_actual,
                debe['cuenta'],
                "",
                debe['monto'],
                "S/",
                costos_var.get(),
                actividad_var.get(),
                glosa_entry.get()
            ))
        
        # Registrar créditos
        for haber in haber_items:
            cursor.execute('''
                INSERT INTO operaciones 
                (operacion_id, fecha, cuenta_debe, cuenta_haber, monto, moneda, costos, actividad, glosa)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                operacion_id,
                fecha_actual,
                "",
                haber['cuenta'],
                haber['monto'],
                "S/",
                costos_var.get(),
                actividad_var.get(),
                glosa_entry.get()
            ))
        
        conn.commit()
        conn.close()
        
        # Limpiar y actualizar
        limpiar_formulario()
        
        if 'tree_operaciones' in globals() and tree_operaciones:
            cargar_datos_operaciones(tree_operaciones)
        
        # RESTABLECER LOS BOTONES - PARTE NUEVA
        ocultar_botones_confirmacion()  # Esto mostrará nuevamente el botón principal
        
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo registrar la operación: {str(e)}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        # Asegurarse de restablecer los botones incluso si hay error
        ocultar_botones_confirmacion()
    
def mostrar_estados_financieros():
    # Limpiar el main_frame
    for widget in main_frame.winfo_children():
        widget.destroy()
    
    # Ocultar formulario_asientos si está visible
    if 'formulario_asientos' in globals() and formulario_asientos.winfo_ismapped():
        formulario_asientos.grid_forget()
    
    # Frame principal
    frame_principal = ttk.Frame(main_frame)
    frame_principal.pack(fill="both", expand=True, padx=40, pady=20)
    
    # Título
    ttk.Label(
        frame_principal,
        text="ESTADOS FINANCIEROS",
        style='Titulo.TLabel',
        font=('Helvetica', 20, 'bold')
    ).pack(pady=(10, 30))
    
    # Contenedor de botones
    botones_frame = ttk.Frame(frame_principal)
    botones_frame.pack(fill="x", pady=(0, 20))
    
    # Configuración de grid
    botones_frame.grid_columnconfigure(0, weight=1)
    botones_frame.grid_columnconfigure(1, weight=1)
    botones_frame.grid_columnconfigure(2, weight=1)
        
    # Estilo para botones más compactos
    estilo = ttk.Style()
    estilo.configure('Compact.TButton', 
                   font=('Helvetica', 12, 'bold'),
                   padding=(15, 10),  # Menor padding vertical
                   anchor="center")
    
    # Botón Balance General - altura específica
    btn_balance = ttk.Button(
        botones_frame,
        text="📊  BALANCE GENERAL",
        style='Compact.TButton',
        bootstyle=PRIMARY,
        command=mostrar_balance_general  # Función asociada
    )
    btn_balance.grid(row=0, column=0, padx=15, pady=34, sticky="nsew")
    botones_frame.rowconfigure(0, minsize=120)  # Altura fija para la fila
    
    # Botón Estado de Resultados
    btn_gyg = ttk.Button(
        botones_frame,
        text="💰  ESTADO DE RESULTADOS",
        style='Compact.TButton',
        bootstyle=SUCCESS,
        command=mostrar_estado_resultados  # Función asociada
    )
    btn_gyg.grid(row=0, column=1, padx=15, pady=34, sticky="nsew")
    
    # Botón Flujo de Efectivo
    btn_flujo = ttk.Button(
        botones_frame,
        text="💵  FLUJO DE EFECTIVO",
        style='Compact.TButton',
        bootstyle=INFO,
        command=mostrar_flujo_efectivo  # Función asociada
    )
    btn_flujo.grid(row=0, column=2, padx=15, pady=34, sticky="nsew")

def mostrar_balance_general():
    for widget in main_frame.winfo_children():
        widget.destroy()
    
    frame = ttk.Frame(main_frame)
    frame.pack(fill="both", expand=True, padx=20, pady=20)
    
    ttk.Label(
        frame,
        text="BALANCE GENERAL",
        style='Titulo.TLabel'
    ).pack(pady=(0, 20))
    
    # Aquí irá el contenido real del balance general
    ttk.Label(
        frame,
        text="Contenido del Balance General se mostrará aquí",
        font=('Helvetica', 12)
    ).pack(pady=50)
    
    ttk.Button(
        frame,
        text="Regresar a Estados Financieros",
        bootstyle=(OUTLINE, SECONDARY),
        command=mostrar_estados_financieros
    ).pack()

def mostrar_estado_resultados(mes=None):
    # Limpiar el main_frame
    for widget in main_frame.winfo_children():
        widget.destroy()

    # Si no se especifica mes, usar el actual
    if mes is None:
        mes = datetime.now().strftime("%m/%Y")
    
    # Actualizar meses disponibles
    actualizar_meses_disponibles()

    # --- CONTENEDOR PRINCIPAL ---
    main_container = ttk.Frame(main_frame)
    main_container.pack(fill="both", expand=True, padx=20, pady=20)
    main_container.grid_columnconfigure(0, weight=1)

    # --- BARRA SUPERIOR CON MESES ---
    top_frame = ttk.Frame(main_container)
    top_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
    top_frame.grid_columnconfigure(1, weight=1)

    # Título
    ttk.Label(
        top_frame,
        text="ESTADO DE RESULTADOS",
        font=('Segoe UI', 14, 'bold'),
        foreground='white'
    ).grid(row=0, column=0, padx=(0, 10), sticky="w")

    # Contenedor para botones de meses (con scroll horizontal)
    meses_container = ttk.Frame(top_frame)
    meses_container.grid(row=0, column=1, sticky="nsew")
    
    canvas = tk.Canvas(meses_container, height=40, highlightthickness=0, bg='#333333')
    scrollbar = ttk.Scrollbar(meses_container, orient="horizontal", command=canvas.xview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all"),
            width=e.width if e.width > meses_container.winfo_width() else meses_container.winfo_width()
        )
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(xscrollcommand=scrollbar.set)
    
    canvas.grid(row=0, column=0, sticky="ew")
    scrollbar.grid(row=1, column=0, sticky="ew")
    meses_container.grid_columnconfigure(0, weight=1)

    # Crear botones de meses (ordenados cronológicamente)
    meses_ordenados = sorted(meses_disponibles, key=lambda x: datetime.strptime(x, "%m/%Y"), reverse=True)
    
    for i, mes_disponible in enumerate(meses_ordenados):
        estilo = 'primary' if mes_disponible == mes else 'secondary'
        
        btn = ttk.Button(
            scrollable_frame,
            text=datetime.strptime(mes_disponible, "%m/%Y").strftime("%B %Y").upper(),
            bootstyle=estilo,
            width=15,
            command=lambda m=mes_disponible: mostrar_estado_resultados(m)
        )
        btn.grid(row=0, column=i, padx=3, pady=2, sticky="nsew")

    # --- CONTENEDOR DE CUADROS ---
    cuadros_container = ttk.Frame(main_container)
    cuadros_container.grid(row=1, column=0, sticky="nsew")
    cuadros_container.grid_columnconfigure(0, weight=1)

    # ===== CUADRO 1: ESTADO DE RESULTADOS =====
    frame_estado = ttk.LabelFrame(
        cuadros_container,
        text=f" ESTADO DE RESULTADOS - {datetime.strptime(mes, '%m/%Y').strftime('%B %Y').upper()} ",
        style='Transparent.TLabelframe',
        padding=10
    )
    frame_estado.grid(row=0, column=0, pady=(0, 15), sticky="nsew", padx=50)
    frame_estado.grid_columnconfigure(0, weight=1)

    # Obtener saldos (filtrados por mes)
    saldo_70 = obtener_saldo_cuenta_mes("70", mes)
    saldo_69 = obtener_saldo_cuenta_mes("69", mes)
    utilidad_bruta = (-1) * (saldo_69 + saldo_70)

    # Treeview
    tree_estado = ttk.Treeview(
        frame_estado,
        columns=("concepto", "monto"),
        show="headings",
        height=3,
        selectmode="none",
        style="Custom.Treeview"
    )
    tree_estado.heading("concepto", text="CONCEPTO", anchor="w")
    tree_estado.heading("monto", text="MONTO (S/.)", anchor="e")
    tree_estado.column("concepto", width=400, anchor="w", stretch=True)
    tree_estado.column("monto", width=200, anchor="e", stretch=False)

    # Insertar datos
    tree_estado.insert("", "end", values=[f"70 - {obtener_nombre_cuenta('70')}", 
                      f"S/. {abs(saldo_70):,.2f}"], tags=("data",))
    tree_estado.insert("", "end", values=[f"69 - {obtener_nombre_cuenta('69')}", 
                      f"S/. {abs(saldo_69):,.2f}"], tags=("data",))
    tree_estado.insert("", "end", values=["UTILIDAD BRUTA", 
                      f"S/. {utilidad_bruta:,.2f}"], tags=("total",))

    tree_estado.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

    # ===== CUADRO 2: GASTOS OPERATIVOS =====
    frame_gastos = ttk.LabelFrame(
        cuadros_container,
        text=f" GASTOS OPERATIVOS - {datetime.strptime(mes, '%m/%Y').strftime('%B %Y').upper()} ",
        style='Transparent.TLabelframe',
        padding=10
    )
    frame_gastos.grid(row=1, column=0, pady=(0, 15), sticky="nsew", padx=50)
    frame_gastos.grid_columnconfigure(0, weight=1)

    # Obtener y procesar gastos (filtrados por mes)
    rangos_gastos = list(range(60, 69)) + list(range(80, 100))
    cuentas_gastos = []
    for codigo in rangos_gastos:
        saldo = obtener_saldo_cuenta_mes(str(codigo), mes)
        if saldo != 0:
            cuentas_gastos.append((f"{codigo} - {obtener_nombre_cuenta(str(codigo))}", saldo))

    total_gastos = sum(saldo for _, saldo in cuentas_gastos)
    utilidad_operativa = - utilidad_bruta + total_gastos

    # Treeview
    tree_gastos = ttk.Treeview(
        frame_gastos,
        columns=("concepto", "monto"),
        show="headings",
        height=min(len(cuentas_gastos) + 2, 8),
        selectmode="none",
        style="Custom.Treeview"
    )
    tree_gastos.heading("concepto", text="CONCEPTO", anchor="w")
    tree_gastos.heading("monto", text="MONTO (S/.)", anchor="e")
    tree_gastos.column("concepto", width=400, anchor="w", stretch=True)
    tree_gastos.column("monto", width=200, anchor="e", stretch=False)

    for cuenta, saldo in cuentas_gastos:
        monto_str = f"S/. {abs(saldo):,.2f}" if saldo >=0 else f"(S/. {abs(saldo):,.2f})"
        tree_gastos.insert("", "end", values=[cuenta, monto_str], tags=("data",))

    tree_gastos.insert("", "end", values=["TOTAL GASTOS", 
                      f"S/. {abs(total_gastos):,.2f}" if total_gastos >=0 else f"(S/. {abs(total_gastos):,.2f})"], 
                      tags=("subtotal",))
    
    tree_gastos.insert("", "end", values=["UTILIDAD OPERATIVA", 
                      f"S/. {utilidad_operativa:,.2f}"], 
                      tags=("final",))

    tree_gastos.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

    # Scrollbar solo si es necesario
    if len(cuentas_gastos) > 6:
        scrollbar_gastos = ttk.Scrollbar(frame_gastos, orient="vertical", command=tree_gastos.yview)
        scrollbar_gastos.grid(row=0, column=1, sticky="ns")
        tree_gastos.configure(yscrollcommand=scrollbar_gastos.set)

    # ===== CUADRO 3: INGRESOS =====
    frame_ingresos = ttk.LabelFrame(
        cuadros_container,
        text=f" INGRESOS - {datetime.strptime(mes, '%m/%Y').strftime('%B %Y').upper()} ",
        style='Transparent.TLabelframe',
        padding=10
    )
    frame_ingresos.grid(row=2, column=0, pady=(0, 15), sticky="nsew", padx=50)
    frame_ingresos.grid_columnconfigure(0, weight=1)

    # Obtener y procesar ingresos (filtrados por mes, con lógica invertida)
    cuentas_ingresos = []
    for codigo in range(71, 80):
        saldo = obtener_saldo_cuenta_mes(str(codigo), mes)
        if saldo != 0:
            # INVERTIMOS EL SIGNO PARA INGRESOS (solicitado)
            saldo_invertido = -saldo
            cuentas_ingresos.append((f"{codigo} - {obtener_nombre_cuenta(str(codigo))}", saldo_invertido))

    total_ingresos = sum(saldo for _, saldo in cuentas_ingresos)
    utilidad_antes_impuestos = total_ingresos - utilidad_operativa

    # Treeview
    tree_ingresos = ttk.Treeview(
        frame_ingresos,
        columns=("concepto", "monto"),
        show="headings",
        height=min(len(cuentas_ingresos) + 2, 8),
        selectmode="none",
        style="Custom.Treeview"
    )
    tree_ingresos.heading("concepto", text="CONCEPTO", anchor="w")
    tree_ingresos.heading("monto", text="MONTO (S/.)", anchor="e")
    tree_ingresos.column("concepto", width=400, anchor="w", stretch=True)
    tree_ingresos.column("monto", width=200, anchor="e", stretch=False)

    for cuenta, saldo in cuentas_ingresos:
        # Mostramos positivo para acreedor, negativo para deudor (solicitado)
        monto_str = f"S/. {saldo:,.2f}" if saldo >=0 else f"(S/. {abs(saldo):,.2f})"
        tree_ingresos.insert("", "end", values=[cuenta, monto_str], tags=("data",))

    tree_ingresos.insert("", "end", values=["TOTAL INGRESOS", 
                          f"S/. {total_ingresos:,.2f}" if total_ingresos >=0 else f"(S/. {abs(total_ingresos):,.2f})"], 
                          tags=("subtotal",))
    
    tree_ingresos.insert("", "end", values=["UTILIDAD ANTES DE IMPUESTOS", 
                          f"S/. {utilidad_antes_impuestos:,.2f}"], 
                          tags=("final",))

    tree_ingresos.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

    # Scrollbar solo si es necesario
    if len(cuentas_ingresos) > 6:
        scrollbar_ingresos = ttk.Scrollbar(frame_ingresos, orient="vertical", command=tree_ingresos.yview)
        scrollbar_ingresos.grid(row=0, column=1, sticky="ns")
        tree_ingresos.configure(yscrollcommand=scrollbar_ingresos.set)

    # --- CONFIGURAR ESTILOS DE FILAS ---
    for tree in [tree_estado, tree_gastos, tree_ingresos]:
        tree.tag_configure("data", foreground='white')
        tree.tag_configure("total", foreground='#4fc3f7', font=('Segoe UI', 10, 'bold'))
        tree.tag_configure("subtotal", foreground='#00c292', font=('Segoe UI', 10, 'bold'))
        tree.tag_configure("final", foreground='white', font=('Segoe UI', 10, 'bold'), background='#3a4f6a')

    # --- BOTÓN DE REGRESO ---
    btn_frame = ttk.Frame(main_container)
    btn_frame.grid(row=2, column=0, pady=(15, 5), sticky="e")
    
    ttk.Button(
        btn_frame,
        text="Regresar a Estados Financieros",
        bootstyle="secondary",
        command=mostrar_estados_financieros
    ).pack(side="right", padx=10)

    # Asegurar scroll al final
    canvas.xview_moveto(1.0)

def mostrar_flujo_efectivo(mes=None):
    # Limpiar el main_frame
    for widget in main_frame.winfo_children():
        widget.destroy()

    # Si no se especifica mes, usar el actual
    if mes is None:
        mes = datetime.now().strftime("%m/%Y")
    
    # Actualizar meses disponibles
    actualizar_meses_disponibles()

    # --- CONTENEDOR PRINCIPAL ---
    main_container = ttk.Frame(main_frame)
    main_container.pack(fill="both", expand=True, padx=20, pady=20)
    main_container.grid_columnconfigure(0, weight=1)

    # --- BARRA SUPERIOR CON MESES ---
    top_frame = ttk.Frame(main_container)
    top_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
    top_frame.grid_columnconfigure(1, weight=1)

    # Título
    ttk.Label(
        top_frame,
        text="FLUJO DE EFECTIVO",
        font=('Segoe UI', 14, 'bold'),
        foreground='white'
    ).grid(row=0, column=0, padx=(0, 10), sticky="w")

    # Contenedor para botones de meses (con scroll horizontal)
    meses_container = ttk.Frame(top_frame)
    meses_container.grid(row=0, column=1, sticky="nsew")
    
    canvas = tk.Canvas(meses_container, height=40, highlightthickness=0, bg='#333333')
    scrollbar = ttk.Scrollbar(meses_container, orient="horizontal", command=canvas.xview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all"),
            width=e.width if e.width > meses_container.winfo_width() else meses_container.winfo_width()
        )
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(xscrollcommand=scrollbar.set)
    
    canvas.grid(row=0, column=0, sticky="ew")
    scrollbar.grid(row=1, column=0, sticky="ew")
    meses_container.grid_columnconfigure(0, weight=1)

    # Crear botones de meses (ordenados cronológicamente)
    meses_ordenados = sorted(meses_disponibles, key=lambda x: datetime.strptime(x, "%m/%Y"), reverse=True)
    
    for i, mes_disponible in enumerate(meses_ordenados):
        estilo = 'primary' if mes_disponible == mes else 'secondary'
        
        btn = ttk.Button(
            scrollable_frame,
            text=datetime.strptime(mes_disponible, "%m/%Y").strftime("%B %Y").upper(),
            bootstyle=estilo,
            width=15,
            command=lambda m=mes_disponible: mostrar_flujo_efectivo(m)
        )
        btn.grid(row=0, column=i, padx=3, pady=2, sticky="nsew")

    # --- CONTENEDOR DE CUADROS ---
    cuadros_container = ttk.Frame(main_container)
    cuadros_container.grid(row=1, column=0, sticky="nsew")
    cuadros_container.grid_columnconfigure(0, weight=1)

    # Obtener el saldo inicial de la cuenta 10 (SI)
    saldo_inicial = obtener_saldo_inicial_cuenta_10(mes)

    # Obtener las operaciones de la cuenta 10 clasificadas por actividad
    operaciones_operacion = obtener_operaciones_cuenta_10_por_actividad(mes, "Operación")
    operaciones_inversion = obtener_operaciones_cuenta_10_por_actividad(mes, "Inversión")
    operaciones_financiamiento = obtener_operaciones_cuenta_10_por_actividad(mes, "Financiamiento")

    # Calcular totales por actividad
    total_debe_operacion = sum(op['monto'] if op['tipo'] == 'debe' else 0 for op in operaciones_operacion)
    total_haber_operacion = sum(op['monto'] if op['tipo'] == 'haber' else 0 for op in operaciones_operacion)
    
    total_debe_inversion = sum(op['monto'] if op['tipo'] == 'debe' else 0 for op in operaciones_inversion)
    total_haber_inversion = sum(op['monto'] if op['tipo'] == 'haber' else 0 for op in operaciones_inversion)
    
    total_debe_financiamiento = sum(op['monto'] if op['tipo'] == 'debe' else 0 for op in operaciones_financiamiento)
    total_haber_financiamiento = sum(op['monto'] if op['tipo'] == 'haber' else 0 for op in operaciones_financiamiento)

    # Calcular saldo final
    saldo_final = saldo_inicial + (total_debe_operacion - total_haber_operacion) + \
                  (total_debe_inversion - total_haber_inversion) + \
                  (total_debe_financiamiento - total_haber_financiamiento)

    # Función para crear un cuadro de actividad
    def crear_cuadro_actividad(frame_parent, titulo, operaciones, total_debe, total_haber):
        frame_actividad = ttk.LabelFrame(
            frame_parent,
            text=f" ACTIVIDADES DE {titulo.upper()} ",
            bootstyle="info",
            padding=10
        )
        frame_actividad.pack(fill="x", pady=(0, 15))
        frame_actividad.grid_columnconfigure(0, weight=1)

        # Treeview
        tree = ttk.Treeview(
            frame_actividad,
            columns=("id", "descripcion", "debe", "haber"),
            show="headings",
            height=min(len(operaciones) + 1, 8),
            selectmode="none"
        )
        
        # Configurar columnas
        tree.heading("id", text="ID", anchor="center")
        tree.heading("descripcion", text="Descripción", anchor="w")
        tree.heading("debe", text="Debe (S/.)", anchor="e")
        tree.heading("haber", text="Haber (S/.)", anchor="e")
        
        tree.column("id", width=60, anchor="center")
        tree.column("descripcion", width=300, anchor="w")
        tree.column("debe", width=120, anchor="e")
        tree.column("haber", width=120, anchor="e")
        
        # Insertar datos
        for op in operaciones:
            debe = f"{op['monto']:,.2f}" if op['tipo'] == 'debe' else ""
            haber = f"{op['monto']:,.2f}" if op['tipo'] == 'haber' else ""
            
            tree.insert("", "end", values=(
                op['id'],
                op['glosa'],
                debe,
                haber
            ))
        
        # Insertar totales
        tree.insert("", "end", values=(
            "TOTAL",
            "",
            f"{total_debe:,.2f}" if total_debe > 0 else "",
            f"({total_haber:,.2f})" if total_haber > 0 else ""
        ), tags=("total",))
        
        tree.grid(row=0, column=0, sticky="nsew")
        
        # Scrollbar si es necesario
        if len(operaciones) > 6:
            scrollbar = ttk.Scrollbar(frame_actividad, orient="vertical", command=tree.yview)
            scrollbar.grid(row=0, column=1, sticky="ns")
            tree.configure(yscrollcommand=scrollbar.set)
        
        # Configurar estilos
        tree.tag_configure("total", foreground='white', font=('Segoe UI', 10, 'bold'), background='#3a4f6a')

    # Crear cuadros para cada tipo de actividad
    if operaciones_operacion:
        crear_cuadro_actividad(
            cuadros_container, 
            "Operación", 
            operaciones_operacion, 
            total_debe_operacion, 
            total_haber_operacion
        )

    if operaciones_inversion:
        crear_cuadro_actividad(
            cuadros_container, 
            "Inversión", 
            operaciones_inversion, 
            total_debe_inversion, 
            total_haber_inversion
        )

    if operaciones_financiamiento:
        crear_cuadro_actividad(
            cuadros_container, 
            "Financiamiento", 
            operaciones_financiamiento, 
            total_debe_financiamiento, 
            total_haber_financiamiento
        )

    # Mostrar saldos
    frame_saldos = ttk.Frame(cuadros_container)
    frame_saldos.pack(fill="x", pady=(10, 0))
    
    # Saldo inicial
    ttk.Label(
        frame_saldos,
        text=f"Saldo Inicial (SI): S/. {saldo_inicial:,.2f}",
        font=('Segoe UI', 10, 'bold')
    ).pack(anchor="w")
    
    # Saldo de actividades
    saldo_operacion = total_debe_operacion - total_haber_operacion
    saldo_inversion = total_debe_inversion - total_haber_inversion
    saldo_financiamiento = total_debe_financiamiento - total_haber_financiamiento
    
    ttk.Label(
        frame_saldos,
        text=f"Flujo Operación: S/. {saldo_operacion:,.2f}",
        font=('Segoe UI', 10)
    ).pack(anchor="w")
    
    ttk.Label(
        frame_saldos,
        text=f"Flujo Inversión: S/. {saldo_inversion:,.2f}",
        font=('Segoe UI', 10)
    ).pack(anchor="w")
    
    ttk.Label(
        frame_saldos,
        text=f"Flujo Financiamiento: S/. {saldo_financiamiento:,.2f}",
        font=('Segoe UI', 10)
    ).pack(anchor="w")
    
    # Saldo final
    ttk.Label(
        frame_saldos,
        text=f"Saldo Efectivo del Periodo Actual: S/. {saldo_final:,.2f}",
        font=('Segoe UI', 10, 'bold'),
        foreground='#4fc3f7'
    ).pack(anchor="w", pady=(5, 0))

    # Botón de regreso
    btn_frame = ttk.Frame(main_container)
    btn_frame.grid(row=2, column=0, pady=(15, 5), sticky="e")
    
    ttk.Button(
        btn_frame,
        text="Regresar a Estados Financieros",
        bootstyle="secondary",
        command=mostrar_estados_financieros
    ).pack(side="right", padx=10)

    # Asegurar scroll al final
    canvas.xview_moveto(1.0)

def obtener_saldo_inicial_cuenta_10(mes):
    """Obtiene el saldo inicial de la cuenta 10 para el mes especificado"""
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    try:
        # Obtener el mes anterior
        mes_num, año = map(int, mes.split('/'))
        if mes_num == 1:
            mes_anterior = f"12/{año-1}"
        else:
            mes_anterior = f"{mes_num-1:02d}/{año}"
        
        # Obtener saldo histórico inicial (SI)
        cursor.execute('''
            SELECT 
                SUM(CASE WHEN cuenta_debe LIKE '10%' AND glosa = 'Saldo inicial' THEN monto ELSE 0 END) -
                SUM(CASE WHEN cuenta_haber LIKE '10%' AND glosa = 'Saldo inicial' THEN monto ELSE 0 END)
            FROM operaciones
        ''')
        saldo_historico = cursor.fetchone()[0] or 0
        
        # Obtener saldo acumulado hasta el mes anterior (excluyendo SI)
        cursor.execute('''
            SELECT 
                SUM(CASE WHEN cuenta_debe LIKE '10%' AND substr(fecha, 4, 7) = ? AND glosa != 'Saldo inicial' THEN monto ELSE 0 END) -
                SUM(CASE WHEN cuenta_haber LIKE '10%' AND substr(fecha, 4, 7) = ? AND glosa != 'Saldo inicial' THEN monto ELSE 0 END)
            FROM operaciones
        ''', (mes_anterior, mes_anterior))
        
        saldo_mes_anterior = cursor.fetchone()[0] or 0
        
        return saldo_historico + saldo_mes_anterior
        
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo obtener saldo inicial: {str(e)}")
        return 0
    finally:
        conn.close()

def obtener_operaciones_cuenta_10_por_actividad(mes, actividad):
    """Obtiene las operaciones de la cuenta 10 filtradas por actividad"""
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT 
                operacion_id as id,
                fecha,
                CASE 
                    WHEN cuenta_debe LIKE '10%' THEN 'debe'
                    WHEN cuenta_haber LIKE '10%' THEN 'haber'
                END as tipo,
                monto,
                CASE WHEN glosa IS NULL THEN '' ELSE glosa END as glosa,
                actividad
            FROM operaciones
            WHERE (cuenta_debe LIKE '10%' OR cuenta_haber LIKE '10%')
            AND substr(fecha, 4, 7) = ?
            AND actividad = ?
            AND (glosa != 'Saldo inicial' OR glosa IS NULL)
            ORDER BY fecha
        ''', (mes, actividad))
        
        columnas = [desc[0] for desc in cursor.description]
        operaciones = []
        
        for fila in cursor.fetchall():
            operacion = dict(zip(columnas, fila))
            operaciones.append({
                'id': operacion['id'],
                'fecha': operacion['fecha'],
                'tipo': operacion['tipo'],
                'monto': operacion['monto'],
                'glosa': operacion['glosa'],
                'actividad': operacion['actividad']
            })
        
        return operaciones
        
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo obtener operaciones: {str(e)}")
        return []
    finally:
        conn.close()

def obtener_saldo_cuenta_mes(codigo_cuenta, mes):
    """Obtiene el saldo real (positivo=deudor, negativo=acreedor) para un mes específico"""
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT 
                SUM(CASE WHEN substr(cuenta_debe, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END) as total_debe,
                SUM(CASE WHEN substr(cuenta_haber, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END) as total_haber
            FROM operaciones
        ''', (codigo_cuenta, mes, codigo_cuenta, mes))
        
        total_debe, total_haber = cursor.fetchone()
        return (total_debe or 0) - (total_haber or 0)
        
    except sqlite3.Error as e:
        messagebox.showerror("Error", f"No se pudo obtener el saldo: {str(e)}")
        return 0
    finally:
        conn.close()

def obtener_saldo_cuenta(codigo_cuenta):
    """Obtiene el saldo real (positivo=deudor, negativo=acreedor)"""
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT 
                SUM(CASE WHEN substr(cuenta_debe, 1, 2) = ? THEN monto ELSE 0 END) as total_debe,
                SUM(CASE WHEN substr(cuenta_haber, 1, 2) = ? THEN monto ELSE 0 END) as total_haber
            FROM operaciones
        ''', (codigo_cuenta, codigo_cuenta))
        
        total_debe, total_haber = cursor.fetchone()
        return (total_debe or 0) - (total_haber or 0)
        
    except sqlite3.Error as e:
        messagebox.showerror("Error", f"No se pudo obtener el saldo: {str(e)}")
        return 0
    finally:
        conn.close()

def mostrar_balance_general(mes=None):
    # Limpiar el main_frame
    for widget in main_frame.winfo_children():
        widget.destroy()

    # Si no se especifica mes, usar el actual
    if mes is None:
        mes = datetime.now().strftime("%m/%Y")
    
    # Actualizar meses disponibles
    actualizar_meses_disponibles()

    # Obtener el mes siguiente para calcular saldos iniciales
    mes_actual_num, año_actual = map(int, mes.split('/'))
    if mes_actual_num == 1:
        mes_anterior = f"12/{año_actual - 1}"
    else:
        mes_anterior = f"{mes_actual_num - 1:02d}/{año_actual}"

    # Frame principal
    main_container = ttk.Frame(main_frame)
    main_container.pack(fill="both", expand=True, padx=20, pady=20)
    main_container.grid_columnconfigure(0, weight=1)

    # Barra superior con título y selección de mes
    top_frame = ttk.Frame(main_container)
    top_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
    top_frame.grid_columnconfigure(1, weight=1)

    # Título
    ttk.Label(
        top_frame,
        text="BALANCE GENERAL",
        font=('Segoe UI', 14, 'bold'),
        foreground='white'
    ).grid(row=0, column=0, padx=(0, 10), sticky="w")

    # Contenedor para botones de meses (con scroll horizontal)
    meses_container = ttk.Frame(top_frame)
    meses_container.grid(row=0, column=1, sticky="nsew")
    
    canvas = tk.Canvas(meses_container, height=40, highlightthickness=0, bg='#333333')
    scrollbar = ttk.Scrollbar(meses_container, orient="horizontal", command=canvas.xview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all"),
            width=e.width if e.width > meses_container.winfo_width() else meses_container.winfo_width()
        )
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(xscrollcommand=scrollbar.set)
    
    canvas.grid(row=0, column=0, sticky="ew")
    scrollbar.grid(row=1, column=0, sticky="ew")
    meses_container.grid_columnconfigure(0, weight=1)

    # Crear botones de meses (ordenados cronológicamente)
    meses_ordenados = sorted(meses_disponibles, key=lambda x: datetime.strptime(x, "%m/%Y"), reverse=True)
    
    for i, mes_disponible in enumerate(meses_ordenados):
        estilo = 'primary' if mes_disponible == mes else 'secondary'
        
        btn = ttk.Button(
            scrollable_frame,
            text=datetime.strptime(mes_disponible, "%m/%Y").strftime("%B %Y").upper(),
            bootstyle=estilo,
            width=15,
            command=lambda m=mes_disponible: mostrar_balance_general(m)
        )
        btn.grid(row=0, column=i, padx=3, pady=2, sticky="nsew")

    # Canvas para scroll vertical del contenido
    content_canvas = tk.Canvas(main_container)
    content_scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=content_canvas.yview)
    content_frame = ttk.Frame(content_canvas)

    content_frame.bind(
        "<Configure>",
        lambda e: content_canvas.configure(
            scrollregion=content_canvas.bbox("all"),
            width=e.width,
            height=min(e.height, 600)  # Limitamos la altura máxima
        )
    )

    content_canvas.create_window((0, 0), window=content_frame, anchor="nw")
    content_canvas.configure(yscrollcommand=content_scrollbar.set)

    content_canvas.grid(row=1, column=0, sticky="nsew")
    content_scrollbar.grid(row=1, column=1, sticky="ns")
    main_container.grid_rowconfigure(1, weight=1)

    # Función para obtener la Utilidad Antes de Impuestos de un mes específico
    def obtener_utilidad_antes_impuestos(mes_consulta):
        conn = sqlite3.connect('data/contabilidad.db')
        cursor = conn.cursor()
        
        # Obtener saldos de cuentas de resultados (70-79)
        total_ingresos = 0
        total_gastos = 0
        total_gastos_2 = 0
        
        # Cuentas de ingresos (69-79)
        for cuenta in range(69, 80):
            cursor.execute('''
                SELECT 
                    SUM(CASE WHEN substr(cuenta_haber, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END) -
                    SUM(CASE WHEN substr(cuenta_debe, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END)
                FROM operaciones
            ''', (f"{cuenta:02d}", mes_consulta, f"{cuenta:02d}", mes_consulta))
            saldo = cursor.fetchone()[0] or 0
            total_ingresos += saldo

        # Cuentas de gastos (60-68)
        for cuenta in range(60, 69):
            cursor.execute('''
                SELECT 
                    SUM(CASE WHEN substr(cuenta_debe, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END) -
                    SUM(CASE WHEN substr(cuenta_haber, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END)
                FROM operaciones
            ''', (f"{cuenta:02d}", mes_consulta, f"{cuenta:02d}", mes_consulta))
            saldo = cursor.fetchone()[0] or 0
            total_gastos += saldo
        
        # Cuentas de gastos 2 (80-99)
        for cuenta in range(80, 100):
            cursor.execute('''
                SELECT 
                    SUM(CASE WHEN substr(cuenta_debe, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END) -
                    SUM(CASE WHEN substr(cuenta_haber, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END)
                FROM operaciones
            ''', (f"{cuenta:02d}", mes_consulta, f"{cuenta:02d}", mes_consulta))
            saldo = cursor.fetchone()[0] or 0
            total_gastos_2 += saldo

        conn.close()
        return - total_ingresos + total_gastos + total_gastos_2

    # Función para obtener saldos históricos iniciales
    def obtener_saldos_historicos_iniciales(codigo_cuenta):
        conn = sqlite3.connect('data/contabilidad.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                SUM(CASE WHEN substr(cuenta_debe, 1, 2) = ? AND glosa = 'Saldo inicial' THEN monto ELSE 0 END) -
                SUM(CASE WHEN substr(cuenta_haber, 1, 2) = ? AND glosa = 'Saldo inicial' THEN monto ELSE 0 END)
            FROM operaciones
        ''', (codigo_cuenta, codigo_cuenta))
        
        saldo = cursor.fetchone()[0] or 0
        conn.close()
        return saldo

    # Función para obtener saldos de un mes específico
    def obtener_saldos_mes(codigo_cuenta, mes_consulta, solo_iniciales=False):
        conn = sqlite3.connect('data/contabilidad.db')
        cursor = conn.cursor()
        
        if solo_iniciales:
            # Para saldos iniciales, combinamos históricos y saldos del mes anterior
            saldo_historico = obtener_saldos_historicos_iniciales(codigo_cuenta)
            
            # Obtener saldos del mes anterior (excluyendo históricos)
            cursor.execute('''
                SELECT 
                    SUM(CASE WHEN substr(cuenta_debe, 1, 2) = ? AND substr(fecha, 4, 7) = ? AND glosa != 'Saldo inicial' THEN monto ELSE 0 END) -
                    SUM(CASE WHEN substr(cuenta_haber, 1, 2) = ? AND substr(fecha, 4, 7) = ? AND glosa != 'Saldo inicial' THEN monto ELSE 0 END)
                FROM operaciones
            ''', (codigo_cuenta, mes_anterior, codigo_cuenta, mes_anterior))
            
            saldo_mes_anterior = cursor.fetchone()[0] or 0
            saldo_total = saldo_historico + saldo_mes_anterior
            
            # APLICAR LA MISMA REGLA PARA LA CUENTA 59 EN SALDOS INICIALES
            if codigo_cuenta == "59":
                utilidad = obtener_utilidad_antes_impuestos(mes_anterior)  # Utilidad del mes anterior
                saldo_total += utilidad
        else:
            # Todos los movimientos del mes (incluyendo saldos iniciales)
            cursor.execute('''
                SELECT 
                    SUM(CASE WHEN substr(cuenta_debe, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END) -
                    SUM(CASE WHEN substr(cuenta_haber, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END)
                FROM operaciones
            ''', (codigo_cuenta, mes_consulta, codigo_cuenta, mes_consulta))
            
            saldo_total = cursor.fetchone()[0] or 0
            
            # Aplicar regla especial para la cuenta 59 (Resultados acumulados)
            if codigo_cuenta == "59":
                utilidad = obtener_utilidad_antes_impuestos(mes_consulta)
                saldo_total += utilidad
        
        conn.close()
        return saldo_total
    
    # Función para crear un cuadro de balance
    def crear_cuadro_balance(frame_parent, titulo, solo_iniciales=False):
        # Frame para el cuadro completo
        frame_balance = ttk.Frame(frame_parent)
        frame_balance.pack(fill="x", pady=(0, 30))
        frame_balance.grid_columnconfigure(0, weight=1)
        frame_balance.grid_columnconfigure(1, weight=1)

        # Título del cuadro con el mes
        mes_nombre = datetime.strptime(mes, "%m/%Y").strftime("%B %Y").upper()
        ttk.Label(
            frame_balance,
            text=f"{titulo} - {mes_nombre}",
            font=('Segoe UI', 12, 'bold'),
            foreground='white'
        ).grid(row=0, column=0, columnspan=2, pady=(0, 10), sticky="w")

        # Frame para activos (columna izquierda)
        frame_activos = ttk.LabelFrame(
            frame_balance,
            text="ACTIVOS",
            bootstyle="info",
            padding=10
        )
        frame_activos.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
        frame_activos.grid_columnconfigure(0, weight=1)

        # Frame para pasivos y patrimonio (columna derecha)
        frame_pasivos_patrimonio = ttk.LabelFrame(
            frame_balance,
            text="PASIVOS Y PATRIMONIO",
            bootstyle="danger",
            padding=10
        )
        frame_pasivos_patrimonio.grid(row=1, column=1, sticky="nsew")
        frame_pasivos_patrimonio.grid_columnconfigure(0, weight=1)

        # Treeview para activos
        tree_activos = ttk.Treeview(
            frame_activos,
            columns=("cuenta", "monto"),
            show="headings",
            height=15,
            selectmode="none",
            style="Custom.Treeview"
        )
        tree_activos.heading("cuenta", text="CUENTA", anchor="w")
        tree_activos.heading("monto", text="MONTO (S/.)", anchor="e")
        tree_activos.column("cuenta", width=300, anchor="w")
        tree_activos.column("monto", width=150, anchor="e")
        tree_activos.grid(row=0, column=0, sticky="nsew")

        # Treeview para pasivos
        tree_pasivos = ttk.Treeview(
            frame_pasivos_patrimonio,
            columns=("cuenta", "monto"),
            show="headings",
            height=8,
            selectmode="none",
            style="Custom.Treeview"
        )
        tree_pasivos.heading("cuenta", text="PASIVOS", anchor="w")
        tree_pasivos.heading("monto", text="MONTO (S/.)", anchor="e")
        tree_pasivos.column("cuenta", width=300, anchor="w")
        tree_pasivos.column("monto", width=150, anchor="e")
        tree_pasivos.grid(row=0, column=0, sticky="nsew", pady=(0, 10))

        # Treeview para patrimonio
        tree_patrimonio = ttk.Treeview(
            frame_pasivos_patrimonio,
            columns=("cuenta", "monto"),
            show="headings",
            height=8,
            selectmode="none",
            style="Custom.Treeview"
        )
        tree_patrimonio.heading("cuenta", text="PATRIMONIO", anchor="w")
        tree_patrimonio.heading("monto", text="MONTO (S/.)", anchor="e")
        tree_patrimonio.column("cuenta", width=300, anchor="w")
        tree_patrimonio.column("monto", width=150, anchor="e")
        tree_patrimonio.grid(row=1, column=0, sticky="nsew")

        # Obtener y mostrar los saldos
        total_activos = 0
        total_pasivos = 0
        total_patrimonio = 0

        # Rango de cuentas para cada categoría
        cuentas_activos = list(range(10, 40))  # 10 a 39
        cuentas_pasivos = list(range(40, 50))  # 40 a 49
        cuentas_patrimonio = list(range(50, 60))  # 50 a 59

        # Procesar activos (10-39)
        for cuenta in cuentas_activos:
            saldo = obtener_saldos_mes(f"{cuenta:02d}", mes, solo_iniciales)
            # Para activos: positivo si es deudor, negativo si es acreedor
            monto_activo = saldo
            if abs(monto_activo) > 0.01:  # Solo mostrar si tiene saldo
                nombre_cuenta = obtener_nombre_cuenta(f"{cuenta:02d}")
                monto_str = f"{monto_activo:,.2f}" if monto_activo >= 0 else f"({abs(monto_activo):,.2f})"
                tree_activos.insert("", "end", values=(f"{cuenta:02d} - {nombre_cuenta}", f"S/. {monto_str}"))
                total_activos += monto_activo

        # Procesar pasivos (40-49)
        for cuenta in cuentas_pasivos:
            saldo = obtener_saldos_mes(f"{cuenta:02d}", mes, solo_iniciales)
            # Para pasivos: negativo si es deudor, positivo si es acreedor (inverso de activos)
            monto_pasivo = -saldo
            if abs(monto_pasivo) > 0.01:
                nombre_cuenta = obtener_nombre_cuenta(f"{cuenta:02d}")
                monto_str = f"{monto_pasivo:,.2f}" if monto_pasivo >= 0 else f"({abs(monto_pasivo):,.2f})"
                tree_pasivos.insert("", "end", values=(f"{cuenta:02d} - {nombre_cuenta}", f"S/. {monto_str}"))
                total_pasivos += monto_pasivo

        # Procesar patrimonio (50-59)
        for cuenta in cuentas_patrimonio:
            saldo = obtener_saldos_mes(f"{cuenta:02d}", mes, solo_iniciales)
            # Para patrimonio: negativo si es deudor, positivo si es acreedor (igual que pasivos)
            monto_patrimonio = -saldo
            if abs(monto_patrimonio) > 0.01:
                nombre_cuenta = obtener_nombre_cuenta(f"{cuenta:02d}")
                monto_str = f"{monto_patrimonio:,.2f}" if monto_patrimonio >= 0 else f"({abs(monto_patrimonio):,.2f})"
                tree_patrimonio.insert("", "end", values=(f"{cuenta:02d} - {nombre_cuenta}", f"S/. {monto_str}"))
                total_patrimonio += monto_patrimonio

        # Agregar totales
        # Total Activos
        total_activos_str = f"{total_activos:,.2f}" if total_activos >= 0 else f"({abs(total_activos):,.2f})"
        tree_activos.insert("", "end", values=("TOTAL ACTIVOS", f"S/. {total_activos_str}"), tags=("total",))
        
        # Total Pasivos
        total_pasivos_str = f"{total_pasivos:,.2f}" if total_pasivos >= 0 else f"({abs(total_pasivos):,.2f})"
        tree_pasivos.insert("", "end", values=("TOTAL PASIVOS", f"S/. {total_pasivos_str}"), tags=("total",))
        
        # Total Patrimonio
        total_patrimonio_str = f"{total_patrimonio:,.2f}" if total_patrimonio >= 0 else f"({abs(total_patrimonio):,.2f})"
        tree_patrimonio.insert("", "end", values=("TOTAL PATRIMONIO", f"S/. {total_patrimonio_str}"), tags=("total",))

        # Agregar "Pasivo + Patrimonio"
        total_pasivo_patrimonio = total_pasivos + total_patrimonio
        total_pasivo_patrimonio_str = f"{total_pasivo_patrimonio:,.2f}" if total_pasivo_patrimonio >= 0 else f"({abs(total_pasivo_patrimonio):,.2f})"
        tree_patrimonio.insert("", "end", values=("PASIVO + PATRIMONIO", f"S/. {total_pasivo_patrimonio_str}"), tags=("total",))

        # Verificar equilibrio contable
        diferencia = total_activos - (total_pasivos + total_patrimonio)
        if abs(diferencia) > 0.01:
            ttk.Label(
                frame_balance,
                text=f"¡ADVERTENCIA! Desequilibrio contable: S/. {abs(diferencia):,.2f}",
                bootstyle="danger",
                font=('Segoe UI', 10, 'bold')
            ).grid(row=2, column=0, columnspan=2, pady=(10, 0))

        # Configurar estilos
        for tree in [tree_activos, tree_pasivos, tree_patrimonio]:
            tree.tag_configure("total", foreground='white', font=('Segoe UI', 10, 'bold'), background='#3a4f6a')

    # Crear los dos cuadros de balance
    crear_cuadro_balance(content_frame, "BALANCE GENERAL INICIAL", solo_iniciales=True)
    crear_cuadro_balance(content_frame, "BALANCE GENERAL FINAL", solo_iniciales=False)

    # Botón de regreso
    btn_frame = ttk.Frame(content_frame)
    btn_frame.pack(pady=(10, 0))

    ttk.Button(
        btn_frame,
        text="Regresar a Estados Financieros",
        bootstyle="secondary",
        command=mostrar_estados_financieros
    ).pack()

    # Ajustar scrolls
    canvas.xview_moveto(1.0)  # Scroll horizontal al final
    content_canvas.yview_moveto(0.0)  # Scroll vertical al inicio

# --- Constantes para estilos ---
COLOR_CABECERA = '4FC3F7'
COLOR_TOTALES = '2A3F54'
ESTILO_BORDE = Side(border_style="thin", color="000000")

# --- ESTILOS GLOBALES (agregar al inicio del módulo de reportes) ---
COLOR_FONDO = "#1a1a2e"          # Fondo oscuro premium
COLOR_TEXTO = "#ffffff"          # Texto blanco
COLOR_PRIMARIO = "#4fc3f7"       # Azul claro moderno
COLOR_SECUNDARIO = "#2a3f54"     # Azul oscuro elegante
COLOR_ACENTO = "#00c292"         # Verde azulado para acentos
COLOR_PANEL = "#16213e"          # Color de paneles

FUENTE_TITULO = ('Segoe UI', 14, 'bold')
FUENTE_SUBTITULO = ('Segoe UI', 11, 'bold')
FUENTE_NORMAL = ('Segoe UI', 10)
FUENTE_DATOS = ('Segoe UI', 9)

ESTILO_CABECERA = {
    'font': FUENTE_SUBTITULO,
    'bg': COLOR_SECUNDARIO,
    'fg': COLOR_TEXTO,
    'relief': 'solid',
    'borderwidth': 1
}

ESTILO_FILA_PAR = {
    'font': FUENTE_DATOS,
    'bg': COLOR_PANEL,
    'fg': COLOR_TEXTO
}

ESTILO_FILA_IMPAR = {
    'font': FUENTE_DATOS,
    'bg': COLOR_FONDO,
    'fg': COLOR_TEXTO
}

ESTILO_TOTALES = {
    'font': FUENTE_SUBTITULO,
    'bg': COLOR_ACENTO,
    'fg': '#000000',
    'relief': 'solid',
    'borderwidth': 1
}

def obtener_saldos_mes(codigo_cuenta, mes):
    """Obtiene el saldo de una cuenta para un mes específico"""
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT 
                SUM(CASE WHEN substr(cuenta_debe, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END) -
                SUM(CASE WHEN substr(cuenta_haber, 1, 2) = ? AND substr(fecha, 4, 7) = ? THEN monto ELSE 0 END)
            FROM operaciones
        ''', (codigo_cuenta, mes, codigo_cuenta, mes))
        
        saldo = cursor.fetchone()[0] or 0
        return saldo
    finally:
        conn.close()

# --- Funciones de Exportación ---
def exportar_a_excel(datos, nombre_archivo, titulo, columnas, estilo_personalizado=True):
    """Exporta datos a Excel manteniendo el estilo de la aplicación"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Configuración general
        ws.sheet_view.showGridLines = False
        
        # Función para convertir colores HEX a ARGB (sin #)
        def hex_to_argb(hex_color):
            if hex_color.startswith('#'):
                return "FF" + hex_color[1:]
            return "FF" + hex_color
        
        # Estilos
        if estilo_personalizado:
            # Convertir colores
            argb_fondo = hex_to_argb(COLOR_FONDO)
            argb_panel = hex_to_argb(COLOR_PANEL)
            argb_primario = hex_to_argb(COLOR_PRIMARIO)
            argb_secundario = hex_to_argb(COLOR_SECUNDARIO)
            argb_acento = hex_to_argb(COLOR_ACENTO)
            
            # Estilo título
            titulo_font = Font(name='Segoe UI', size=14, bold=True, color=hex_to_argb(COLOR_TEXTO))
            titulo_fill = PatternFill(start_color=argb_secundario, end_color=argb_secundario, fill_type="solid")
            
            # Estilo cabecera
            cabecera_font = Font(name='Segoe UI', size=10, bold=True, color=hex_to_argb(COLOR_TEXTO))
            cabecera_fill = PatternFill(start_color=argb_primario, end_color=argb_primario, fill_type="solid")
            
            # Estilo totales
            totales_fill = PatternFill(start_color=argb_acento, end_color=argb_acento, fill_type="solid")
            
            # Bordes
            borde_fino = Side(border_style="thin", color="FFFFFF")
            borde = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
        else:
            # Estilos por defecto si no queremos personalización
            titulo_font = Font(bold=True)
            titulo_fill = PatternFill()
            cabecera_font = Font(bold=True)
            cabecera_fill = PatternFill()
            totales_fill = PatternFill()
            borde = Border()
        
        # Título
        ws.merge_cells('A1:{}1'.format(chr(65 + len(columnas) - 1)))
        celda_titulo = ws['A1']
        celda_titulo.value = titulo
        celda_titulo.font = titulo_font
        celda_titulo.fill = titulo_fill
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        
        # Cabeceras
        for col_num, (col_name, col_width) in enumerate(columnas, 1):
            col_letra = chr(64 + col_num)
            ws.column_dimensions[col_letra].width = col_width
            celda = ws.cell(row=2, column=col_num, value=col_name)
            celda.font = cabecera_font
            celda.fill = cabecera_fill
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = borde
        
        # Datos
        for row_num, row_data in enumerate(datos, 3):
            for col_num, cell_value in enumerate(row_data, 1):
                celda = ws.cell(row=row_num, column=col_num, value=cell_value)
                celda.font = Font(name='Segoe UI', size=9)
                celda.border = borde
                
                # Alternar colores de fila
                if row_num % 2 == 0:
                    celda.fill = PatternFill(start_color=COLOR_PANEL[1:], end_color=COLOR_PANEL[1:], fill_type="solid")
                else:
                    celda.fill = PatternFill(start_color=COLOR_FONDO[1:], end_color=COLOR_FONDO[1:], fill_type="solid")
                
                # Formato números
                if isinstance(cell_value, (float, int)):
                    celda.number_format = '#,##0.00'
                    celda.alignment = Alignment(horizontal="right")
                else:
                    celda.alignment = Alignment(horizontal="left")
        
        # Totales
        if datos:
            total_row = len(datos) + 3
            for col_num in range(1, len(columnas) + 1):
                celda = ws.cell(row=total_row, column=col_num)
                celda.fill = totales_fill
                celda.border = borde
                
                if col_num == 1:
                    celda.value = "TOTAL"
                    celda.font = Font(name='Segoe UI', bold=True)
                elif columnas[col_num-1][0].lower() in ['monto', 'debe', 'haber', 'saldo']:
                    col_letra = chr(64 + col_num)
                    formula = f"=SUM({col_letra}3:{col_letra}{total_row-1})"
                    celda.value = formula
                    celda.number_format = '#,##0.00'
        
        # Guardar
        if not nombre_archivo.endswith('.xlsx'):
            nombre_archivo += '.xlsx'
        
        ruta = os.path.join('reportes', nombre_archivo)
        os.makedirs('reportes', exist_ok=True)
        wb.save(ruta)
        return ruta
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar a Excel: {str(e)}")
        return None

def exportar_a_pdf(datos, nombre_archivo, titulo, columnas):
    """Exporta a PDF manteniendo el estilo de la aplicación"""
    try:
        if not nombre_archivo.endswith('.pdf'):
            nombre_archivo += '.pdf'
        
        ruta = os.path.join('reportes', nombre_archivo)
        os.makedirs('reportes', exist_ok=True)
        
        # Configuración del documento
        doc = SimpleDocTemplate(
            ruta,
            pagesize=letter,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        
        elementos = []
        estilos = getSampleStyleSheet()
        
        # Estilo título (usamos Helvetica que es estándar en PDF)
        titulo_style = ParagraphStyle(
            'Titulo',
            parent=estilos['Title'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=colors.HexColor(COLOR_TEXTO),
            backColor=colors.HexColor(COLOR_SECUNDARIO),
            alignment=1,  # Centrado
            spaceAfter=20
        )
        
        # Fondo del título
        elementos.append(Table(
            [[Paragraph(titulo, titulo_style)]],
            style=[
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor(COLOR_SECUNDARIO)),
                ('TEXTCOLOR', (0,0), (-1,-1), colors.HexColor(COLOR_TEXTO)),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 14),
                ('BOX', (0,0), (-1,-1), 1, colors.HexColor(COLOR_PRIMARIO)),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ],
            colWidths=[doc.width]
        ))
        
        # Preparar datos para la tabla
        datos_tabla = [[Paragraph(col[0], estilos['Normal']) for col in columnas]]  # Cabeceras
        
        for fila in datos:
            datos_tabla.append([
                Paragraph(str(item), estilos['Normal']) if isinstance(item, str) else item
                for item in fila
            ])
        
        # Crear tabla con estilo
        tabla = Table(datos_tabla, repeatRows=1)
        
        # Estilo de la tabla (usamos Helvetica)
        estilo_tabla = TableStyle([
            # Cabecera
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor(COLOR_PRIMARIO)),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor(COLOR_TEXTO)),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            
            # Filas alternas
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor(COLOR_FONDO)),
            ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor(COLOR_TEXTO)),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            
            # Bordes
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor(COLOR_PRIMARIO)),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ])
        
        # Aplicar fondo alterno a filas
        for i, fila in enumerate(datos_tabla[1:], 1):
            if i % 2 == 0:
                estilo_tabla.add('BACKGROUND', (0,i), (-1,i), colors.HexColor(COLOR_PANEL))
        
        # Resaltar totales
        if any("TOTAL" in str(fila[0]) for fila in datos):
            estilo_tabla.add('BACKGROUND', (0,-1), (-1,-1), colors.HexColor(COLOR_ACENTO))
            estilo_tabla.add('TEXTCOLOR', (0,-1), (-1,-1), colors.black)
            estilo_tabla.add('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold')
        
        tabla.setStyle(estilo_tabla)
        elementos.append(tabla)
        
        # Pie de página (simplificado)
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        elementos.append(Paragraph(
            f"Generado el {fecha}",
            ParagraphStyle(
                'Pie',
                parent=estilos['Normal'],
                fontName='Helvetica-Oblique',
                fontSize=8,
                textColor=colors.HexColor(COLOR_PRIMARIO),
                alignment=2  # Derecha
            )
        ))
        
        doc.build(elementos)
        return ruta
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar a PDF: {str(e)}")
        return None

def exportar_balance_general(formato):
    """Exporta el balance general con el estilo de la aplicación"""
    try:
        mes_actual = datetime.now().strftime("%m/%Y")
        datos = []
        total_activo = total_pasivo = total_patrimonio = 0
        
        # Procesar activos (10-39)
        for cuenta in range(10, 40):
            saldo = obtener_saldos_mes(f"{cuenta:02d}", mes_actual)
            if abs(saldo) > 0.01:
                nombre = obtener_nombre_cuenta(f"{cuenta:02d}")
                datos.append([
                    f"{cuenta:02d}", 
                    nombre, 
                    saldo if saldo > 0 else 0, 
                    abs(saldo) if saldo < 0 else 0
                ])
                if saldo > 0:
                    total_activo += saldo
        
        # Separador
        datos.append(["", "TOTAL ACTIVO", total_activo, ""])
        
        # Procesar pasivos (40-49)
        for cuenta in range(40, 50):
            saldo = obtener_saldos_mes(f"{cuenta:02d}", mes_actual)
            if abs(saldo) > 0.01:
                nombre = obtener_nombre_cuenta(f"{cuenta:02d}")
                datos.append([
                    f"{cuenta:02d}", 
                    nombre, 
                    "", 
                    abs(saldo) if saldo < 0 else saldo
                ])
                if saldo < 0:
                    total_pasivo += abs(saldo)
        
        # Separador
        datos.append(["", "TOTAL PASIVO", "", total_pasivo])
        
        # Procesar patrimonio (50-59)
        for cuenta in range(50, 60):
            saldo = obtener_saldos_mes(f"{cuenta:02d}", mes_actual)
            if abs(saldo) > 0.01:
                nombre = obtener_nombre_cuenta(f"{cuenta:02d}")
                datos.append([
                    f"{cuenta:02d}", 
                    nombre, 
                    "", 
                    abs(saldo) if saldo < 0 else saldo
                ])
                if saldo < 0:
                    total_patrimonio += abs(saldo)
        
        # Totales finales
        datos.append(["", "TOTAL PATRIMONIO", "", total_patrimonio])
        datos.append(["", "TOTAL PASIVO + PATRIMONIO", "", total_pasivo + total_patrimonio])
        
        columnas = [
            ("Código", 10),
            ("Descripción", 50),
            ("Activo", 15),
            ("Pasivo/Patrimonio", 20)
        ]
        
        fecha = datetime.now().strftime("%Y%m%d_%H%M")
        nombre_archivo = f"Balance_General_{fecha}"
        titulo = "BALANCE GENERAL - SISTEMA CONTABLE"
        
        if formato == 'excel':
            ruta = exportar_a_excel(datos, nombre_archivo, titulo, columnas)
        else:
            ruta = exportar_a_pdf(datos, nombre_archivo, titulo, columnas)
        
        if ruta:
            messagebox.showinfo("Éxito", f"Balance General exportado correctamente:\n{ruta}")
            os.startfile(ruta)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar Balance General: {str(e)}")

def exportar_flujo_efectivo(formato):
    """Exporta el flujo de efectivo al formato especificado"""
    try:
        # Obtener datos del flujo de efectivo
        mes_actual = datetime.now().strftime("%m/%Y")
        datos = []
        
        # Flujo de actividades operativas (10-19)
        for cuenta in range(10, 20):
            saldo = obtener_saldos_mes(f"{cuenta:02d}", mes_actual)
            if abs(saldo) > 0.01:
                nombre = obtener_nombre_cuenta(f"{cuenta:02d}")
                datos.append(["Operativas", f"{cuenta:02d} - {nombre}", saldo])
        
        # Flujo de actividades de inversión (20-29)
        for cuenta in range(20, 30):
            saldo = obtener_saldos_mes(f"{cuenta:02d}", mes_actual)
            if abs(saldo) > 0.01:
                nombre = obtener_nombre_cuenta(f"{cuenta:02d}")
                datos.append(["Inversión", f"{cuenta:02d} - {nombre}", saldo])
        
        # Flujo de actividades de financiación (30-39)
        for cuenta in range(30, 40):
            saldo = obtener_saldos_mes(f"{cuenta:02d}", mes_actual)
            if abs(saldo) > 0.01:
                nombre = obtener_nombre_cuenta(f"{cuenta:02d}")
                datos.append(["Financiación", f"{cuenta:02d} - {nombre}", saldo])
        
        # Agregar totales
        total_operativas = sum(fila[2] for fila in datos if fila[0] == "Operativas")
        total_inversion = sum(fila[2] for fila in datos if fila[0] == "Inversión")
        total_financiacion = sum(fila[2] for fila in datos if fila[0] == "Financiación")
        
        datos.append(["TOTAL", "Actividades Operativas", total_operativas])
        datos.append(["TOTAL", "Actividades de Inversión", total_inversion])
        datos.append(["TOTAL", "Actividades de Financiación", total_financiacion])
        datos.append(["TOTAL GENERAL", "Variación de Efectivo", 
                     total_operativas + total_inversion + total_financiacion])
        
        columnas = [
            ("Tipo Actividad", 20),
            ("Cuenta", 50),
            ("Monto", 20)
        ]
        
        fecha = datetime.now().strftime("%Y%m%d")
        nombre_archivo = f"Flujo_Efectivo_{fecha}"
        titulo = "ESTADO DE FLUJO DE EFECTIVO"
        
        if formato == 'excel':
            ruta = exportar_a_excel(datos, nombre_archivo, titulo, columnas)
        else:
            ruta = exportar_a_pdf(datos, nombre_archivo, titulo, columnas)
        
        if ruta:
            messagebox.showinfo("Éxito", f"Flujo de Efectivo exportado correctamente:\n{ruta}")
            os.startfile(ruta)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar Flujo de Efectivo: {str(e)}")

# --- Funciones específicas para cada reporte ---
def exportar_libro_diario(formato, mes):
    try:
        # Conexión a la base de datos con filtro por mes
        conn = sqlite3.connect('data/contabilidad.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                operacion_id,
                substr(fecha, 1, 10) as fecha_str,
                CASE 
                    WHEN cuenta_debe != '' THEN cuenta_debe
                    ELSE cuenta_haber
                END as cuenta,
                CASE 
                    WHEN cuenta_debe != '' THEN monto
                    ELSE 0
                END as debe,
                CASE 
                    WHEN cuenta_haber != '' THEN monto
                    ELSE 0
                END as haber,
                CASE WHEN glosa IS NULL THEN '' ELSE glosa END as glosa
            FROM operaciones
            WHERE (glosa != 'Saldo inicial' OR glosa IS NULL)
            AND substr(fecha, 4, 7) = ?
            ORDER BY fecha DESC
        ''', (mes,))
        
        datos = cursor.fetchall()
        conn.close()
        
        # Formatear los datos para exportación
        datos_formateados = []
        for fila in datos:
            # Asegurar que la fecha esté en formato dd/mm/YYYY
            try:
                fecha_obj = datetime.strptime(fila[1], "%d/%m/%Y")
                fecha_str = fecha_obj.strftime("%d/%m/%Y")
            except ValueError:
                fecha_str = fila[1]  # Si falla el parseo, mantener el original
            
            datos_formateados.append((
                fila[0],        # ID
                fecha_str,      # Fecha
                fila[2],        # Cuenta
                fila[3],        # Debe
                fila[4],        # Haber
                fila[5]         # Glosa
            ))
        
        # Calcular totales
        total_debe = sum(fila[3] for fila in datos_formateados)
        total_haber = sum(fila[4] for fila in datos_formateados)
        
        if formato == 'excel':
            exportar_libro_diario_excel(datos_formateados, mes_actual, total_debe, total_haber)
        else:
            exportar_libro_diario_pdf(datos_formateados, mes_actual, total_debe, total_haber)
            
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar el Libro Diario: {str(e)}") 

def exportar_libro_diario_excel(datos, mes_actual, total_debe, total_haber):
    """Exporta a Excel con el formato específico"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro Diario"
        
        # Configuración de columnas
        columnas = [
            ("ID", 6.29),
            ("Fecha", 16.0),
            ("Cuenta", 45.71),
            ("Debe", 17.86),
            ("Haber", 17.86),
            ("Glosa", 45.71)
        ]
        
        # Estilos
        titulo_style = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color='404040', end_color='404040', fill_type="solid")
        encabezado_style = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
        encabezado_fill = PatternFill(start_color='808080', end_color='808080', fill_type="solid")
        dato_style = Font(name='Segoe UI', size=10, color='000000')
        dato_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
        total_style = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        total_fill = PatternFill(start_color='404040', end_color='404040', fill_type="solid")
        borde = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))
        
        # Título
        ws.merge_cells('A1:F1')
        celda_titulo = ws['A1']
        celda_titulo.value = f"LIBRO DIARIO - {mes_actual}"
        celda_titulo.font = titulo_style
        celda_titulo.fill = titulo_fill
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24
        
        # Encabezados
        for col_num, (col_name, width) in enumerate(columnas, 1):
            col_letra = get_column_letter(col_num)
            ws.column_dimensions[col_letra].width = width
            
            celda = ws.cell(row=2, column=col_num, value=col_name)
            celda.font = encabezado_style
            celda.fill = encabezado_fill
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = borde
        
        ws.row_dimensions[2].height = 19.5
        
        # Datos
        for row_num, row_data in enumerate(datos, 3):
            for col_num, cell_value in enumerate(row_data, 1):
                celda = ws.cell(row=row_num, column=col_num)
                
                # Formatear valores especiales
                if col_num == 4 or col_num == 5:  # Columnas Debe y Haber
                    if cell_value == 0 or cell_value == 0.0:
                        celda.value = "-"
                    else:
                        celda.value = cell_value
                else:
                    celda.value = cell_value
                
                celda.font = dato_style
                celda.fill = dato_fill
                celda.border = borde
                
                # Alineación
                if col_num in [3, 6]:  # Cuenta y Glosa alineados a la izquierda
                    celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Ajustar altura de fila automáticamente
            ws.row_dimensions[row_num].height = None
        
        # Totales
        total_row = len(datos) + 3
        ws.merge_cells(f'A{total_row}:C{total_row}')
        celda_total = ws.cell(row=total_row, column=1, value="TOTAL")
        celda_total.font = total_style
        celda_total.fill = total_fill
        celda_total.alignment = Alignment(horizontal="center", vertical="center")
        celda_total.border = borde
        
        # Formatear totales (mostrar "-" si es cero)
        total_debe_cell = ws.cell(row=total_row, column=4, value=total_debe if total_debe != 0 else "-")
        total_debe_cell.font = total_style
        total_debe_cell.fill = total_fill
        total_debe_cell.border = borde
        
        total_haber_cell = ws.cell(row=total_row, column=5, value=total_haber if total_haber != 0 else "-")
        total_haber_cell.font = total_style
        total_haber_cell.fill = total_fill
        total_haber_cell.border = borde
        
        ws.cell(row=total_row, column=6, value="").fill = total_fill
        ws.cell(row=total_row, column=6).border = borde
        
        # Aplicar formato de número a montos (solo si no son cero)
        for row in ws.iter_rows(min_row=3, max_row=total_row, min_col=4, max_col=5):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value != 0:
                    cell.number_format = '#,##0.00'
        
        # Guardar archivo
        os.makedirs('reportes', exist_ok=True)
        fecha = datetime.now().strftime("%Y%m%d_%H%M")
        ruta = os.path.join('reportes', f"Libro_Diario_{fecha}.xlsx")
        wb.save(ruta)
        
        messagebox.showinfo("Éxito", f"Libro Diario exportado correctamente:\n{ruta}")
        os.startfile(ruta)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar a Excel: {str(e)}")

def exportar_libro_diario_pdf(datos, mes_actual, total_debe, total_haber):
    """Exporta a PDF en formato A4 con ajuste proporcional"""
    try:
        # Configuración del documento
        fecha = datetime.now().strftime("%Y%m%d_%H%M")
        ruta = os.path.join('reportes', f"Libro_Diario_{fecha}.pdf")
        os.makedirs('reportes', exist_ok=True)
        
        # Tamaño A4 en puntos (1 pulgada = 72 puntos)
        ancho_A4 = 595  # 210mm
        alto_A4 = 842   # 297mm
        
        doc = SimpleDocTemplate(
            ruta,
            pagesize=(ancho_A4, alto_A4),
            leftMargin=40,
            rightMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        
        # Estilos
        estilos = getSampleStyleSheet()
        
        # Estilo para el título
        titulo_style = ParagraphStyle(
            'Titulo',
            parent=estilos['Title'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=colors.white,
            alignment=TA_CENTER,
            backColor=colors.HexColor('#404040'),
            spaceAfter=20
        )
        
        # Estilo para encabezados de columna
        encabezado_style = ParagraphStyle(
            'Encabezado',
            parent=estilos['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.white,
            alignment=TA_CENTER,
            backColor=colors.HexColor('#808080'),
            leading=14
        )
        
        # Estilo para datos normales
        dato_style = ParagraphStyle(
            'Dato',
            parent=estilos['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.black,
            alignment=TA_CENTER,
            leading=12
        )
        
        # Estilo para datos alineados a la izquierda
        dato_left_style = ParagraphStyle(
            'DatoLeft',
            parent=dato_style,
            alignment=TA_LEFT
        )
        
        # Estilo para totales
        total_style = ParagraphStyle(
            'Total',
            parent=estilos['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.white,
            alignment=TA_CENTER,
            backColor=colors.HexColor('#808080')
        )
        
        # Contenido
        elementos = []
        
        # Título del documento
        titulo = Paragraph(f"LIBRO DIARIO - {mes_actual}", titulo_style)
        elementos.append(titulo)
        
        # Calcular anchos de columna proporcionales
        # Distribución porcentual aproximada: ID(8%), Fecha(15%), Cuenta(32%), Debe(15%), Haber(15%), Glosa(15%)
        ancho_disponible = ancho_A4 - 80  # Restamos márgenes
        anchos_columnas = [
            ancho_disponible * 0.08,  # ID
            ancho_disponible * 0.15,  # Fecha
            ancho_disponible * 0.32,  # Cuenta
            ancho_disponible * 0.15,  # Debe
            ancho_disponible * 0.15,  # Haber
            ancho_disponible * 0.15   # Glosa
        ]
        
        # Datos para la tabla
        datos_tabla = []
        
        # Encabezados
        encabezados = [
            Paragraph("ID", encabezado_style),
            Paragraph("Fecha", encabezado_style),
            Paragraph("Cuenta", encabezado_style),
            Paragraph("Debe", encabezado_style),
            Paragraph("Haber", encabezado_style),
            Paragraph("Glosa", encabezado_style)
        ]
        datos_tabla.append(encabezados)
        
        # Filas de datos
        for fila in datos:
            # Formatear valores
            id_str = str(fila[0]) if fila[0] is not None else ""
            fecha_str = str(fila[1]) if fila[1] is not None else ""
            cuenta_str = str(fila[2]) if fila[2] is not None else ""
            debe_str = "-" if fila[3] == 0 else f"{fila[3]:,.2f}"
            haber_str = "-" if fila[4] == 0 else f"{fila[4]:,.2f}"
            glosa_str = str(fila[5]) if fila[5] is not None else ""
            
            datos_tabla.append([
                Paragraph(id_str, dato_style),
                Paragraph(fecha_str, dato_style),
                Paragraph(cuenta_str, dato_left_style),
                Paragraph(debe_str, dato_style),
                Paragraph(haber_str, dato_style),
                Paragraph(glosa_str, dato_left_style)
            ])
        
        # Fila de totales
        total_debe_str = "-" if total_debe == 0 else f"{total_debe:,.2f}"
        total_haber_str = "-" if total_haber == 0 else f"{total_haber:,.2f}"
        
        datos_tabla.append([
            Paragraph("TOTAL", total_style),
            Paragraph("", total_style),
            Paragraph("", total_style),
            Paragraph(total_debe_str, total_style),
            Paragraph(total_haber_str, total_style),
            Paragraph("", total_style)
        ])
        
        # Crear tabla
        tabla = Table(datos_tabla, colWidths=anchos_columnas, repeatRows=1)
        
        # Aplicar estilos a la tabla
        estilo_tabla = TableStyle([
            # Estilo para encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#808080')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#808080')),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            
            # Estilo para datos
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('ALIGN', (3, 1), (4, -2), 'RIGHT'),  # Alinear números a la derecha
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#404040')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#404040')),
            
            # Estilo para fila de totales
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#808080')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#808080')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            
            # Combinar celdas para el total
            ('SPAN', (0, -1), (2, -1))
        ])
        
        # Alternar colores de fila para mejor legibilidad
        for i in range(1, len(datos_tabla)-1):
            if i % 2 == 0:
                estilo_tabla.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f5f5f5'))
            else:
                estilo_tabla.add('BACKGROUND', (0, i), (-1, i), colors.white)
        
        tabla.setStyle(estilo_tabla)
        elementos.append(tabla)
        
        # Pie de página
        fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M")
        pie_pagina = Paragraph(
            f"Generado el {fecha_generacion} - Sistema Contable Premium",
            ParagraphStyle(
                'Pie',
                parent=estilos['Normal'],
                fontName='Helvetica-Oblique',
                fontSize=8,
                textColor=colors.HexColor('#808080'),
                alignment=TA_RIGHT
            )
        )
        elementos.append(pie_pagina)
        
        # Construir el PDF
        doc.build(elementos)
        
        messagebox.showinfo("Éxito", f"Libro Diario exportado correctamente a PDF:\n{ruta}")
        os.startfile(ruta)
        
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar a PDF: {str(e)}")

def exportar_libro_mayor(formato):
    """Exporta el libro mayor al formato especificado"""
    conn = sqlite3.connect('data/contabilidad.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT 
            substr(cuenta_debe, 1, 2) as "Cuenta",
            cuenta_debe as "Nombre Cuenta",
            SUM(monto) as "Total Debe",
            0 as "Total Haber"
        FROM operaciones
        WHERE cuenta_debe != ''
        GROUP BY substr(cuenta_debe, 1, 2)
        
        UNION ALL
        
        SELECT 
            substr(cuenta_haber, 1, 2) as "Cuenta",
            cuenta_haber as "Nombre Cuenta",
            0 as "Total Debe",
            SUM(monto) as "Total Haber"
        FROM operaciones
        WHERE cuenta_haber != ''
        GROUP BY substr(cuenta_haber, 1, 2)
        
        ORDER BY "Cuenta"
    ''')
    
    datos = cursor.fetchall()
    conn.close()
    
    columnas = [
        ("Cuenta", 10),
        ("Nombre Cuenta", 40),
        ("Total Debe", 15),
        ("Total Haber", 15)
    ]
    
    fecha = datetime.now().strftime("%Y%m%d")
    nombre_archivo = f"Libro_Mayor_{fecha}"
    titulo = "LIBRO MAYOR CONTABLE"
    
    if formato == 'excel':
        ruta = exportar_a_excel(datos, nombre_archivo, titulo, columnas)
    else:
        ruta = exportar_a_pdf(datos, nombre_archivo, titulo, columnas)
    
    if ruta:
        messagebox.showinfo("Éxito", f"Reporte exportado correctamente:\n{ruta}")
        os.startfile(ruta)

# --- Modificar la función mostrar_reportes ---
def mostrar_reportes():
    # Limpiar el main_frame
    for widget in main_frame.winfo_children():
        widget.destroy()

    # Obtener meses disponibles
    actualizar_meses_disponibles()
    mes_actual = datetime.now().strftime("%m/%Y")

    # Frame principal
    frame_principal = ttk.Frame(main_frame)
    frame_principal.pack(fill="both", expand=True, padx=20, pady=20)
    frame_principal.grid_columnconfigure(0, weight=1)

    # Título
    ttk.Label(
        frame_principal,
        text="EXPORTAR REPORTES",
        style='Titulo.TLabel'
    ).grid(row=0, column=0, pady=(0, 20), sticky="w")

    # --- Selector de Meses (Scroll Horizontal) ---
    meses_frame = ttk.Frame(frame_principal)
    meses_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 20))
    meses_frame.grid_columnconfigure(0, weight=1)

    # Canvas para scroll horizontal
    canvas = tk.Canvas(meses_frame, height=40, highlightthickness=0, bg='#333333')
    scrollbar = ttk.Scrollbar(meses_frame, orient="horizontal", command=canvas.xview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all"),
            width=e.width if e.width > meses_frame.winfo_width() else meses_frame.winfo_width()
        )
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(xscrollcommand=scrollbar.set)

    canvas.grid(row=0, column=0, sticky="ew")
    scrollbar.grid(row=1, column=0, sticky="ew")
    meses_frame.grid_columnconfigure(0, weight=1)

    # Botones de meses (ordenados del más reciente al más antiguo)
    meses_ordenados = sorted(meses_disponibles, key=lambda x: datetime.strptime(x, "%m/%Y"), reverse=True)
    mes_seleccionado = tk.StringVar(value=mes_actual)  # Mes actual por defecto

    for i, mes_disponible in enumerate(meses_ordenados):
        btn = ttk.Radiobutton(
            scrollable_frame,
            text=datetime.strptime(mes_disponible, "%m/%Y").strftime("%B %Y").upper(),
            variable=mes_seleccionado,
            value=mes_disponible,
            bootstyle="toolbutton",
            width=15
        )
        btn.grid(row=0, column=i, padx=3, pady=2, sticky="nsew")

    # --- Botones de Exportación (Excel/PDF) ---
    export_frame = ttk.Frame(frame_principal)
    export_frame.grid(row=2, column=0, sticky="nsew", pady=(10, 0))
    export_frame.grid_columnconfigure(0, weight=1)

    # Función para crear botones de exportación con el mes seleccionado
    def crear_boton_exportacion(parent, texto, comando_exportar):
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill="x", pady=5)
        
        ttk.Label(btn_frame, text=texto, width=20).pack(side="left", padx=5)
        
        ttk.Button(
            btn_frame,
            text="Excel",
            bootstyle=(SUCCESS, OUTLINE),
            command=lambda: comando_exportar('excel', mes_seleccionado.get()),
            width=10
        ).pack(side="left", padx=2)
        
        ttk.Button(
            btn_frame,
            text="PDF",
            bootstyle=(DANGER, OUTLINE),
            command=lambda: comando_exportar('pdf', mes_seleccionado.get()),
            width=10
        ).pack(side="left", padx=2)

    # Botones para cada reporte
    crear_boton_exportacion(export_frame, "Libro Diario:", exportar_libro_diario)
    crear_boton_exportacion(export_frame, "Libro Mayor:", exportar_libro_mayor)
    crear_boton_exportacion(export_frame, "Balance General:", exportar_balance_general)
    crear_boton_exportacion(export_frame, "Flujo de Efectivo:", exportar_flujo_efectivo)

    # Botón de regreso
    ttk.Button(
        frame_principal,
        text="Regresar",
        bootstyle=(OUTLINE, SECONDARY),
        command=lambda: mostrar_seccion("inicio")
    ).grid(row=3, column=0, pady=20)

# --- Barra de estado premium con grid ---
status_bar = ttk.Frame(root, bootstyle="dark")
status_bar.grid(row=2, column=0, columnspan=2, sticky="nsew")
status_bar.grid_columnconfigure(0, weight=1)

ttk.Label(
    status_bar,
    text="Sesión: ADMIN | Última actualización: Hoy 15:30",
    bootstyle="inverse-dark"
).grid(row=0, column=0, sticky="w", padx=10)

ttk.Label(
    status_bar,
    text="© 2023 Sistema Contable Premium v2.0",
    bootstyle="inverse-dark"
).grid(row=0, column=1, sticky="e", padx=10)

# Crear el formulario pero ocultarlo inicialmente
formulario_asientos = crear_formulario_asientos()
formulario_asientos.grid_forget()  # Ocultar inicialmente
panel_inicio = None

# Mostrar la pantalla de inicio
mostrar_seccion("inicio")

# --- Barra de estado ---
status_bar = ttk.Frame(root, bootstyle="dark")
status_bar.grid(row=2, column=0, columnspan=2, sticky="nsew")

# Inicializar base de datos y mostrar sección de inicio
inicializar_base_datos()
operaciones_registradas = cargar_operaciones_db()

root.mainloop()